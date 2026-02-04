#!/usr/bin/env python3
"""
CGAS Module 1: Chloroplast Genome Assembly and Quality Control Pipeline
=========================================================================

This module integrates:
1. Read quality control (Fastp)
2. Chloroplast genome assembly (GetOrganelle)
3. Assembly validation and completeness assessment
4. Read mapping and coverage depth analysis
5. Comprehensive reporting

Author: Abdullah
"""

import os
import sys
import re
import subprocess
import json
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, asdict
from collections import defaultdict
import argparse
from datetime import datetime


@dataclass
class ReadPair:
    """Store information about a read pair or single-end read"""
    sample_name: str
    forward_read: Path
    reverse_read: Optional[Path] = None
    is_paired: bool = True
    
    def __post_init__(self):
        self.is_paired = self.reverse_read is not None


@dataclass
class FastpStats:
    """Store Fastp QC statistics"""
    sample_name: str
    total_reads_before: int
    total_reads_after: int
    total_bases_before: int
    total_bases_after: int
    q20_rate_before: float
    q20_rate_after: float
    q30_rate_before: float
    q30_rate_after: float
    gc_content: float
    duplication_rate: float
    
    def to_dict(self):
        return asdict(self)


@dataclass
class AssemblyStats:
    """Store GetOrganelle assembly statistics"""
    sample_name: str
    num_contigs: int
    is_complete: bool
    is_circular: bool
    total_length: int
    lsc_length: Optional[int] = None
    ssc_length: Optional[int] = None
    ir_length: Optional[int] = None
    has_flipflop: bool = False
    assembly_path: Optional[Path] = None
    
    def to_dict(self):
        data = asdict(self)
        if self.assembly_path:
            data['assembly_path'] = str(self.assembly_path)
        return data


@dataclass
class MappingStats:
    """Store read mapping and coverage statistics"""
    sample_name: str
    total_reads: int
    mapped_reads: int
    mapping_rate: float
    mean_coverage: float
    median_coverage: float
    cp_reads_saved: int
    
    def to_dict(self):
        return asdict(self)


class CGASModule1:
    """Main class for CGAS Module 1 pipeline"""
    
    # Supported read file patterns
    FORWARD_PATTERNS = [
        r'(.+?)_1\.fastq$',
        r'(.+?)_1\.fastq\.gz$',
        r'(.+?)_1\.fq$',
        r'(.+?)_1\.fq\.gz$',
        r'(.+?)_1\.clean\.fq\.gz$',
        r'(.+?)_R1\.fastq$',
        r'(.+?)_R1\.fastq\.gz$',
        r'(.+?)_R1\.fq$',
        r'(.+?)_R1\.fq\.gz$',
    ]
    
    def __init__(self, 
                 input_dir: str,
                 output_dir: str,
                 threads: int = 8,
                 fastp_path: str = "fastp",
                 getorganelle_path: str = "get_organelle_from_reads.py",
                 bwa_path: str = "bwa",
                 samtools_path: str = "samtools",
                 genome_type: str = "embplant_pt",
                 k_values: str = "21,45,65,85,105",
                 rounds: int = 15,
                 use_nohup: bool = False,
                 log_level: str = "INFO",
                 force_mapping: bool = False,
                 trim_poly_g: bool = False,
                 cut_mean_quality: int = 20,
                 skip_existing: bool = True):
        """
        Initialize CGAS Module 1
        
        Args:
            input_dir: Directory containing raw FASTQ files
            output_dir: Directory for all outputs
            threads: Number of threads for parallel processing
            fastp_path: Path to fastp executable
            getorganelle_path: Path to GetOrganelle script
            bwa_path: Path to BWA executable
            samtools_path: Path to samtools executable
            genome_type: Genome type for GetOrganelle (default: embplant_pt)
            k_values: K-mer values for assembly
            rounds: Number of GetOrganelle rounds
            use_nohup: Use nohup for long-running assemblies
            log_level: Logging level (DEBUG, INFO, WARNING, ERROR)
            force_mapping: Force mapping even if assembly is incomplete
            trim_poly_g: Enable poly-G tail trimming (for NovaSeq data)
            cut_mean_quality: Mean quality requirement for sliding window cutting
            skip_existing: Skip samples that have already been processed (default: True)
        """
        self.input_dir = Path(input_dir)
        self.output_dir = Path(output_dir)
        self.threads = threads
        self.fastp_path = fastp_path
        self.getorganelle_path = getorganelle_path
        self.bwa_path = bwa_path
        self.samtools_path = samtools_path
        self.genome_type = genome_type
        self.k_values = k_values
        self.rounds = rounds
        self.use_nohup = use_nohup
        self.force_mapping = force_mapping
        self.trim_poly_g = trim_poly_g
        self.cut_mean_quality = cut_mean_quality
        self.skip_existing = skip_existing
        
        # Create output directories
        self.qc_dir = self.output_dir / "01_QC"
        self.clean_reads_dir = self.output_dir / "02_clean_reads"
        self.assembly_dir = self.output_dir / "03_assemblies"
        self.mapping_dir = self.output_dir / "04_mapping"
        self.cp_reads_dir = self.output_dir / "05_cp_reads"
        self.reports_dir = self.output_dir / "06_reports"
        self.assembled_genomes_dir = self.output_dir / "07_assembled_genomes"
        
        for directory in [self.qc_dir, self.clean_reads_dir, self.assembly_dir, 
                         self.mapping_dir, self.cp_reads_dir, self.reports_dir,
                         self.assembled_genomes_dir]:
            directory.mkdir(parents=True, exist_ok=True)
        
        # Setup logging
        self._setup_logging(log_level)
        
        # Storage for results
        self.fastp_results: Dict[str, FastpStats] = {}
        self.assembly_results: Dict[str, AssemblyStats] = {}
        self.mapping_results: Dict[str, MappingStats] = {}
        self.failed_samples: Dict[str, str] = {}  # sample_name: error_message
        self.problematic_assemblies: Dict[str, str] = {}  # sample_name: reason (for >2 contigs)
    
    def _setup_logging(self, log_level: str):
        """Setup logging configuration"""
        log_file = self.output_dir / f"cgas_module1_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        
        logging.basicConfig(
            level=getattr(logging, log_level.upper()),
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler(sys.stdout)
            ]
        )
        self.logger = logging.getLogger(__name__)
        self.logger.info(f"CGAS Module 1 initialized. Log file: {log_file}")
    
    def _check_fastp_complete(self, sample_name: str) -> bool:
        """Check if Fastp QC has been completed for this sample"""
        if not self.skip_existing:
            return False
        
        # Check for clean reads and JSON report
        clean_r1 = self.clean_reads_dir / f"{sample_name}_clean_R1.fq.gz"
        json_report = self.qc_dir / sample_name / f"{sample_name}_fastp.json"
        
        if clean_r1.exists() and json_report.exists():
            self.logger.info(f"✓ Fastp already complete for {sample_name} - skipping")
            return True
        return False
    
    def _check_assembly_complete(self, sample_name: str) -> bool:
        """Check if GetOrganelle assembly has been completed for this sample"""
        if not self.skip_existing:
            return False
        
        # Check for assembly output directory and FASTA files
        assembly_dir = self.assembly_dir / f"{sample_name}_assembly"
        
        if not assembly_dir.exists():
            return False
        
        # Look for assembly FASTA files
        fasta_files = list(assembly_dir.glob("*.path_sequence.fasta"))
        if not fasta_files:
            fasta_files = list(assembly_dir.glob("*complete*.fasta"))
        if not fasta_files:
            fasta_files = list(assembly_dir.glob("*.fasta"))
        
        if fasta_files:
            self.logger.info(f"✓ Assembly already complete for {sample_name} - skipping")
            return True
        return False
    
    def _check_mapping_complete(self, sample_name: str) -> bool:
        """Check if read mapping has been completed for this sample"""
        if not self.skip_existing:
            return False
        
        # Check for sorted BAM file and CP reads
        sorted_bam = self.mapping_dir / sample_name / f"{sample_name}.sorted.bam"
        cp_reads_dir = self.cp_reads_dir / sample_name
        
        if sorted_bam.exists() and cp_reads_dir.exists():
            cp_r1 = cp_reads_dir / f"{sample_name}_cp_R1.fq.gz"
            if cp_r1.exists():
                self.logger.info(f"✓ Mapping already complete for {sample_name} - skipping")
                return True
        return False
    
    def find_read_pairs(self) -> List[ReadPair]:
        """
        Scan input directory and identify read pairs or single-end reads
        
        Returns:
            List of ReadPair objects
        """
        self.logger.info(f"Scanning input directory: {self.input_dir}")
        read_pairs = []
        processed_samples = set()
        
        # Get all files in input directory
        all_files = list(self.input_dir.glob("*"))
        
        for file_path in all_files:
            if not file_path.is_file():
                continue
            
            filename = file_path.name
            
            # Try to match forward read patterns
            for pattern in self.FORWARD_PATTERNS:
                match = re.match(pattern, filename)
                if match:
                    sample_name = match.group(1)
                    
                    # Skip if already processed
                    if sample_name in processed_samples:
                        continue
                    
                    # Find corresponding reverse read
                    reverse_read = self._find_reverse_read(file_path)
                    
                    if reverse_read and reverse_read.exists():
                        read_pairs.append(ReadPair(
                            sample_name=sample_name,
                            forward_read=file_path,
                            reverse_read=reverse_read,
                            is_paired=True
                        ))
                        self.logger.info(f"Found paired-end sample: {sample_name}")
                    else:
                        read_pairs.append(ReadPair(
                            sample_name=sample_name,
                            forward_read=file_path,
                            reverse_read=None,
                            is_paired=False
                        ))
                        self.logger.info(f"Found single-end sample: {sample_name}")
                    
                    processed_samples.add(sample_name)
                    break
        
        self.logger.info(f"Total samples found: {len(read_pairs)}")
        return read_pairs
    
    def _find_reverse_read(self, forward_path: Path) -> Optional[Path]:
        """Find corresponding reverse read for a forward read"""
        forward_name = forward_path.name
        
        # Define replacement patterns
        replacements = [
            ('_1.fastq', '_2.fastq'),
            ('_1.fastq.gz', '_2.fastq.gz'),
            ('_1.fq', '_2.fq'),
            ('_1.fq.gz', '_2.fq.gz'),
            ('_1.clean.fq.gz', '_2.clean.fq.gz'),
            ('_R1.fastq', '_R2.fastq'),
            ('_R1.fastq.gz', '_R2.fastq.gz'),
            ('_R1.fq', '_R2.fq'),
            ('_R1.fq.gz', '_R2.fq.gz'),
        ]
        
        for old, new in replacements:
            if old in forward_name:
                reverse_name = forward_name.replace(old, new)
                reverse_path = forward_path.parent / reverse_name
                if reverse_path.exists():
                    return reverse_path
        
        return None
    
    def run_fastp(self, read_pair: ReadPair) -> FastpStats:
        """
        Run Fastp quality control on reads
        
        Args:
            read_pair: ReadPair object containing read information
            
        Returns:
            FastpStats object with QC metrics
        """
        # Check if already completed
        if self._check_fastp_complete(read_pair.sample_name):
            # Load existing results
            json_report = self.qc_dir / read_pair.sample_name / f"{read_pair.sample_name}_fastp.json"
            stats = self._parse_fastp_json(json_report, read_pair.sample_name)
            self.fastp_results[read_pair.sample_name] = stats
            return stats
        
        self.logger.info(f"Running Fastp QC for {read_pair.sample_name}")
        
        # Output files
        sample_qc_dir = self.qc_dir / read_pair.sample_name
        sample_qc_dir.mkdir(exist_ok=True)
        
        clean_r1 = self.clean_reads_dir / f"{read_pair.sample_name}_clean_R1.fq.gz"
        json_report = sample_qc_dir / f"{read_pair.sample_name}_fastp.json"
        html_report = sample_qc_dir / f"{read_pair.sample_name}_fastp.html"
        
        # Build fastp command
        cmd = [
            self.fastp_path,
            "-i", str(read_pair.forward_read),
            "-o", str(clean_r1),
            "-j", str(json_report),
            "-h", str(html_report),
            "-w", str(self.threads),
            "--cut_mean_quality", str(self.cut_mean_quality),
        ]
        
        # Add poly-G trimming for NovaSeq data
        if self.trim_poly_g:
            cmd.append("--trim_poly_g")
        
        # Add paired-end specific options
        if read_pair.is_paired:
            cmd.append("--detect_adapter_for_pe")
            clean_r2 = self.clean_reads_dir / f"{read_pair.sample_name}_clean_R2.fq.gz"
            cmd.extend(["-I", str(read_pair.reverse_read), "-O", str(clean_r2)])
        
        try:
            subprocess.run(cmd, check=True, capture_output=True, text=True)
            self.logger.info(f"Fastp completed for {read_pair.sample_name}")
        except subprocess.CalledProcessError as e:
            self.logger.error(f"Fastp failed for {read_pair.sample_name}: {e.stderr}")
            raise
        
        # Parse JSON report
        stats = self._parse_fastp_json(json_report, read_pair.sample_name)
        self.fastp_results[read_pair.sample_name] = stats
        
        return stats
    
    def _parse_fastp_json(self, json_path: Path, sample_name: str) -> FastpStats:
        """Parse Fastp JSON report"""
        with open(json_path, 'r') as f:
            data = json.load(f)
        
        summary = data.get('summary', {})
        before_filtering = summary.get('before_filtering', {})
        after_filtering = summary.get('after_filtering', {})
        
        return FastpStats(
            sample_name=sample_name,
            total_reads_before=before_filtering.get('total_reads', 0),
            total_reads_after=after_filtering.get('total_reads', 0),
            total_bases_before=before_filtering.get('total_bases', 0),
            total_bases_after=after_filtering.get('total_bases', 0),
            q20_rate_before=before_filtering.get('q20_rate', 0.0),
            q20_rate_after=after_filtering.get('q20_rate', 0.0),
            q30_rate_before=before_filtering.get('q30_rate', 0.0),
            q30_rate_after=after_filtering.get('q30_rate', 0.0),
            gc_content=after_filtering.get('gc_content', 0.0),
            duplication_rate=data.get('duplication', {}).get('rate', 0.0)
        )
    
    def run_getorganelle(self, sample_name: str) -> AssemblyStats:
        """
        Run GetOrganelle assembly
        
        Args:
            sample_name: Sample name
            
        Returns:
            AssemblyStats object
        """
        # Check if already completed
        if self._check_assembly_complete(sample_name):
            # Load existing results
            assembly_out = self.assembly_dir / f"{sample_name}_assembly"
            stats = self._parse_getorganelle_output(assembly_out, sample_name)
            self.assembly_results[sample_name] = stats
            return stats
        
        self.logger.info(f"Running GetOrganelle assembly for {sample_name}")
        
        # Find clean reads
        clean_r1 = self.clean_reads_dir / f"{sample_name}_clean_R1.fq.gz"
        clean_r2 = self.clean_reads_dir / f"{sample_name}_clean_R2.fq.gz"
        
        # Check if paired or single
        is_paired = clean_r2.exists()
        
        # Output directory
        assembly_out = self.assembly_dir / f"{sample_name}_assembly"
        
        # Build GetOrganelle command
        cmd = [
            self.getorganelle_path,
            "-o", str(assembly_out),
            "-F", self.genome_type,
            "-R", str(self.rounds),
            "-k", self.k_values,
            "-t", str(self.threads),
            "--overwrite"
        ]
        
        # Add read files based on paired or single-end
        if is_paired:
            cmd.extend(["-1", str(clean_r1), "-2", str(clean_r2)])
            self.logger.info(f"Using paired-end reads for {sample_name}")
        else:
            cmd.extend(["-u", str(clean_r1)])
            self.logger.info(f"Using single-end reads for {sample_name}")
            self.logger.warning(
                f"⚠ Single-end mode detected for {sample_name}. "
                f"GetOrganelle may require more rounds (-R) or higher coverage for successful assembly."
            )
        
        # Add nohup if requested
        if self.use_nohup:
            log_file = self.assembly_dir / f"{sample_name}_getorganelle.log"
            cmd = ["nohup"] + cmd + [">", str(log_file), "2>&1", "&"]
        
        try:
            if self.use_nohup:
                # Run with shell for nohup
                subprocess.run(" ".join(cmd), shell=True, check=True)
                self.logger.info(f"GetOrganelle started in background for {sample_name}")
            else:
                subprocess.run(cmd, check=True, capture_output=True, text=True)
                self.logger.info(f"GetOrganelle completed for {sample_name}")
        except subprocess.CalledProcessError as e:
            self.logger.error(f"GetOrganelle failed for {sample_name}: {e}")
            raise
        
        # Parse assembly results
        stats = self._parse_getorganelle_output(assembly_out, sample_name)
        self.assembly_results[sample_name] = stats
        
        # If assembly is complete, copy to assembled_genomes directory
        if stats.is_complete:
            self._copy_assembled_genome(sample_name, assembly_out, stats)
        
        return stats
    
    def _copy_assembled_genome(self, sample_name: str, assembly_dir: Path, stats: AssemblyStats):
        """
        Copy complete assembled genomes to 07_assembled_genomes with clean naming
        Only copies up to 2 contigs (typical for SSC flip-flop)
        Creates warning file for assemblies with >2 contigs
        
        Args:
            sample_name: Sample name
            assembly_dir: GetOrganelle output directory
            stats: Assembly statistics
        """
        # Find path_sequence.fasta files
        fasta_files = sorted(assembly_dir.glob("*.path_sequence.fasta"))
        
        if not fasta_files:
            self.logger.warning(f"No path_sequence.fasta files found for {sample_name}")
            return
        
        num_contigs = len(fasta_files)
        
        # Check if assembly has too many contigs
        if num_contigs > 2:
            self.logger.warning(f"⚠ {sample_name} has {num_contigs} contigs (expected ≤2 for complete assembly)")
            self.problematic_assemblies[sample_name] = f"{num_contigs} contigs (expected 1-2)"
            return
        
        # Copy and rename contigs
        for idx, fasta_file in enumerate(fasta_files, 1):
            # New filename: SampleName_1.fasta, SampleName_2.fasta
            new_filename = f"{sample_name}_{idx}.fasta"
            output_path = self.assembled_genomes_dir / new_filename
            
            # Read original FASTA and update header
            try:
                header_count = 0
                with open(fasta_file, 'r') as f_in:
                    for line in f_in:
                        if line.startswith('>'):
                            header_count += 1
                
                # Check if file has multiple headers (fragmented assembly)
                if header_count > 1:
                    self.logger.warning(f"⚠ {sample_name}: {fasta_file.name} has {header_count} headers (fragmented assembly)")
                    self.problematic_assemblies[sample_name] = f"Fragmented assembly ({header_count} sequences in {fasta_file.name})"
                    return  # Stop processing this sample
                
                # If exactly 1 header, proceed with copying
                with open(fasta_file, 'r') as f_in, open(output_path, 'w') as f_out:
                    for line in f_in:
                        if line.startswith('>'):
                            # Replace with clean header
                            new_header = f">{sample_name}_{idx}\n"
                            f_out.write(new_header)
                        else:
                            f_out.write(line)
                
                self.logger.info(f"✓ Copied and renamed: {fasta_file.name} → {new_filename}")
                
            except Exception as e:
                self.logger.error(f"Failed to copy {fasta_file.name}: {e}")
    
    def _parse_getorganelle_output(self, assembly_dir: Path, sample_name: str) -> AssemblyStats:
        """
        Parse GetOrganelle output to assess assembly completeness
        
        Improved checks for:
        1. SSC flip-flop (2 contigs with LSC, SSC, IR annotations AND similar sizes)
        2. Single circular contig (from header or GFA file)
        3. Multiple incomplete contigs
        """
        self.logger.info(f"Parsing GetOrganelle output for {sample_name}")
        
        # Look for assembly files - prioritize path_sequence.fasta
        fasta_files = list(assembly_dir.glob("*.path_sequence.fasta"))
        if not fasta_files:
            fasta_files = list(assembly_dir.glob("*complete*.fasta"))
        if not fasta_files:
            fasta_files = list(assembly_dir.glob("*.fasta")) + list(assembly_dir.glob("*.fa"))
        
        if not fasta_files:
            self.logger.warning(f"No assembly files found for {sample_name}")
            return AssemblyStats(
                sample_name=sample_name,
                num_contigs=0,
                is_complete=False,
                is_circular=False,
                total_length=0
            )
        
        # Use the most specific file available
        assembly_file = fasta_files[0]
        self.logger.info(f"Using assembly file: {assembly_file.name}")
        
        # Parse FASTA file
        contigs = self._parse_fasta(assembly_file)
        num_contigs = len(contigs)
        
        # Check for circularity in FASTA headers
        is_circular = any('circular' in header.lower() for header, _ in contigs)
        
        # Check GFA file for circularity if not found in headers
        if not is_circular:
            gfa_files = list(assembly_dir.glob("*.gfa"))
            if gfa_files:
                is_circular = self._check_gfa_circularity(gfa_files[0])
        
        # Calculate total length
        total_length = sum(len(seq) for _, seq in contigs)
        
        # Parse GetOrganelle log for LSC/SSC/IR information
        log_file = assembly_dir / "get_org.log.txt"
        lsc_length, ssc_length, ir_length = None, None, None
        
        if log_file.exists():
            lsc_length, ssc_length, ir_length = self._parse_getorganelle_log(log_file)
        
        # Enhanced SSC flip-flop detection
        has_flipflop = False
        if num_contigs == 2 and lsc_length and ssc_length and ir_length:
            # Check if contig sizes are consistent with flip-flop
            expected_size = lsc_length + ssc_length + (2 * ir_length)
            contig_sizes = sorted([len(seq) for _, seq in contigs])
            
            # Both contigs should be approximately the same size and equal to expected genome size
            size_diff = abs(contig_sizes[0] - contig_sizes[1])
            avg_size = sum(contig_sizes) / 2
            
            if size_diff < 100 and abs(avg_size - expected_size) < 500:
                has_flipflop = True
                self.logger.info(f"SSC flip-flop detected: 2 contigs, sizes {contig_sizes}, expected {expected_size}")
            else:
                self.logger.warning(f"Two contigs found but sizes don't match flip-flop pattern: {contig_sizes} vs expected {expected_size}")
        
        # Determine completeness
        is_complete = False
        if has_flipflop:
            is_complete = True
            self.logger.info(f"Assembly COMPLETE: SSC flip-flop detected")
        elif num_contigs == 1 and is_circular:
            is_complete = True
            self.logger.info(f"Assembly COMPLETE: Single circular contig")
        elif num_contigs == 1 and lsc_length and ssc_length and ir_length:
            # Single contig with structure info - likely complete even if not marked circular
            expected_size = lsc_length + ssc_length + (2 * ir_length)
            if abs(total_length - expected_size) < 500:
                is_complete = True
                self.logger.info(f"Assembly COMPLETE: Single contig with correct structure")
        else:
            self.logger.warning(f"Assembly potentially INCOMPLETE: {num_contigs} contigs, circular={is_circular}")
        
        return AssemblyStats(
            sample_name=sample_name,
            num_contigs=num_contigs,
            is_complete=is_complete,
            is_circular=is_circular,
            total_length=total_length,
            lsc_length=lsc_length,
            ssc_length=ssc_length,
            ir_length=ir_length,
            has_flipflop=has_flipflop,
            assembly_path=assembly_file
        )
    
    def _parse_fasta(self, fasta_path: Path) -> List[Tuple[str, str]]:
        """Parse FASTA file and return list of (header, sequence) tuples"""
        contigs = []
        current_header = None
        current_seq = []
        
        with open(fasta_path, 'r') as f:
            for line in f:
                line = line.strip()
                if line.startswith('>'):
                    if current_header:
                        contigs.append((current_header, ''.join(current_seq)))
                    current_header = line[1:]
                    current_seq = []
                else:
                    current_seq.append(line)
            
            if current_header:
                contigs.append((current_header, ''.join(current_seq)))
        
        return contigs
    
    def _check_gfa_circularity(self, gfa_path: Path) -> bool:
        """
        Check if assembly graph shows circularity
        Looks for loops in GFA links
        """
        try:
            with open(gfa_path, 'r') as f:
                for line in f:
                    if line.startswith('L'):  # Link line
                        # GFA link format: L <from> <from_orient> <to> <to_orient> <overlap>
                        parts = line.strip().split('\t')
                        if len(parts) >= 5:
                            from_seg = parts[1]
                            to_seg = parts[3]
                            # If a segment links to itself, it's circular
                            if from_seg == to_seg:
                                return True
            return False
        except Exception as e:
            self.logger.warning(f"Could not parse GFA file {gfa_path}: {e}")
            return False
    
    def _parse_getorganelle_log(self, log_file: Path) -> Tuple[Optional[int], Optional[int], Optional[int]]:
        """Parse GetOrganelle log to extract LSC, SSC, IR lengths"""
        lsc_length = None
        ssc_length = None
        ir_length = None
        
        with open(log_file, 'r') as f:
            content = f.read()
            
            # First try to parse from "Detecting large repeats" line
            # Format: Total:LSC:SSC:Repeat(bp) = 160373:89521:20300:25276
            repeat_match = re.search(r'Total:LSC:SSC:Repeat\(bp\)\s*=\s*(\d+):(\d+):(\d+):(\d+)', content)
            if repeat_match:
                total_length = int(repeat_match.group(1))
                lsc_length = int(repeat_match.group(2))
                ssc_length = int(repeat_match.group(3))
                ir_length = int(repeat_match.group(4))
                self.logger.info(f"Parsed from 'Detecting large repeats': LSC={lsc_length}, SSC={ssc_length}, IR={ir_length}")
                return lsc_length, ssc_length, ir_length
            
            # Fallback: Look for patterns like "LSC: 12345 bp"
            lsc_match = re.search(r'LSC[:\s]+(\d+)', content, re.IGNORECASE)
            ssc_match = re.search(r'SSC[:\s]+(\d+)', content, re.IGNORECASE)
            ir_match = re.search(r'IR[ab]?[:\s]+(\d+)', content, re.IGNORECASE)
            
            if lsc_match:
                lsc_length = int(lsc_match.group(1))
            if ssc_match:
                ssc_length = int(ssc_match.group(1))
            if ir_match:
                ir_length = int(ir_match.group(1))
        
        return lsc_length, ssc_length, ir_length
    
    def run_mapping(self, sample_name: str) -> MappingStats:
        """
        Map clean reads to assembled chloroplast genome and calculate coverage
        
        Args:
            sample_name: Sample name
            
        Returns:
            MappingStats object
        """
        # Check if already completed
        if self._check_mapping_complete(sample_name):
            # Load existing results by recalculating from BAM
            sorted_bam = self.mapping_dir / sample_name / f"{sample_name}.sorted.bam"
            assembly_stats = self.assembly_results.get(sample_name)
            if assembly_stats:
                stats = self._calculate_coverage(sample_name, sorted_bam, assembly_stats)
                self.mapping_results[sample_name] = stats
                return stats
        
        self.logger.info(f"Running read mapping for {sample_name}")
        
        # Get assembly and clean reads
        assembly_stats = self.assembly_results.get(sample_name)
        if not assembly_stats or not assembly_stats.assembly_path:
            self.logger.error(f"No assembly found for {sample_name}")
            raise ValueError(f"No assembly found for {sample_name}")
        
        # Check if we should proceed with mapping
        if not assembly_stats.is_complete and not self.force_mapping:
            self.logger.warning(f"⚠ Assembly for {sample_name} is INCOMPLETE - Skipping mapping")
            self.logger.warning(f"  Use --force-mapping flag to map incomplete assemblies anyway")
            return MappingStats(
                sample_name=sample_name,
                total_reads=0,
                mapped_reads=0,
                mapping_rate=0.0,
                mean_coverage=0.0,
                median_coverage=0.0,
                cp_reads_saved=0
            )
        
        if not assembly_stats.is_complete and self.force_mapping:
            self.logger.info(f"Assembly for {sample_name} is incomplete, but proceeding with mapping (--force-mapping)")
        else:
            self.logger.info(f"Assembly for {sample_name} is complete, proceeding with mapping")
        
        assembly_fasta = assembly_stats.assembly_path
        clean_r1 = self.clean_reads_dir / f"{sample_name}_clean_R1.fq.gz"
        clean_r2 = self.clean_reads_dir / f"{sample_name}_clean_R2.fq.gz"
        is_paired = clean_r2.exists()
        
        # Diagnostic logging
        self.logger.info(f"Assembly FASTA: {assembly_fasta}")
        self.logger.info(f"Assembly exists: {assembly_fasta.exists()}")
        self.logger.info(f"Clean R1: {clean_r1}")
        self.logger.info(f"Clean R1 exists: {clean_r1.exists()}")
        self.logger.info(f"Clean R2: {clean_r2}")
        self.logger.info(f"Clean R2 exists: {clean_r2.exists()}")
        self.logger.info(f"Is paired-end: {is_paired}")
        
        # Check if assembly file exists
        if not assembly_fasta.exists():
            self.logger.error(f"Assembly FASTA file not found: {assembly_fasta}")
            return MappingStats(
                sample_name=sample_name,
                total_reads=0,
                mapped_reads=0,
                mapping_rate=0.0,
                mean_coverage=0.0,
                median_coverage=0.0,
                cp_reads_saved=0
            )
        
        # Check if clean reads exist
        if not clean_r1.exists():
            self.logger.error(f"Clean reads not found for {sample_name}: {clean_r1}")
            return MappingStats(
                sample_name=sample_name,
                total_reads=0,
                mapped_reads=0,
                mapping_rate=0.0,
                mean_coverage=0.0,
                median_coverage=0.0,
                cp_reads_saved=0
            )
        
        # Create mapping output directory
        sample_map_dir = self.mapping_dir / sample_name
        sample_map_dir.mkdir(exist_ok=True)
        
        # Index reference
        self.logger.info(f"Indexing reference for {sample_name}")
        try:
            result = subprocess.run([self.bwa_path, "index", str(assembly_fasta)], 
                          check=True, capture_output=True, text=True)
        except subprocess.CalledProcessError as e:
            self.logger.error(f"BWA indexing failed for {sample_name}")
            self.logger.error(f"Error: {e.stderr}")
            raise
        
        # Map reads
        sam_file = sample_map_dir / f"{sample_name}.sam"
        bam_file = sample_map_dir / f"{sample_name}.bam"
        sorted_bam = sample_map_dir / f"{sample_name}.sorted.bam"
        
        self.logger.info(f"Mapping reads for {sample_name} (paired: {is_paired})")
        try:
            with open(sam_file, 'w') as sam_out:
                if is_paired:
                    self.logger.info(f"Running: bwa mem -t {self.threads} {assembly_fasta} {clean_r1} {clean_r2}")
                    result = subprocess.run(
                        [self.bwa_path, "mem", "-t", str(self.threads), 
                         str(assembly_fasta), str(clean_r1), str(clean_r2)],
                        stdout=sam_out, stderr=subprocess.PIPE, text=True, check=True
                    )
                else:
                    self.logger.info(f"Running: bwa mem -t {self.threads} {assembly_fasta} {clean_r1}")
                    result = subprocess.run(
                        [self.bwa_path, "mem", "-t", str(self.threads), 
                         str(assembly_fasta), str(clean_r1)],
                        stdout=sam_out, stderr=subprocess.PIPE, text=True, check=True
                    )
            
            self.logger.info(f"BWA mapping completed for {sample_name}")
            
        except subprocess.CalledProcessError as e:
            self.logger.error(f"BWA mapping failed for {sample_name}")
            self.logger.error(f"Error: {e.stderr}")
            raise
        except Exception as e:
            self.logger.error(f"Unexpected error during mapping for {sample_name}: {e}")
            raise
        
        # Convert to BAM and sort
        self.logger.info(f"Converting SAM to BAM for {sample_name}")
        try:
            subprocess.run([self.samtools_path, "view", "-b", str(sam_file), "-o", str(bam_file)],
                          check=True, capture_output=True, text=True)
            subprocess.run([self.samtools_path, "sort", str(bam_file), "-o", str(sorted_bam)],
                          check=True, capture_output=True, text=True)
            subprocess.run([self.samtools_path, "index", str(sorted_bam)], 
                          check=True, capture_output=True, text=True)
            
            self.logger.info(f"BAM sorting and indexing completed for {sample_name}")
            
        except subprocess.CalledProcessError as e:
            self.logger.error(f"SAMtools processing failed for {sample_name}")
            self.logger.error(f"Error: {e.stderr}")
            raise
        
        # Extract mapped reads
        self._extract_cp_reads(sample_name, sorted_bam, clean_r1, clean_r2, is_paired)
        
        # Calculate coverage statistics
        stats = self._calculate_coverage(sample_name, sorted_bam, assembly_stats)
        self.mapping_results[sample_name] = stats
        
        # Cleanup
        sam_file.unlink()
        bam_file.unlink()
        
        return stats
    
    def _extract_cp_reads(self, sample_name: str, bam_file: Path, 
                         clean_r1: Path, clean_r2: Path, is_paired: bool):
        """Extract reads that mapped to chloroplast genome"""
        self.logger.info(f"Extracting chloroplast reads for {sample_name}")
        
        cp_reads_sample_dir = self.cp_reads_dir / sample_name
        cp_reads_sample_dir.mkdir(exist_ok=True)
        
        # Get mapped read names
        mapped_bam = cp_reads_sample_dir / "mapped.bam"
        subprocess.run([self.samtools_path, "view", "-b", "-F", "4", 
                       str(bam_file), "-o", str(mapped_bam)], check=True)
        
        # Convert to FASTQ
        cp_r1 = cp_reads_sample_dir / f"{sample_name}_cp_R1.fq.gz"
        
        if is_paired:
            cp_r2 = cp_reads_sample_dir / f"{sample_name}_cp_R2.fq.gz"
            subprocess.run([self.samtools_path, "fastq", "-1", str(cp_r1), 
                          "-2", str(cp_r2), "-0", "/dev/null", "-s", "/dev/null",
                          str(mapped_bam)], check=True)
            self.logger.info(f"Extracted paired-end CP reads: {cp_r1.name}, {cp_r2.name}")
        else:
            # For single-end: samtools fastq doesn't support -o flag
            # Use stdout redirection instead
            with open(cp_r1, "wb") as out_fq:
                subprocess.run([self.samtools_path, "fastq", str(mapped_bam)],
                             stdout=out_fq, check=True)
            self.logger.info(f"Extracted single-end CP reads: {cp_r1.name}")
        
        mapped_bam.unlink()
        self.logger.info(f"Chloroplast reads extracted for {sample_name}")
    
    def _calculate_coverage(self, sample_name: str, bam_file: Path, 
                           assembly_stats: AssemblyStats) -> MappingStats:
        """Calculate coverage statistics from BAM file"""
        # Get mapping statistics
        flagstat_output = subprocess.run(
            [self.samtools_path, "flagstat", str(bam_file)],
            capture_output=True, text=True, check=True
        ).stdout
        
        # Parse flagstat output
        total_reads = 0
        mapped_reads = 0
        for line in flagstat_output.split('\n'):
            if 'in total' in line:
                total_reads = int(line.split()[0])
            elif 'mapped (' in line and 'primary' not in line:
                mapped_reads = int(line.split()[0])
        
        mapping_rate = (mapped_reads / total_reads * 100) if total_reads > 0 else 0.0
        
        # Calculate coverage depth
        depth_output = subprocess.run(
            [self.samtools_path, "depth", str(bam_file)],
            capture_output=True, text=True, check=True
        ).stdout
        
        depths = []
        for line in depth_output.strip().split('\n'):
            if line:
                parts = line.split('\t')
                if len(parts) >= 3:
                    depths.append(int(parts[2]))
        
        mean_coverage = sum(depths) / len(depths) if depths else 0.0
        median_coverage = sorted(depths)[len(depths)//2] if depths else 0.0
        
        return MappingStats(
            sample_name=sample_name,
            total_reads=total_reads,
            mapped_reads=mapped_reads,
            mapping_rate=mapping_rate,
            mean_coverage=mean_coverage,
            median_coverage=median_coverage,
            cp_reads_saved=mapped_reads
        )
    
    def generate_reports(self):
        """Generate comprehensive summary reports"""
        self.logger.info("Generating summary reports")
        
        # Problematic assemblies report (>2 contigs)
        self._generate_problematic_assemblies_report()
        
        # Individual sample reports
        for sample_name in self.fastp_results.keys():
            self._generate_sample_report(sample_name)
        
        # Overall summary report
        self._generate_summary_report()
        
        self.logger.info(f"Reports generated in {self.reports_dir}")
    
    def _generate_problematic_assemblies_report(self):
        """Generate report for assemblies with >2 contigs"""
        if not self.problematic_assemblies:
            return
        
        report_file = self.reports_dir / "00_PROBLEMATIC_ASSEMBLIES.txt"
        
        with open(report_file, 'w') as f:
            f.write("=" * 80 + "\n")
            f.write("PROBLEMATIC ASSEMBLIES REPORT\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Total problematic assemblies: {len(self.problematic_assemblies)}\n\n")
            f.write("These samples were NOT copied to 07_assembled_genomes for the following reasons:\n\n")
            f.write("ISSUE TYPES:\n")
            f.write("  1. >2 contigs: More than 2 path_sequence.fasta files\n")
            f.write("     (Expected: 1 circular or 2 for SSC flip-flop)\n\n")
            f.write("  2. Fragmented assembly: Multiple sequences (headers) in single FASTA file\n")
            f.write("     (Expected: 1 header per path_sequence.fasta for complete genome)\n\n")
            f.write("-" * 80 + "\n\n")
            
            for sample_name in sorted(self.problematic_assemblies.keys()):
                reason = self.problematic_assemblies[sample_name]
                assembly_dir = self.assembly_dir / f"{sample_name}_assembly"
                
                f.write(f"Sample: {sample_name}\n")
                f.write(f"Issue: {reason}\n")
                f.write(f"Location: {assembly_dir}\n")
                f.write(f"Action: Manual inspection recommended\n")
                f.write("-" * 80 + "\n")
        
        self.logger.warning(f"⚠ {len(self.problematic_assemblies)} problematic assemblies. See {report_file}")
        
        # Also create a simple list
        list_file = self.reports_dir / "problematic_assemblies_list.txt"
        with open(list_file, 'w') as f:
            f.write("# Problematic assemblies (not copied to 07_assembled_genomes)\n")
            for sample_name in sorted(self.problematic_assemblies.keys()):
                f.write(f"{sample_name}\n")
    
    def _generate_sample_report(self, sample_name: str):
        """Generate detailed report for a single sample"""
        report_file = self.reports_dir / f"{sample_name}_report.txt"
        
        fastp_stats = self.fastp_results.get(sample_name)
        assembly_stats = self.assembly_results.get(sample_name)
        mapping_stats = self.mapping_results.get(sample_name)
        
        with open(report_file, 'w') as f:
            f.write("=" * 80 + "\n")
            f.write(f"CGAS Module 1 Report: {sample_name}\n")
            f.write("=" * 80 + "\n\n")
            
            # Fastp QC section
            if fastp_stats:
                f.write("1. QUALITY CONTROL (Fastp)\n")
                f.write("-" * 80 + "\n")
                f.write(f"Total reads before QC:  {fastp_stats.total_reads_before:,}\n")
                f.write(f"Total reads after QC:   {fastp_stats.total_reads_after:,}\n")
                f.write(f"Total bases before QC:  {fastp_stats.total_bases_before/1e9:.2f} Gb\n")
                f.write(f"Total bases after QC:   {fastp_stats.total_bases_after/1e9:.2f} Gb\n")
                f.write(f"Q20 rate before QC:     {fastp_stats.q20_rate_before*100:.2f}%\n")
                f.write(f"Q20 rate after QC:      {fastp_stats.q20_rate_after*100:.2f}%\n")
                f.write(f"Q30 rate before QC:     {fastp_stats.q30_rate_before*100:.2f}%\n")
                f.write(f"Q30 rate after QC:      {fastp_stats.q30_rate_after*100:.2f}%\n")
                f.write(f"GC content:             {fastp_stats.gc_content*100:.2f}%\n")
                f.write(f"Duplication rate:       {fastp_stats.duplication_rate*100:.2f}%\n")
                f.write("\n")
            
            # Assembly section
            if assembly_stats:
                f.write("2. CHLOROPLAST GENOME ASSEMBLY (GetOrganelle)\n")
                f.write("-" * 80 + "\n")
                f.write(f"Number of contigs:      {assembly_stats.num_contigs}\n")
                f.write(f"Assembly status:        {'COMPLETE' if assembly_stats.is_complete else 'INCOMPLETE'}\n")
                f.write(f"Circular:               {'Yes' if assembly_stats.is_circular else 'No'}\n")
                f.write(f"Total length:           {assembly_stats.total_length:,} bp\n")
                
                if assembly_stats.has_flipflop:
                    f.write(f"SSC flip-flop detected: Yes\n")
                
                if assembly_stats.lsc_length:
                    f.write(f"LSC length:             {assembly_stats.lsc_length:,} bp\n")
                if assembly_stats.ssc_length:
                    f.write(f"SSC length:             {assembly_stats.ssc_length:,} bp\n")
                if assembly_stats.ir_length:
                    f.write(f"IR length:              {assembly_stats.ir_length:,} bp\n")
                
                if assembly_stats.assembly_path:
                    f.write(f"Assembly file:          {assembly_stats.assembly_path}\n")
                f.write("\n")
            
            # Mapping section
            if mapping_stats:
                f.write("3. READ MAPPING AND COVERAGE\n")
                f.write("-" * 80 + "\n")
                f.write(f"Total reads:            {mapping_stats.total_reads:,}\n")
                f.write(f"Mapped reads:           {mapping_stats.mapped_reads:,}\n")
                f.write(f"Mapping rate:           {mapping_stats.mapping_rate:.2f}%\n")
                f.write(f"Mean coverage:          {mapping_stats.mean_coverage:.2f}x\n")
                f.write(f"Median coverage:        {mapping_stats.median_coverage:.2f}x\n")
                f.write(f"CP reads saved:         {mapping_stats.cp_reads_saved:,}\n")
                f.write("\n")
            
            # Recommendations
            f.write("4. RECOMMENDATIONS\n")
            f.write("-" * 80 + "\n")
            
            if assembly_stats:
                if assembly_stats.is_complete:
                    f.write("✓ Assembly is COMPLETE and ready for downstream CGAS analyses.\n")
                else:
                    f.write("⚠ Assembly may be INCOMPLETE. Please review manually:\n")
                    f.write(f"  - Check assembly output in: {self.assembly_dir / f'{sample_name}_assembly'}\n")
                    f.write("  - Consider adjusting GetOrganelle parameters\n")
                    f.write("  - Verify read quality and coverage\n")
            
            if mapping_stats and mapping_stats.total_reads > 0:
                if mapping_stats.mean_coverage < 30:
                    f.write(f"⚠ Low coverage ({mapping_stats.mean_coverage:.1f}x). Consider:\n")
                    f.write("  - Sequencing deeper\n")
                    f.write("  - Checking for contamination\n")
                elif mapping_stats.mean_coverage > 1000:
                    f.write(f"⚠ Very high coverage ({mapping_stats.mean_coverage:.1f}x).\n")
                    f.write("  - This is generally good but may indicate chloroplast enrichment\n")
            
            f.write("\n" + "=" * 80 + "\n")
    
    def _generate_summary_report(self):
        """Generate summary table for all samples in TSV and Excel formats"""
        summary_file = self.reports_dir / "00_SUMMARY_ALL_SAMPLES.tsv"
        excel_file = self.reports_dir / "00_SUMMARY_ALL_SAMPLES.xlsx"
        
        # Prepare headers and data
        headers = [
            "Sample",
            "Reads_Before_QC",
            "Reads_After_QC",
            "Bases_After_QC_Gb",
            "Q20_Rate_%",
            "Q30_Rate_%",
            "Num_Contigs",
            "Assembly_Status",
            "Circular",
            "Mapped_Reads",
            "Mapping_Rate_%",
            "Mean_Coverage",
            "Median_Coverage"
        ]
        
        # Collect all data rows
        data_rows = []
        for sample_name in sorted(self.fastp_results.keys()):
            fastp = self.fastp_results.get(sample_name)
            assembly = self.assembly_results.get(sample_name)
            mapping = self.mapping_results.get(sample_name)
            
            row = [
                sample_name,
                str(fastp.total_reads_before) if fastp else "NA",
                str(fastp.total_reads_after) if fastp else "NA",
                f"{fastp.total_bases_after/1e9:.2f}" if fastp else "NA",
                f"{fastp.q20_rate_after*100:.2f}" if fastp else "NA",
                f"{fastp.q30_rate_after*100:.2f}" if fastp else "NA",
                str(assembly.num_contigs) if assembly else "NA",
                "COMPLETE" if assembly and assembly.is_complete else "INCOMPLETE",
                "Yes" if assembly and assembly.is_circular else "No",
                str(mapping.mapped_reads) if mapping else "NA",
                f"{mapping.mapping_rate:.2f}" if mapping else "NA",
                f"{mapping.mean_coverage:.2f}" if mapping else "NA",
                f"{mapping.median_coverage:.2f}" if mapping else "NA"
            ]
            data_rows.append(row)
        
        # Write TSV file
        with open(summary_file, 'w') as f:
            f.write("\t".join(headers) + "\n")
            for row in data_rows:
                f.write("\t".join(row) + "\n")
        
        self.logger.info(f"Summary TSV report saved to {summary_file}")
        
        # Write Excel file with formatting
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "CGAS Module 1 Summary"
            
            # Write headers with formatting
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Write data rows
            for row_num, data_row in enumerate(data_rows, 2):
                for col_num, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Color code Assembly_Status column (now column 8, was column 8 before too)
                    if col_num == 8:  # Assembly_Status column
                        if value == "COMPLETE":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif value == "INCOMPLETE":
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = "A2"
            
            # Save Excel file
            wb.save(excel_file)
            self.logger.info(f"Summary Excel report saved to {excel_file}")
            
        except ImportError:
            self.logger.warning("openpyxl not installed. Skipping Excel report generation.")
            self.logger.warning("Install with: pip install openpyxl")
        except Exception as e:
            self.logger.error(f"Failed to generate Excel report: {e}")
        
        # Generate failed samples summary if any
        if self.failed_samples:
            failed_file = self.reports_dir / "00_FAILED_SAMPLES.txt"
            with open(failed_file, 'w') as f:
                f.write("=" * 80 + "\n")
                f.write("FAILED SAMPLES SUMMARY\n")
                f.write("=" * 80 + "\n\n")
                f.write(f"Total failed samples: {len(self.failed_samples)}\n\n")
                
                for sample, error in self.failed_samples.items():
                    f.write(f"Sample: {sample}\n")
                    f.write(f"Error: {error}\n")
                    f.write("-" * 80 + "\n")
            
            self.logger.warning(f"{len(self.failed_samples)} samples failed. See {failed_file}")
    
    def run_pipeline(self):
        """Run the complete Module 1 pipeline"""
        self.logger.info("Starting CGAS Module 1 pipeline")
        
        # Step 1: Find all read pairs
        read_pairs = self.find_read_pairs()
        
        if not read_pairs:
            self.logger.error("No read files found in input directory")
            return
        
        # Step 2-6: Process each sample
        for read_pair in read_pairs:
            try:
                self.logger.info(f"\n{'='*80}")
                self.logger.info(f"Processing sample: {read_pair.sample_name}")
                self.logger.info(f"{'='*80}\n")
                
                # Step 2: Quality control
                self.run_fastp(read_pair)
                
                # Step 3: Assembly
                self.run_getorganelle(read_pair.sample_name)
                
                # Step 4 & 5: Mapping and coverage (combined)
                try:
                    self.run_mapping(read_pair.sample_name)
                except Exception as e:
                    if "incomplete" in str(e).lower() and not self.force_mapping:
                        self.logger.warning(f"Skipping mapping for {read_pair.sample_name}: {e}")
                    else:
                        raise
                
            except Exception as e:
                self.logger.error(f"Error processing {read_pair.sample_name}: {e}")
                self.failed_samples[read_pair.sample_name] = str(e)
                continue
        
        # Step 6: Generate reports
        self.generate_reports()
        
        # Final summary
        self.logger.info("\n" + "="*80)
        self.logger.info("CGAS Module 1 pipeline completed!")
        self.logger.info(f"Total samples processed: {len(read_pairs)}")
        self.logger.info(f"Successful: {len(read_pairs) - len(self.failed_samples)}")
        self.logger.info(f"Failed: {len(self.failed_samples)}")
        
        # Count assembled genomes
        assembled_count = len(list(self.assembled_genomes_dir.glob("*.fasta")))
        self.logger.info(f"\nAssembled genomes ready for annotation: {assembled_count} FASTA files")
        self.logger.info(f"Location: {self.assembled_genomes_dir}")
        
        if self.problematic_assemblies:
            self.logger.warning(f"\n⚠ Problematic assemblies (not copied): {len(self.problematic_assemblies)}")
            self.logger.warning(f"  See {self.reports_dir / '00_PROBLEMATIC_ASSEMBLIES.txt'}")
        
        self.logger.info(f"\nAll results saved to: {self.output_dir}")
        self.logger.info("="*80)


def main():
    """Main entry point for command-line usage"""
    parser = argparse.ArgumentParser(
        description="CGAS Module 1: Chloroplast Genome Assembly and QC Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage (automatically skips completed samples)
  python cgas_module1.py -i raw_reads/ -o cgas_output/
  
  # Force rerun everything
  python cgas_module1.py -i raw_reads/ -o cgas_output/ --force-rerun
  
  # With custom threads and nohup
  python cgas_module1.py -i raw_reads/ -o cgas_output/ -t 16 --nohup
  
  # Custom k-mer values
  python cgas_module1.py -i raw_reads/ -o cgas_output/ -k 21,45,65,85,105,125
  
  # Force mapping for incomplete assemblies
  python cgas_module1.py -i raw_reads/ -o cgas_output/ --force-mapping
  
  # NovaSeq data with poly-G trimming
  python cgas_module1.py -i raw_reads/ -o cgas_output/ --trim-poly-g
        """
    )
    
    # Required arguments
    parser.add_argument("-i", "--input", required=True,
                       help="Input directory containing FASTQ files")
    parser.add_argument("-o", "--output", required=True,
                       help="Output directory for all results")
    
    # Optional arguments
    parser.add_argument("-t", "--threads", type=int, default=8,
                       help="Number of threads (default: 8)")
    parser.add_argument("-F", "--genome-type", default="embplant_pt",
                       help="Genome type for GetOrganelle (default: embplant_pt)")
    parser.add_argument("-k", "--kmers", default="21,45,65,85,105",
                       help="K-mer values for assembly (default: 21,45,65,85,105)")
    parser.add_argument("-R", "--rounds", type=int, default=15,
                       help="Number of GetOrganelle rounds (default: 15)")
    parser.add_argument("--nohup", action="store_true",
                       help="Use nohup for long-running assemblies")
    parser.add_argument("--log-level", default="INFO",
                       choices=["DEBUG", "INFO", "WARNING", "ERROR"],
                       help="Logging level (default: INFO)")
    
    # Assembly and mapping options
    parser.add_argument("--force-mapping", action="store_true",
                       help="Force read mapping even for incomplete assemblies")
    parser.add_argument("--skip-existing", action="store_true", default=True,
                       help="Skip samples that have already been processed (default: True)")
    parser.add_argument("--force-rerun", action="store_true",
                       help="Force rerun all steps even if outputs exist (overrides --skip-existing)")
    
    # Fastp options
    parser.add_argument("--trim-poly-g", action="store_true",
                       help="Enable poly-G tail trimming (for NovaSeq data)")
    parser.add_argument("--cut-mean-quality", type=int, default=20,
                       help="Mean quality requirement for sliding window (default: 20)")
    
    # Tool paths
    parser.add_argument("--fastp", default="fastp",
                       help="Path to fastp executable (default: fastp)")
    parser.add_argument("--getorganelle", default="get_organelle_from_reads.py",
                       help="Path to GetOrganelle script (default: get_organelle_from_reads.py)")
    parser.add_argument("--bwa", default="bwa",
                       help="Path to BWA executable (default: bwa)")
    parser.add_argument("--samtools", default="samtools",
                       help="Path to samtools executable (default: samtools)")
    
    args = parser.parse_args()
    
    # Handle force-rerun flag
    skip_existing = args.skip_existing and not args.force_rerun
    
    # Initialize and run pipeline
    pipeline = CGASModule1(
        input_dir=args.input,
        output_dir=args.output,
        threads=args.threads,
        fastp_path=args.fastp,
        getorganelle_path=args.getorganelle,
        bwa_path=args.bwa,
        samtools_path=args.samtools,
        genome_type=args.genome_type,
        k_values=args.kmers,
        rounds=args.rounds,
        use_nohup=args.nohup,
        log_level=args.log_level,
        force_mapping=args.force_mapping,
        trim_poly_g=args.trim_poly_g,
        cut_mean_quality=args.cut_mean_quality,
        skip_existing=skip_existing
    )
    
    pipeline.run_pipeline()


if __name__ == "__main__":
    main()

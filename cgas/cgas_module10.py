#!/usr/bin/env python3
"""
CGAS Module 10: SNP/Substitution Analysis with Graphical Visualization (Enhanced)
==================================================================================

This enhanced version includes:
- Automatic selection of representative genomes when >20 samples
- Alphabetical ordering of species in plots and legends
- Fully dynamic plot dimensions based on sample count
- Improved readability across all sample sizes

Author: Abdullah (Enhanced version)
Version: 2.1 (Module 10 - Enhanced Dynamic Plotting)
Date: February 2026

Dependencies:
    Python: openpyxl, numpy
    R: ggplot2, dplyr, tidyr, reshape2

Usage:
    python cgas_module10_enhanced.py
    python cgas_module10_enhanced.py -i alignments/
    python cgas_module10_enhanced.py --no-figures  # Skip R visualization

Output:
    Module10_SNP_Analysis/
        - Individual substitution files: [alignment_name]_Substitutions.xlsx
        - Merged analysis: Complete_Substitution_Analysis.xlsx
        - Merged CSV: Complete_Substitution_Analysis.csv (for R)
        - Figures/ (if R available):
            - Substitution_Types.pdf and .png
            - Substitution_Types_horizontal.pdf and .png
            - Ts_Tv_Ratio.pdf and .png

Notes:
    - FASTA files must contain exactly 2 sequences (reference and query)
    - Sequences must be pre-aligned and of equal length
    - Gaps (-) are ignored in the analysis
    - When >20 samples exist, 20 representative samples are auto-selected
"""

import os
import sys
import argparse
import subprocess
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError as e:
    print(f"Error: Required package not installed: {e}")
    print("Please install required packages using:")
    print("pip install openpyxl")
    sys.exit(1)

try:
    import numpy as np
    NUMPY_AVAILABLE = True
except ImportError:
    NUMPY_AVAILABLE = False
    print("Warning: numpy not installed. Representative selection will use simple method.")


# ============================================================================
# CONSTANTS
# ============================================================================

OUTPUT_FOLDER = "Module10_SNP_Analysis"
FIGURES_FOLDER = "Figures"
MAX_SAMPLES_FOR_PLOT = 20  # Maximum samples to plot


# ============================================================================
# SUBSTITUTION TYPE DEFINITIONS
# ============================================================================

# All possible nucleotide substitutions
SUBSTITUTION_TYPES = [
    'A_to_G', 'G_to_A', 'T_to_G', 'G_to_T',
    'A_to_C', 'C_to_A', 'C_to_T', 'T_to_C',
    'G_to_C', 'C_to_G', 'A_to_T', 'T_to_A'
]

# Transition substitutions (purine-purine or pyrimidine-pyrimidine)
TRANSITIONS = ['A_to_G', 'G_to_A', 'C_to_T', 'T_to_C']

# Transversion substitutions (purine-pyrimidine or pyrimidine-purine)
TRANSVERSIONS = ['T_to_G', 'G_to_T', 'A_to_C', 'C_to_A', 
                 'G_to_C', 'C_to_G', 'A_to_T', 'T_to_A']


def format_substitution_name(sub_type: str) -> str:
    """
    Convert substitution type from 'A_to_G' format to 'A→G' format.
    
    Parameters:
    -----------
    sub_type : str
        Substitution type in format 'BASE_to_BASE'
        
    Returns:
    --------
    str
        Formatted substitution name with arrow (e.g., 'A→G')
    """
    if '_to_' in sub_type:
        from_base, to_base = sub_type.split('_to_')
        return f"{from_base}→{to_base}"
    return sub_type


def extract_species_name(sequence_name: str) -> str:
    """
    Extract binomial species name (genus + species epithet) from sequence header.
    
    Parameters:
    -----------
    sequence_name : str
        Full sequence header from FASTA file
        
    Returns:
    --------
    str
        Formatted binomial name (e.g., "Hibiscus hamabo")
    """
    # Split by common delimiters and get first parts
    parts = sequence_name.replace('_', ' ').split()
    
    # Take first two words (genus + species)
    if len(parts) >= 2:
        genus = parts[0].capitalize()
        species = parts[1].lower()
        return f"{genus} {species}"
    elif len(parts) == 1:
        return parts[0].capitalize()
    else:
        return sequence_name


def format_species_comparison(seq_name1: str, seq_name2: str) -> str:
    """
    Format two species names as a pairwise comparison.
    
    Parameters:
    -----------
    seq_name1 : str
        First sequence header
    seq_name2 : str
        Second sequence header
        
    Returns:
    --------
    str
        Formatted comparison (e.g., "Hibiscus hamabo × Hibiscus moscheutos")
    """
    species1 = extract_species_name(seq_name1)
    species2 = extract_species_name(seq_name2)
    return f"{species1} × {species2}"


# ============================================================================
# SEQUENCE LOADING
# ============================================================================

def load_fasta_sequences(fasta_file: str) -> Tuple[List[str], List[str]]:
    """
    Load sequences from a FASTA file.
    
    Parameters:
    -----------
    fasta_file : str
        Path to the FASTA file
        
    Returns:
    --------
    Tuple[List[str], List[str]]
        Tuple of (sequence_names, sequences)
        
    Raises:
    -------
    ValueError
        If file doesn't contain exactly 2 sequences
        If sequences are not of equal length
    """
    sequence_names = []
    sequences = []
    current_sequence = ''
    current_name = ''
    
    try:
        with open(fasta_file, 'r') as file:
            for line in file:
                line = line.strip()
                
                if line.startswith('>'):
                    # Save previous sequence
                    if current_sequence:
                        sequences.append(current_sequence.upper())
                        sequence_names.append(current_name)
                    
                    # Start new sequence
                    current_name = line[1:].strip()
                    current_sequence = ''
                else:
                    current_sequence += line
            
            # Add last sequence
            if current_sequence:
                sequences.append(current_sequence.upper())
                sequence_names.append(current_name)
    
    except Exception as e:
        raise ValueError(f"Error reading FASTA file: {e}")
    
    # Validate sequences
    if len(sequences) != 2:
        raise ValueError(f"Expected 2 sequences, found {len(sequences)}. "
                        "This script requires pairwise alignments.")
    
    if len(sequences[0]) != len(sequences[1]):
        raise ValueError(f"Sequences have different lengths: {len(sequences[0])} vs {len(sequences[1])}. "
                        "Sequences must be pre-aligned.")
    
    return sequence_names, sequences


# ============================================================================
# SUBSTITUTION ANALYSIS
# ============================================================================

def analyze_substitutions(sequences: List[str]) -> Dict:
    """
    Analyze nucleotide substitutions between two aligned sequences.
    
    Parameters:
    -----------
    sequences : List[str]
        List containing exactly 2 aligned sequences
        
    Returns:
    --------
    Dict
        Dictionary containing substitution counts and positions
    """
    seq1, seq2 = sequences[0], sequences[1]
    
    # Initialize results dictionary
    results = {sub_type: {'count': 0, 'positions': []} 
              for sub_type in SUBSTITUTION_TYPES}
    
    # Analyze each position
    for pos, (base1, base2) in enumerate(zip(seq1, seq2), start=1):
        # Skip gaps
        if base1 == '-' or base2 == '-':
            continue
        
        # Skip matches
        if base1 == base2:
            continue
        
        # Record substitution
        sub_type = f"{base1}_to_{base2}"
        if sub_type in results:
            results[sub_type]['count'] += 1
            results[sub_type]['positions'].append(pos)
    
    # Calculate Ts/Tv
    ts_count = sum(results[t]['count'] for t in TRANSITIONS)
    tv_count = sum(results[t]['count'] for t in TRANSVERSIONS)
    tstv_ratio = ts_count / tv_count if tv_count > 0 else float('inf')
    
    results['stats'] = {
        'transitions': ts_count,
        'transversions': tv_count,
        'tstv_ratio': tstv_ratio,
        'total_substitutions': ts_count + tv_count
    }
    
    return results


# ============================================================================
# EXCEL FILE CREATION
# ============================================================================

def create_substitution_file(results: Dict, output_file: str, sequence_names: List[str]):
    """Create Excel file with substitution analysis."""
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Substitution Analysis"
    
    # Headers
    sheet['A1'] = "Substitution Type"
    sheet['B1'] = "Count"
    sheet['C1'] = "Positions"
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ['A1', 'B1', 'C1']:
        sheet[cell].fill = header_fill
        sheet[cell].font = header_font
        sheet[cell].alignment = Alignment(horizontal='center')
    
    # Write data
    row = 2
    for sub_type in SUBSTITUTION_TYPES:
        sheet[f'A{row}'] = format_substitution_name(sub_type)
        sheet[f'B{row}'] = results[sub_type]['count']
        positions = results[sub_type]['positions']
        sheet[f'C{row}'] = ', '.join(map(str, positions[:20]))
        if len(positions) > 20:
            sheet[f'C{row}'].value += f", ... ({len(positions)} total)"
        row += 1
    
    # Add summary statistics
    row += 1
    sheet[f'A{row}'] = "SUMMARY"
    sheet[f'A{row}'].font = Font(bold=True)
    row += 1
    
    stats = results['stats']
    sheet[f'A{row}'] = "Total Transitions"
    sheet[f'B{row}'] = stats['transitions']
    row += 1
    
    sheet[f'A{row}'] = "Total Transversions"
    sheet[f'B{row}'] = stats['transversions']
    row += 1
    
    sheet[f'A{row}'] = "Ts/Tv Ratio"
    sheet[f'B{row}'] = round(stats['tstv_ratio'], 3) if stats['tstv_ratio'] != float('inf') else "Inf"
    row += 1
    
    sheet[f'A{row}'] = "Total Substitutions"
    sheet[f'B{row}'] = stats['total_substitutions']
    
    # Auto-adjust column widths
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 80
    
    workbook.save(output_file)


# ============================================================================
# FILE DISCOVERY
# ============================================================================

def find_fasta_files(directory: str) -> List[str]:
    """Find all FASTA files in the specified directory."""
    fasta_extensions = ['.fasta', '.fa', '.fna', '.fas']
    fasta_files = []
    
    for file in os.listdir(directory):
        if any(file.lower().endswith(ext) for ext in fasta_extensions):
            fasta_files.append(os.path.join(directory, file))
    
    return sorted(fasta_files)


# ============================================================================
# PROCESSING
# ============================================================================

def process_single_fasta_file(fasta_file: str, output_folder: str) -> Tuple[Optional[str], Optional[Dict], Optional[List[str]]]:
    """Process a single FASTA file."""
    
    print(f"\nProcessing: {os.path.basename(fasta_file)}")
    
    try:
        # Load sequences
        sequence_names, sequences = load_fasta_sequences(fasta_file)
        print(f"  → Loaded 2 sequences:")
        print(f"     Seq 1: {extract_species_name(sequence_names[0])}")
        print(f"     Seq 2: {extract_species_name(sequence_names[1])}")
        print(f"  → Length: {len(sequences[0]):,} bp")
        
        # Analyze substitutions
        results = analyze_substitutions(sequences)
        total_subs = results['stats']['total_substitutions']
        print(f"  → Found {total_subs:,} substitutions")
        print(f"     Ts: {results['stats']['transitions']}, Tv: {results['stats']['transversions']}")
        print(f"     Ts/Tv: {results['stats']['tstv_ratio']:.3f}")
        
        # Create output file
        base_name = Path(fasta_file).stem
        output_file = os.path.join(output_folder, f"{base_name}_Substitutions.xlsx")
        create_substitution_file(results, output_file, sequence_names)
        print(f"  ✓ Saved: {os.path.basename(output_file)}")
        
        return output_file, results, sequence_names
    
    except Exception as e:
        print(f"  ✗ Error: {e}")
        return None, None, None


# ============================================================================
# MERGING RESULTS
# ============================================================================

def merge_substitution_files(sub_files: List[str], species_names: List[str], 
                            output_file: str) -> Tuple[Optional[Dict], Optional[List[str]]]:
    """Merge multiple substitution files into a single Excel file."""
    
    print(f"\n{'='*70}")
    print("MERGING RESULTS")
    print(f"{'='*70}")
    
    # Load all data
    all_data = {}
    file_names = []
    
    for sub_file, species_name in zip(sub_files, species_names):
        try:
            wb = openpyxl.load_workbook(sub_file)
            sheet = wb.active
            
            data = {}
            for row in range(2, 14):  # 12 substitution types
                sub_type = sheet[f'A{row}'].value
                count = sheet[f'B{row}'].value
                if sub_type and count is not None:
                    # Convert back to underscore format
                    sub_type = sub_type.replace('→', '_to_')
                    data[sub_type] = count
            
            # Get Ts/Tv ratio
            tstv_row = None
            for row in range(14, 20):
                cell_val = sheet[f'A{row}'].value
                if cell_val and "Ts/Tv" in str(cell_val):
                    tstv_row = row
                    break
            
            if tstv_row:
                tstv_value = sheet[f'B{tstv_row}'].value
                if tstv_value == "Inf":
                    tstv_value = float('inf')
                else:
                    try:
                        tstv_value = float(tstv_value)
                    except:
                        tstv_value = 0.0
                data['Ts/Tv Ratio'] = tstv_value
            
            all_data[species_name] = data
            file_names.append(species_name)
            
            wb.close()
            
        except Exception as e:
            print(f"  ⚠ Warning: Could not read {os.path.basename(sub_file)}: {e}")
    
    if not all_data:
        print("  ✗ No data to merge")
        return None, None
    
    # Create merged Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Merged Substitutions"
    
    # Headers
    sheet['A1'] = "Substitution Type"
    sheet['B1'] = "Category"
    
    col = 3
    for name in file_names:
        sheet.cell(row=1, column=col).value = name
        col += 1
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Write substitution data
    row = 2
    for sub_type in SUBSTITUTION_TYPES:
        sheet.cell(row=row, column=1).value = format_substitution_name(sub_type)
        sheet.cell(row=row, column=2).value = "Substitution"
        
        col = 3
        for name in file_names:
            value = all_data[name].get(sub_type, 0)
            sheet.cell(row=row, column=col).value = value
            col += 1
        
        row += 1
    
    # Write Ts/Tv ratio
    sheet.cell(row=row, column=1).value = "Ts/Tv Ratio"
    sheet.cell(row=row, column=2).value = "Ratio"
    
    col = 3
    for name in file_names:
        value = all_data[name].get('Ts/Tv Ratio', 0)
        sheet.cell(row=row, column=col).value = value if value != float('inf') else "Inf"
        col += 1
    
    # Auto-adjust column widths
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    for i in range(len(file_names)):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i + 3)].width = 15
    
    workbook.save(output_file)
    print(f"  ✓ Merged file created: {os.path.basename(output_file)}")
    print(f"  → Contains {len(file_names)} sample(s)")
    
    return all_data, file_names


# ============================================================================
# CSV CREATION FOR R
# ============================================================================

def create_csv_for_r(all_data: Dict, file_names: List[str], csv_file: str):
    """Create CSV file formatted for R visualization."""
    
    print(f"\nCreating CSV for R visualization...")
    
    with open(csv_file, 'w') as f:
        # Header
        f.write("Substitution_Type,Category," + ",".join(file_names) + "\n")
        
        # Substitution data
        for sub_type in SUBSTITUTION_TYPES:
            f.write(f"{sub_type},Substitution")
            for name in file_names:
                value = all_data[name].get(sub_type, 0)
                f.write(f",{value}")
            f.write("\n")
        
        # Ts/Tv ratio
        f.write("Ts/Tv Ratio,Ratio")
        for name in file_names:
            value = all_data[name].get('Ts/Tv Ratio', 0)
            if value == float('inf'):
                f.write(",Inf")
            else:
                f.write(f",{value:.6f}")
        f.write("\n")
    
    print(f"  ✓ CSV created: {os.path.basename(csv_file)}")


# ============================================================================
# REPRESENTATIVE SAMPLE SELECTION
# ============================================================================

def select_representative_samples(all_data: Dict, file_names: List[str], 
                                 max_samples: int = MAX_SAMPLES_FOR_PLOT) -> Tuple[List[str], str]:
    """
    Select representative samples when total exceeds max_samples.
    
    Selection strategy:
    1. Calculate variation coefficient for each sample
    2. Cluster samples into groups
    3. Select representatives from each group
    4. Sort alphabetically
    
    Parameters:
    -----------
    all_data : Dict
        Dictionary of all sample data
    file_names : List[str]
        List of all sample names
    max_samples : int
        Maximum number of samples to select
        
    Returns:
    --------
    Tuple[List[str], str]
        (selected sample names sorted alphabetically, selection method description)
    """
    
    if len(file_names) <= max_samples:
        # Return all samples sorted alphabetically
        return sorted(file_names), "all_samples"
    
    print(f"\n  → Selecting {max_samples} representative samples from {len(file_names)} total...")
    
    # Extract substitution counts for each sample
    sample_vectors = []
    for name in file_names:
        vector = [all_data[name].get(sub_type, 0) for sub_type in SUBSTITUTION_TYPES]
        sample_vectors.append(vector)
    
    if NUMPY_AVAILABLE:
        # Use numpy for sophisticated selection
        vectors = np.array(sample_vectors)
        
        # Calculate variation metrics
        totals = vectors.sum(axis=1)
        
        # Stratified selection based on total substitution counts
        sorted_indices = np.argsort(totals)
        
        # Select evenly distributed samples
        step = len(file_names) / max_samples
        selected_indices = [int(i * step) for i in range(max_samples)]
        selected_indices = [sorted_indices[i] for i in selected_indices]
        
        selected_names = [file_names[i] for i in selected_indices]
        method = "stratified_by_variation"
    
    else:
        # Simpler selection without numpy
        # Calculate total substitutions per sample
        totals = []
        for name in file_names:
            total = sum(all_data[name].get(sub_type, 0) for sub_type in SUBSTITUTION_TYPES)
            totals.append((total, name))
        
        # Sort by total
        totals.sort()
        
        # Select evenly spaced samples
        step = len(totals) / max_samples
        selected_names = [totals[int(i * step)][1] for i in range(max_samples)]
        method = "stratified_simple"
    
    # Sort alphabetically
    selected_names.sort()
    
    print(f"  → Selected {len(selected_names)} samples using {method} method")
    print(f"  → Samples will be ordered alphabetically in plots")
    
    return selected_names, method


# ============================================================================
# R VISUALIZATION
# ============================================================================

def check_r_packages():
    """Check if required R packages are installed."""
    
    r_check_script = '''
    packages <- c("ggplot2", "dplyr", "tidyr", "reshape2")
    missing <- packages[!packages %in% installed.packages()[,"Package"]]
    if(length(missing) > 0) {
        cat(paste("MISSING:", paste(missing, collapse=", ")))
        quit(status=1)
    } else {
        cat("OK")
        quit(status=0)
    }
    '''
    
    try:
        result = subprocess.run(['Rscript', '-e', r_check_script],
                              capture_output=True,
                              text=True,
                              timeout=30)
        
        if result.returncode == 0:
            return True, []
        else:
            missing = result.stdout.replace("MISSING:", "").strip().split(", ")
            return False, missing
    
    except Exception as e:
        return False, [str(e)]


def generate_r_visualizations(csv_file: str, output_folder: str, 
                              total_samples: int, all_data: Dict = None, 
                              file_names: List[str] = None):
    """Generate R-based visualizations."""
    
    print(f"\n{'='*70}")
    print("GENERATING VISUALIZATIONS")
    print(f"{'='*70}")
    
    # Check R installation
    try:
        result = subprocess.run(['Rscript', '--version'],
                              capture_output=True,
                              timeout=10)
        if result.returncode != 0:
            print("  ✗ R is not installed or not in PATH")
            print("    Please install R from: https://www.r-project.org/")
            return
    except Exception:
        print("  ✗ R is not installed or not in PATH")
        return
    
    print("  ✓ R is installed")
    
    # Check R packages
    packages_ok, missing = check_r_packages()
    if not packages_ok:
        print(f"  ✗ Missing R packages: {', '.join(missing)}")
        # Format package names with quotes
        quoted_packages = ', '.join([f'"{p}"' for p in missing])
        print(f"    Install with: install.packages(c({quoted_packages}))")
        return
    else:
        print("  ✓ All required R packages are installed")
    
    # Create figures directory
    figures_folder = os.path.join(output_folder, FIGURES_FOLDER)
    os.makedirs(figures_folder, exist_ok=True)
    
    # Determine which samples to plot
    if total_samples > MAX_SAMPLES_FOR_PLOT and all_data and file_names:
        selected_samples, method = select_representative_samples(all_data, file_names, MAX_SAMPLES_FOR_PLOT)
        num_samples_to_plot = len(selected_samples)
        
        # Create filtered CSV
        filtered_csv = os.path.join(output_folder, "Complete_Substitution_Analysis_filtered.csv")
        create_filtered_csv(all_data, selected_samples, filtered_csv)
        csv_for_plot = filtered_csv
    else:
        # Sort all samples alphabetically
        if file_names:
            selected_samples = sorted(file_names)
            # Create sorted CSV
            filtered_csv = os.path.join(output_folder, "Complete_Substitution_Analysis_sorted.csv")
            create_filtered_csv(all_data, selected_samples, filtered_csv)
            csv_for_plot = filtered_csv
        else:
            selected_samples = None
            csv_for_plot = csv_file
        
        num_samples_to_plot = total_samples
    
    # Create R script
    r_script_file = os.path.join(output_folder, "generate_snp_plots.R")
    r_script_content = create_enhanced_r_script(
        os.path.basename(csv_for_plot), 
        num_samples_to_plot
    )
    
    with open(r_script_file, 'w') as f:
        f.write(r_script_content)
    
    print(f"  → R script created: {os.path.basename(r_script_file)}")
    
    # Execute R script
    print(f"  → Executing R script...")
    try:
        result = subprocess.run(['Rscript', r_script_file],
                              capture_output=True,
                              text=True,
                              timeout=120,
                              cwd=output_folder)
        
        # Always show R output for debugging
        if result.stdout:
            print(result.stdout)
        
        if result.returncode == 0:
            print(f"  ✓ R visualization completed successfully")
            print(f"\n  Generated figures in {FIGURES_FOLDER}/:")
            print(f"    - Substitution_Types_vertical.pdf and .png")
            print(f"    - Substitution_Types_horizontal.pdf and .png")
            print(f"    - Ts_Tv_Ratio.pdf and .png")
            
            # Verify files exist
            import glob
            pdf_files = glob.glob(os.path.join(figures_folder, "*.pdf"))
            png_files = glob.glob(os.path.join(figures_folder, "*.png"))
            if pdf_files or png_files:
                print(f"\n  Files created:")
                for f in sorted(pdf_files + png_files):
                    size = os.path.getsize(f)
                    print(f"    ✓ {os.path.basename(f)} ({size:,} bytes)")
            else:
                print(f"\n  ⚠ Warning: No figure files found in {FIGURES_FOLDER}/")
        else:
            print(f"  ✗ R script execution failed (return code: {result.returncode})")
            if result.stderr:
                print(f"\n  Error output:")
                print(result.stderr[:1000])
    
    except subprocess.TimeoutExpired:
        print(f"  ✗ R script timed out (>120s)")
    except Exception as e:
        print(f"  ✗ Error executing R script: {e}")


def create_filtered_csv(all_data: Dict, selected_samples: List[str], csv_file: str):
    """Create CSV with only selected samples, sorted alphabetically."""
    
    with open(csv_file, 'w') as f:
        # Header
        f.write("Substitution_Type,Category," + ",".join(selected_samples) + "\n")
        
        # Substitution data
        for sub_type in SUBSTITUTION_TYPES:
            f.write(f"{sub_type},Substitution")
            for name in selected_samples:
                value = all_data[name].get(sub_type, 0)
                f.write(f",{value}")
            f.write("\n")
        
        # Ts/Tv ratio
        f.write("Ts/Tv Ratio,Ratio")
        for name in selected_samples:
            value = all_data[name].get('Ts/Tv Ratio', 0)
            if value == float('inf'):
                f.write(",Inf")
            else:
                f.write(f",{value:.6f}")
        f.write("\n")


def create_enhanced_r_script(csv_file: str, num_samples: int) -> str:
    """
    Create enhanced R script with fully dynamic plot sizing.
    
    This version implements:
    - Dynamic plot dimensions based on sample count
    - Adaptive font sizes
    - Both vertical and horizontal layouts
    - Publication-quality styling
    - Alphabetical ordering (handled in CSV creation)
    """
    
    script = f'''
# CGAS Module 10: Enhanced SNP/Substitution Visualization Script
# Generated automatically with dynamic plotting capabilities

suppressPackageStartupMessages({{
  library(ggplot2)
  library(dplyr)
  library(tidyr)
  library(reshape2)
}})

cat("\\n========================================\\n")
cat("CGAS MODULE 10: ENHANCED VISUALIZATION\\n")
cat("========================================\\n\\n")

# Read data
cat("Reading SNP data from CSV...\\n")
data <- read.csv("{csv_file}", check.names=FALSE)

# Safety check: Empty CSV
if (nrow(data) == 0) {{
  stop("ERROR: CSV file is empty - no data to plot")
}}

cat(paste("  ✓ Loaded:", nrow(data), "rows\\n\\n"))

# Create figures directory
if (!dir.exists("Figures")) {{
  dir.create("Figures")
}}

# ============================================================================
# DYNAMIC PLOT SIZING
# ============================================================================

# Get sample columns
sample_cols <- names(data)[!(names(data) %in% c("Substitution_Type", "Category"))]
num_samples <- length(sample_cols)

cat(paste("  → Number of samples:", num_samples, "\\n"))
cat(paste("  → Samples are alphabetically ordered\\n\\n"))

# Calculate dynamic dimensions
if (num_samples == 1) {{
  plot_width_vert <- 4
  plot_width_horiz <- 8
  plot_height_vert <- 6
  plot_height_horiz <- 3
  base_font <- 14
  axis_font <- 12
  axis_title_font <- 13
  title_font <- 16
  legend_font <- 11
  bar_width <- 0.5
}} else if (num_samples == 2) {{
  plot_width_vert <- 5
  plot_width_horiz <- 8
  plot_height_vert <- 6
  plot_height_horiz <- 4
  base_font <- 13
  axis_font <- 11
  axis_title_font <- 12
  title_font <- 15
  legend_font <- 10
  bar_width <- 0.6
}} else if (num_samples <= 5) {{
  plot_width_vert <- 6 + (num_samples - 2) * 0.8
  plot_width_horiz <- 9
  plot_height_vert <- 6
  plot_height_horiz <- 4 + (num_samples - 2) * 0.3
  base_font <- 12
  axis_font <- 10
  axis_title_font <- 11
  title_font <- 14
  legend_font <- 10
  bar_width <- 0.7
}} else if (num_samples <= 10) {{
  plot_width_vert <- 8.4 + (num_samples - 5) * 0.7
  plot_width_horiz <- 10
  plot_height_vert <- 6.5
  plot_height_horiz <- 6 + (num_samples - 5) * 0.35
  base_font <- 11
  axis_font <- 9
  axis_title_font <- 10
  title_font <- 13
  legend_font <- 9
  bar_width <- 0.7
}} else if (num_samples <= 15) {{
  plot_width_vert <- 11.9 + (num_samples - 10) * 0.5
  plot_width_horiz <- 11
  plot_height_vert <- 7
  plot_height_horiz <- 7.75 + (num_samples - 10) * 0.3
  base_font <- 10
  axis_font <- 8.5
  axis_title_font <- 9.5
  title_font <- 12
  legend_font <- 8.5
  bar_width <- 0.7
}} else if (num_samples <= 20) {{
  plot_width_vert <- 14.4 + (num_samples - 15) * 0.4
  plot_width_horiz <- 12
  plot_height_vert <- 7.5
  plot_height_horiz <- 9.25 + (num_samples - 15) * 0.25
  base_font <- 9.5
  axis_font <- 8
  axis_title_font <- 9
  title_font <- 11.5
  legend_font <- 8
  bar_width <- 0.7
}} else {{
  # Fallback for >20 (should not occur with filtering)
  plot_width_vert <- 16
  plot_width_horiz <- 13
  plot_height_vert <- 8
  plot_height_horiz <- 12
  base_font <- 9
  axis_font <- 7.5
  axis_title_font <- 8.5
  title_font <- 11
  legend_font <- 7.5
  bar_width <- 0.65
}}

cat(paste("  → Vertical plot dimensions:", round(plot_width_vert, 1), "×", 
          plot_height_vert, "inches\\n"))
cat(paste("  → Horizontal plot dimensions:", round(plot_width_horiz, 1), "×", 
          round(plot_height_horiz, 1), "inches\\n\\n"))

# ============================================================================
# DATA PREPARATION
# ============================================================================

# Filter substitution data
sub_data <- data[data$Category == "Substitution", ]

if (nrow(sub_data) == 0) {{
  stop("ERROR: No substitution data found in CSV")
}}

# Define complementary pair mapping
pair_map <- data.frame(
  Original = c("A_to_G", "G_to_A", "C_to_T", "T_to_C", "C_to_G", "G_to_C",
               "A_to_T", "T_to_A", "A_to_C", "C_to_A", "T_to_G", "G_to_T"),
  Combined = c("A↔G", "A↔G", "C↔T", "C↔T", "C↔G", "C↔G",
               "A↔T", "A↔T", "A↔C", "A↔C", "T↔G", "T↔G"),
  Type = c("Transition", "Transition", "Transition", "Transition",
           "Transversion", "Transversion", "Transversion", "Transversion",
           "Transversion", "Transversion", "Transversion", "Transversion"),
  stringsAsFactors = FALSE
)

# Merge and combine pairs
sub_data <- merge(sub_data, pair_map,
                  by.x = "Substitution_Type", by.y = "Original", all.x = TRUE)

# Sum complementary pairs
combined_data <- sub_data %>%
  group_by(Combined, Type) %>%
  summarise(across(all_of(sample_cols), sum, na.rm = TRUE), .groups = "drop")

# Convert to long format
plot_data <- combined_data %>%
  pivot_longer(cols = all_of(sample_cols), names_to = "Sample", values_to = "Count")

# Order substitution types (transitions first)
sub_order <- c("A↔G", "C↔T", "A↔C", "T↔G", "C↔G", "A↔T")
plot_data$Combined <- factor(plot_data$Combined, levels = sub_order)

# Samples should already be alphabetically ordered from CSV
# Preserve that order
plot_data$Sample <- factor(plot_data$Sample, levels = sample_cols)

if (nrow(plot_data) == 0) {{
  stop("ERROR: No data to plot after processing")
}}

# ============================================================================
# COLOR SCHEME
# ============================================================================

# Publication-quality color scheme for 6 substitution types
sub_colors <- c(
  "A↔G" = "#3498DB",    # Blue (Transition)
  "C↔T" = "#E74C3C",    # Red (Transition)
  "A↔C" = "#95A5A6",    # Gray (Transversion)
  "T↔G" = "#2ECC71",    # Green (Transversion)
  "C↔G" = "#F39C12",    # Orange (Transversion)
  "A↔T" = "#9B59B6"     # Purple (Transversion)
)

# ============================================================================
# PUBLICATION-QUALITY THEME
# ============================================================================

pub_theme <- theme_bw(base_size = base_font) +
  theme(
    panel.grid = element_blank(),
    axis.line = element_line(color = "black", linewidth = 0.8),
    axis.ticks = element_line(color = "black", linewidth = 0.6),
    axis.ticks.length = unit(0.25, "cm"),
    axis.text = element_text(color = "black"),
    axis.title = element_text(face = "bold"),
    legend.title = element_text(face = "bold", size = legend_font),
    legend.text = element_text(face = "italic", size = legend_font),
    legend.position = "right",
    legend.key.size = unit(0.8, "cm"),
    plot.title = element_text(size = title_font, face = "bold", hjust = 0.5),
    plot.margin = margin(t = 10, r = 10, b = 10, l = 10)
  )

# ============================================================================
# FIGURE 1: VERTICAL BAR PLOT (GROUPED BARS)
# ============================================================================

cat("Generating vertical grouped bar plot...\\n")

p_vertical <- ggplot(plot_data, aes(x = Combined, y = Count, fill = Sample)) +
  geom_bar(
    stat = "identity",
    position = position_dodge(width = 0.85),
    width = bar_width,
    color = "black",
    linewidth = 0.3
  ) +
  scale_fill_manual(
    values = colorRampPalette(c("#3498DB", "#E74C3C", "#2ECC71", "#F39C12", 
                                 "#9B59B6", "#1ABC9C", "#34495E", "#E67E22"))(num_samples),
    name = "Species"
  ) +
  labs(
    x = "Substitution type",
    y = "Number of substitutions",
    title = "Nucleotide Substitution Patterns"
  ) +
  pub_theme +
  theme(
    axis.text.x = element_text(angle = 0, hjust = 0.5, size = axis_font),
    axis.text.y = element_text(size = axis_font),
    axis.title = element_text(size = axis_title_font)
  ) +
  scale_y_continuous(expand = expansion(mult = c(0, 0.05)))

# Save vertical plot
pdf("Figures/Substitution_Types_vertical.pdf", width = plot_width_vert, height = plot_height_vert)
print(p_vertical)
dev.off()

ggsave("Figures/Substitution_Types_vertical.png", plot = p_vertical, 
       width = plot_width_vert, height = plot_height_vert, dpi = 600)

cat("  ✓ Vertical plot saved\\n\\n")

# ============================================================================
# FIGURE 2: HORIZONTAL BAR PLOT (GROUPED BARS)
# ============================================================================

cat("Generating horizontal grouped bar plot...\\n")

p_horizontal <- ggplot(plot_data, aes(x = Combined, y = Count, fill = Sample)) +
  geom_bar(
    stat = "identity",
    position = position_dodge(width = 0.85),
    width = bar_width,
    color = "black",
    linewidth = 0.3
  ) +
  scale_fill_manual(
    values = colorRampPalette(c("#3498DB", "#E74C3C", "#2ECC71", "#F39C12", 
                                 "#9B59B6", "#1ABC9C", "#34495E", "#E67E22"))(num_samples),
    name = "Species"
  ) +
  coord_flip() +
  labs(
    x = "Substitution type",
    y = "Number of substitutions",
    title = "Nucleotide Substitution Patterns"
  ) +
  pub_theme +
  theme(
    axis.text.x = element_text(angle = 0, hjust = 0.5, size = axis_font),
    axis.text.y = element_text(size = axis_font),
    axis.title = element_text(size = axis_title_font)
  ) +
  scale_y_continuous(expand = expansion(mult = c(0, 0.05)))

# Save horizontal plot
pdf("Figures/Substitution_Types_horizontal.pdf", width = plot_width_horiz, height = plot_height_horiz)
print(p_horizontal)
dev.off()

ggsave("Figures/Substitution_Types_horizontal.png", plot = p_horizontal, 
       width = plot_width_horiz, height = plot_height_horiz, dpi = 600)

cat("  ✓ Horizontal plot saved\\n\\n")

# ============================================================================
# FIGURE 3: TS/TV RATIO
# ============================================================================

cat("Generating Ts/Tv ratio graph...\\n")

# Extract Ts/Tv data
tstv_data <- data[data$Substitution_Type == "Ts/Tv Ratio", ]

if (nrow(tstv_data) == 0) {{
  stop("ERROR: No Ts/Tv Ratio found in CSV")
}}

# Convert to long format
tstv_long <- tstv_data %>%
  select(-Category, -Substitution_Type) %>%
  pivot_longer(cols = everything(), names_to = "Sample", values_to = "Ratio")

# Convert Inf to numeric (will be handled in plotting)
tstv_long$Ratio <- as.numeric(tstv_long$Ratio)

# Remove NA values
tstv_long <- tstv_long[!is.na(tstv_long$Ratio) & is.finite(tstv_long$Ratio), ]

if (nrow(tstv_long) == 0) {{
  cat("  ⚠ Warning: All Ts/Tv ratios are NA or Inf - skipping Ts/Tv plot\\n")
}} else {{
  # Preserve alphabetical order
  tstv_long$Sample <- factor(tstv_long$Sample, levels = sample_cols)
  
  # Create bar plot
  p_tstv <- ggplot(tstv_long, aes(x = Sample, y = Ratio)) +
    geom_bar(stat = "identity", fill = "#2ECC71", color = "black", 
             linewidth = 0.5, width = bar_width) +
    geom_hline(yintercept = 0, color = "black", linewidth = 0.5) +
    labs(
      x = "",
      y = "Ts/Tv Ratio",
      title = "Transition/Transversion Ratio"
    ) +
    pub_theme +
    theme(
      axis.text.x = element_text(angle = 90, hjust = 1, vjust = 0.5, 
                                 size = axis_font, face = "italic"),
      axis.text.y = element_text(size = axis_font),
      axis.title.y = element_text(size = axis_title_font),
      plot.title = element_text(size = title_font),
      legend.position = "none"
    )
  
  # Safe Y-axis scaling
  max_y <- max(tstv_long$Ratio, na.rm = TRUE)
  if (!is.finite(max_y) || max_y <= 0) {{
    max_y <- 2.0
  }}
  
  p_tstv <- p_tstv + scale_y_continuous(
    expand = expansion(mult = c(0, 0.1)),
    breaks = seq(0, ceiling(max_y * 5) / 5, by = 0.2)
  )
  
  # Save Ts/Tv plot
  pdf("Figures/Ts_Tv_Ratio.pdf", width = plot_width_vert, height = plot_height_vert)
  print(p_tstv)
  dev.off()
  
  ggsave("Figures/Ts_Tv_Ratio.png", plot = p_tstv, 
         width = plot_width_vert, height = plot_height_vert, dpi = 600)
  
  cat("  ✓ Ts/Tv ratio plot saved\\n\\n")
}}

cat("========================================\\n")
cat("All visualizations completed!\\n")
cat("========================================\\n")
cat("\\nSummary:\\n")
cat("  - Complementary pairs combined (e.g., C→G + G→C = C↔G)\\n")
cat("  - Total substitution pairs: 6\\n")
cat("  - Transitions: A↔G, C↔T\\n")
cat("  - Transversions: A↔T, C↔G, A↔C, T↔G\\n")
cat("  - Samples ordered alphabetically\\n")
cat(paste("  - Plot dimensions adapted for", num_samples, "samples\\n"))
'''
    
    return script


# ============================================================================
# MAIN PIPELINE
# ============================================================================

def main():
    """Main pipeline execution function with command-line arguments."""
    
    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description='CGAS Module 10: Enhanced SNP/Substitution Analysis',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  %(prog)s                          # Process current directory
  %(prog)s -i alignments/           # Process specific directory
  %(prog)s -i data/ -o results/     # Custom input and output
  %(prog)s --no-figures             # Skip R visualization

Features:
  - Automatic selection of 20 representative samples when >20 total
  - Alphabetical ordering of species in plots and legends
  - Fully dynamic plot dimensions (1-20+ samples)
  - Publication-quality vertical and horizontal layouts

For more information, see the CGAS documentation.
        '''
    )
    
    parser.add_argument('-i', '--input',
                       type=str,
                       default='.',
                       help='Input directory containing FASTA alignment files (default: current directory)')
    
    parser.add_argument('-o', '--output',
                       type=str,
                       default=None,
                       help='Output directory (default: Module10_SNP_Analysis in input directory)')
    
    parser.add_argument('--no-figures',
                       action='store_true',
                       help='Skip R figure generation')
    
    args = parser.parse_args()
    
    # Set up directories
    input_dir = os.path.abspath(args.input)
    
    if args.output:
        output_dir = os.path.abspath(args.output)
    else:
        output_dir = os.path.join(input_dir, OUTPUT_FOLDER)
    
    print(f"\n{'='*70}")
    print("CGAS MODULE 10: ENHANCED SNP/SUBSTITUTION ANALYSIS")
    print(f"{'='*70}")
    print(f"\nInput directory: {input_dir}")
    print(f"Output directory: {output_dir}")
    
    # Create output folder
    os.makedirs(output_dir, exist_ok=True)
    print(f"Output folder: {os.path.basename(output_dir)}/")
    
    # Find FASTA files
    fasta_files = find_fasta_files(input_dir)
    
    if not fasta_files:
        print("\n❌ ERROR: No FASTA files found!")
        print("Please place FASTA alignment files (.fasta, .fa, .fna) in the working directory.")
        return
    
    print(f"\nFound {len(fasta_files)} FASTA file(s):")
    for ff in fasta_files:
        print(f"  - {os.path.basename(ff)}")
    
    # Process each FASTA file
    print(f"\n{'='*70}")
    print("ANALYZING NUCLEOTIDE SUBSTITUTIONS")
    print(f"{'='*70}")
    
    sub_files = []
    species_comparisons = []
    
    for fasta_file in fasta_files:
        try:
            output_file, results, sequence_names = process_single_fasta_file(fasta_file, output_dir)
            if output_file:
                sub_files.append(output_file)
                # Create species comparison string
                if sequence_names:
                    comparison = format_species_comparison(sequence_names[0], sequence_names[1])
                    species_comparisons.append(comparison)
                else:
                    # Fallback to filename if sequence names not available
                    species_comparisons.append(Path(fasta_file).stem)
        except Exception as e:
            print(f"  ❌ ERROR processing {os.path.basename(fasta_file)}: {e}")
            continue
    
    # Merge results
    all_data = None
    file_names = None
    if sub_files:
        merged_file = os.path.join(output_dir, "Complete_Substitution_Analysis.xlsx")
        all_data, file_names = merge_substitution_files(sub_files, species_comparisons, merged_file)
        
        # Create CSV for R visualization
        if not args.no_figures and all_data and file_names:
            csv_file = os.path.join(output_dir, "Complete_Substitution_Analysis.csv")
            create_csv_for_r(all_data, file_names, csv_file)
            
            # Generate R visualizations
            generate_r_visualizations(csv_file, output_dir, len(file_names), 
                                     all_data, file_names)
    
    # Summary
    print(f"\n{'='*70}")
    print("ANALYSIS COMPLETE")
    print(f"{'='*70}")
    print(f"✓ Successfully processed: {len(sub_files)} alignment(s)")
    print(f"✓ Individual substitution files: {len(sub_files)}")
    if sub_files:
        print(f"✓ Merged analysis file: {os.path.basename(merged_file)}")
    if not args.no_figures and all_data:
        print(f"✓ R visualizations: {FIGURES_FOLDER}/")
        if len(file_names) > MAX_SAMPLES_FOR_PLOT:
            print(f"  → {MAX_SAMPLES_FOR_PLOT} representative samples plotted from {len(file_names)} total")
        print(f"  → All samples ordered alphabetically")
    print(f"\nAll output files saved in: {output_dir}/")
    print(f"{'='*70}\n")


# ============================================================================
# SCRIPT ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠ Analysis interrupted by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\n\n❌ FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

#!/usr/bin/env python3
"""
CGAS Module 11: Gene and tRNA Intron Extraction, Analysis, and Visualization
=============================================================================

This script extracts and analyzes intron positions and lengths from both gene
(CDS/gene features) and tRNA features in GenBank format files. Results are saved
in Excel format with separate sheets for gene introns and tRNA introns, featuring
publication-quality formatting.

Additionally, this module generates high-quality matplotlib figures:
  - Individual gene-structure figures (one per species / GenBank file)
  - A combined size-range figure comparing min-max intron (and optionally exon)
    sizes across all species

Key Features:
1. Extracts intron data from CDS/gene features
2. Extracts intron data from tRNA features
3. Validates intron lengths (1-15000 bp)
4. Uses Roman numerals for multiple introns
5. Groups output by species
6. Generates publication-ready Excel files with:
   - Proper formatting (headers, italics for genes, species grouping)
   - Separate sheets for genes and tRNAs
   - Abbreviation footnotes
   - Auto-adjusted column widths
7. Generates individual gene-structure figures (per species file) showing the
   SAME genes that appear in the Excel -- both use identical extraction logic
8. Generates combined intron/exon size-range comparison figures across all species

Author: Abdullah
Version: 3.1 (Module 11 - unified extraction + figure generation)
Date: February 2026

Dependencies:
    Python: biopython, pandas, openpyxl, matplotlib

Usage:
    python cgas_module11.py
    python cgas_module11.py -i genbank_files/
    python cgas_module11.py -i data/ -o results/

Output:
    Module11_Intron_Analysis/
        - intron_data_YYYYMMDD_HHMMSS.xlsx
          * Sheet 1: Gene Introns (CDS/gene features)
          * Sheet 2: tRNA Introns (tRNA features)
          * Sheet 3: Gene Intron Size Comparison (if multiple species)
          * Sheet 4: tRNA Intron Size Comparison (if multiple species)
        - figures/
          * <species>_gene_structure.png   (one per GenBank file)
          * combined_gene_intron_size_ranges.png
          * combined_trna_intron_size_ranges.png
          * combined_gene_exon_size_ranges.png  (if internal exon data available)

Notes:
    - Processes all GenBank files (.gb, .gbff, .genbank, .gbk) in input directory
    - Intron length must be 1-15000 bp
    - Handles multiple introns per gene with Roman numeral notation
    - Groups results by species for clarity
    - IR (Inverted Repeat) regions: Genes appearing twice (e.g., trnA-UGC, trnI-GAU)
      with identical intron sizes are automatically deduplicated in comparative analysis
    - Figures show exactly the same genes as the Excel sheets because both are
      produced from a single shared extraction pass per GenBank file
"""

import os
import sys
import math
import argparse
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Dict, Optional

try:
    from Bio import SeqIO
    from Bio.SeqFeature import CompoundLocation
except ImportError as e:
    print(f"Error: BioPython not installed: {e}")
    print("Please install required packages using:")
    print("pip install biopython")
    sys.exit(1)

try:
    import pandas as pd
    import numpy as np
except ImportError as e:
    print(f"Error: pandas/numpy not installed: {e}")
    print("Please install required packages using:")
    print("pip install pandas numpy")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError as e:
    print(f"Error: openpyxl not installed: {e}")
    print("Please install required packages using:")
    print("pip install openpyxl")
    sys.exit(1)

import json
import subprocess



# ============================================================================
# CONSTANTS
# ============================================================================

OUTPUT_FOLDER     = "Module11_Intron_Analysis"
FIGURES_FOLDER    = "figures"
MAX_INTRON_LENGTH = 15000
VALID_EXTENSIONS  = ('.gb', '.gbff', '.genbank', '.gbk')

# Figure colour palette
COLOUR_EXON   = '#4472C4'   # blue
COLOUR_INTRON = '#ED7D31'   # orange
FIG_DPI       = 300


# ============================================================================
# LOGGING CONFIGURATION
# ============================================================================

def setup_logging():
    """Configure logging with proper encoding for all platforms."""
    if sys.platform == 'win32':
        try:
            if hasattr(sys.stdout, 'reconfigure'):
                sys.stdout.reconfigure(encoding='utf-8')
        except (AttributeError, OSError):
            pass

    logging.basicConfig(
        level=logging.INFO,
        format='%(message)s',
        handlers=[logging.StreamHandler(sys.stdout)]
    )
    return logging.getLogger(__name__)


logger = setup_logging()


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def to_roman(num: int) -> str:
    """Convert integer to Roman numeral for intron numbering."""
    val  = [1000, 900, 500, 400, 100, 90, 50, 40, 10, 9, 5, 4, 1]
    syms = ['M', 'CM', 'D', 'CD', 'C', 'XC', 'L', 'XL', 'X', 'IX', 'V', 'IV', 'I']
    roman_num = ''
    for i in range(len(val)):
        count = int(num / val[i])
        if count:
            roman_num += syms[i] * count
            num -= val[i] * count
    return roman_num


def find_genbank_files(directory: str) -> List[str]:
    """Find all GenBank files in the specified directory."""
    genbank_files = []
    for file in os.listdir(directory):
        if file.endswith(VALID_EXTENSIONS):
            genbank_files.append(os.path.join(directory, file))
    return sorted(genbank_files)


def safe_filename(name: str) -> str:
    """Convert a species/gene name into a safe file-system name."""
    for ch in (' ', '/', '\\', ':', '*', '?', '"', '<', '>', '|'):
        name = name.replace(ch, '_')
    return name


# ============================================================================
# UNIFIED EXTRACTION  (single pass -> Excel rows + figure structures)
# ============================================================================
#
# KEY DESIGN DECISION:
#   Both the Excel output and the figure generation share the same extraction
#   function.  This guarantees the figure always shows exactly the same genes
#   as the spreadsheet -- no separate re-parsing with different logic.
#
# Excel-row format (unchanged from module 11 v2):
#   [filename, species, gene, label, start, end, length, ...]
#
# Figure-entry format (new, produced alongside):
#   {
#       'gene'   : str,
#       'exons'  : [(start, end), ...],          # genomic order
#       'introns': [(start, end, length), ...],
#       'strand' : '+' or '-',
#   }
#   One dict per feature with valid introns.  IR duplicates are preserved
#   as separate list entries so the figure shows every occurrence.

def extract_gene_introns_from_genbank(gb_file: str):
    """
    Single-pass extraction of CDS/gene introns.

    Identical intron filtering logic to module 11 v2.  Additionally captures
    exon coordinates and inter-exon gap types for figure drawing.

    gap_sequence
    ------------
    For each consecutive exon pair (i → i+1) this records whether the gap is:
      ('intron', ptr)  – a regular intron (within MAX_INTRON_LENGTH); ptr is
                         the index into valid_introns for that gap.
      ('trans', None)  – a trans-spliced junction (gap > MAX_INTRON_LENGTH or
                         /trans_splicing qualifier present); drawn as a visual
                         break, not as an intron bar.

    Trans-spliced genes (e.g. rps12) that have BOTH a trans-splice gap AND at
    least one regular intron are now included in both Excel and figure output.
    Previously they were silently dropped when any gap exceeded MAX_INTRON_LENGTH.

    Returns
    -------
    species      : str
    excel_rows   : List[List]
        [filename, species, gene, label, start, end, length, ...]
    fig_entries  : List[dict]
        One entry per feature that has ≥1 valid intron OR a trans-splice gap.
    """
    excel_rows  = []
    fig_entries = []
    filename    = os.path.splitext(os.path.basename(gb_file))[0]
    species     = 'Unknown species'

    try:
        for record in SeqIO.parse(gb_file, "genbank"):
            species = record.annotations.get('organism', 'Unknown species')

            for feature in record.features:
                if feature.type not in ('CDS', 'gene'):
                    continue

                gene = feature.qualifiers.get('gene', ['Unknown'])[0]

                if not isinstance(feature.location, CompoundLocation):
                    continue

                is_trans = 'trans_splicing' in feature.qualifiers

                # ── Determine biological exon order ──────────────────────────
                # For trans-spliced genes (e.g. rps12): use the join() annotation
                # order, reversed for complement strand, as the biological order
                # is defined by the annotation (not simple coordinate sorting).
                #
                # For all normal cis-spliced genes:
                #   + strand → sort exons by ascending coordinate (leftmost = E1)
                #   − strand → sort exons by descending coordinate (rightmost = E1)
                # This strictly follows transcription direction, independent of
                # how the join() parts happen to be ordered in the annotation.
                overall_strand = feature.location.strand   # -1, +1, or None
                join_parts     = list(feature.location.parts)   # join-order

                if is_trans:
                    # Trans-spliced (e.g. rps12): biologically, the LSC exon is always
                    # Exon 1 (5' end), regardless of strand or annotation join order.
                    # The IR exons follow in transcription order, which on the minus
                    # strand means descending genomic coordinate (highest coord = E2).
                    #
                    # Algorithm:
                    #   1. Find the "outlier" exon — the one far away from the others
                    #      (the LSC exon). Identified as the part whose distance to the
                    #      group median is greatest (i.e. the trans-splice partner).
                    #   2. That outlier = E1.
                    #   3. Remaining parts (IR exons) sorted descending on minus strand
                    #      (ascending on plus strand) to follow transcription direction.
                    if len(join_parts) >= 2:
                        starts = [int(p.start) for p in join_parts]
                        median_start = sorted(starts)[len(starts) // 2]
                        # The LSC exon is the one furthest from the median
                        outlier_idx = max(range(len(starts)),
                                          key=lambda i: abs(starts[i] - median_start))
                        lsc_part  = join_parts[outlier_idx]
                        ir_parts  = [p for i, p in enumerate(join_parts)
                                     if i != outlier_idx]
                        # Sort IR exons in transcription order
                        if overall_strand == -1:
                            ir_parts_sorted = sorted(ir_parts,
                                                     key=lambda p: p.start,
                                                     reverse=True)  # descending -> E2, E3
                        else:
                            ir_parts_sorted = sorted(ir_parts,
                                                     key=lambda p: p.start)  # ascending
                        bio_parts = [lsc_part] + ir_parts_sorted
                    else:
                        bio_parts = join_parts  # fallback for 2-exon trans-spliced
                else:
                    # Cis-spliced: sort strictly by transcription direction
                    if overall_strand == -1:
                        # Negative strand → transcription right-to-left
                        # Rightmost exon (highest start coord) = Exon 1
                        bio_parts = sorted(join_parts,
                                           key=lambda p: p.start,
                                           reverse=True)
                    else:
                        # Positive strand → transcription left-to-right
                        # Leftmost exon (lowest start coord) = Exon 1
                        bio_parts = sorted(join_parts,
                                           key=lambda p: p.start)

                # Map (start, end) → biological exon number (1-indexed)
                exon_number_map: Dict[tuple, int] = {}
                for bio_idx, part in enumerate(bio_parts, 1):
                    exon_number_map[(int(part.start), int(part.end))] = bio_idx

                # Sort by coordinate for visual left-to-right display
                exon_parts = sorted(feature.location.parts, key=lambda x: x.start)

                try:
                    strand = '+' if (overall_strand is not None and overall_strand >= 0) else '-'
                except (TypeError, AttributeError):
                    strand = '+'

                exons = [(int(p.start), int(p.end)) for p in exon_parts]

                # Build gap_sequence and valid_introns in a single pass
                valid_introns = []
                gap_sequence  = []   # one entry per consecutive exon pair

                for i in range(len(exon_parts) - 1):
                    intron_start  = int(exon_parts[i].end) + 1
                    intron_end    = int(exon_parts[i + 1].start)
                    intron_length = intron_end - intron_start + 1

                    if 0 < intron_length <= MAX_INTRON_LENGTH:
                        gap_sequence.append(('intron', len(valid_introns)))
                        valid_introns.append((intron_start, intron_end, intron_length))
                    else:
                        # Large gap → treat as trans-spliced junction
                        gap_sequence.append(('trans', None))

                # Include gene if it has regular introns OR trans-splice gaps
                has_regular = bool(valid_introns)
                has_trans   = any(g[0] == 'trans' for g in gap_sequence)

                if not has_regular and not has_trans:
                    continue

                # Excel row (only regular introns recorded, unchanged format)
                if has_regular:
                    num_introns = len(valid_introns)
                    intron_info = []
                    for idx, (i_s, i_e, i_l) in enumerate(valid_introns, 1):
                        label = "Intron" if num_introns == 1 else f"Intron {to_roman(idx)}"
                        intron_info.extend([label, i_s, i_e, i_l])
                    excel_rows.append([filename, species, gene] + intron_info)

                # Figure entry — only include if the feature has at least one regular
                # intron. Trans-splice-only entries (e.g. the rps12 E1|trans|E2 span
                # without any cis-intron) are excluded to avoid a spurious third row.
                if not has_regular:
                    continue

                fig_entries.append({
                    'gene'           : gene,
                    'exons'          : exons,
                    'introns'        : valid_introns,
                    'strand'         : strand,
                    'gap_sequence'   : gap_sequence,
                    'exon_number_map': exon_number_map,  # (start,end) -> bio exon number
                })

    except Exception as e:
        logger.warning(f"  Warning (CDS/gene): {os.path.basename(gb_file)}: {e}")

    return species, excel_rows, fig_entries


def extract_tRNA_introns_from_genbank(gb_file: str):
    """
    Single-pass extraction of tRNA introns.

    Identical intron filtering logic to module 11 v2.  Additionally captures
    exon coordinates for figure drawing.

    Returns
    -------
    species      : str
    excel_rows   : List[List]
    fig_entries  : List[dict]
    """
    excel_rows  = []
    fig_entries = []
    filename    = os.path.splitext(os.path.basename(gb_file))[0]
    species     = 'Unknown species'

    try:
        for record in SeqIO.parse(gb_file, "genbank"):
            species = record.annotations.get('organism', 'Unknown species')

            for feature in record.features:
                if feature.type != 'tRNA':
                    continue

                gene = feature.qualifiers.get(
                    'gene', feature.qualifiers.get('product', ['Unknown'])
                )[0]

                if not isinstance(feature.location, CompoundLocation):
                    continue

                overall_strand = feature.location.strand
                join_parts     = list(feature.location.parts)

                # tRNA genes are cis-spliced: sort by transcription direction.
                # + strand → ascending (leftmost = E1)
                # − strand → descending (rightmost = E1)
                if overall_strand == -1:
                    bio_parts = sorted(join_parts,
                                       key=lambda p: p.start,
                                       reverse=True)
                else:
                    bio_parts = sorted(join_parts,
                                       key=lambda p: p.start)

                exon_number_map: Dict[tuple, int] = {}
                for bio_idx, part in enumerate(bio_parts, 1):
                    exon_number_map[(int(part.start), int(part.end))] = bio_idx

                exon_parts = sorted(feature.location.parts, key=lambda x: x.start)

                try:
                    strand = '+' if (overall_strand is not None and overall_strand >= 0) else '-'
                except (TypeError, AttributeError):
                    strand = '+'

                exons = [(int(p.start), int(p.end)) for p in exon_parts]

                # tRNA intron formula from module 11 v2 (end+1 / start-1)
                # Also build gap_sequence for consistency with CDS extraction
                valid_introns = []
                gap_sequence  = []

                for i in range(len(exon_parts) - 1):
                    intron_start  = int(exon_parts[i].end) + 1
                    intron_end    = int(exon_parts[i + 1].start) - 1
                    intron_length = int(intron_end - intron_start + 1)

                    if 0 < intron_length <= MAX_INTRON_LENGTH:
                        gap_sequence.append(('intron', len(valid_introns)))
                        valid_introns.append((int(intron_start),
                                              int(intron_end),
                                              intron_length))
                    else:
                        gap_sequence.append(('trans', None))

                if not valid_introns:
                    continue

                # Excel row
                num_introns = len(valid_introns)
                intron_info = []
                for idx, (i_s, i_e, i_l) in enumerate(valid_introns, 1):
                    label = "Intron" if num_introns == 1 else f"Intron {to_roman(idx)}"
                    intron_info.extend([label, i_s, i_e, i_l])

                excel_rows.append([filename, species, gene] + intron_info)

                # Figure entry
                fig_entries.append({
                    'gene'           : gene,
                    'exons'          : exons,
                    'introns'        : valid_introns,
                    'strand'         : strand,
                    'gap_sequence'   : gap_sequence,
                    'exon_number_map': exon_number_map,
                })

    except Exception as e:
        logger.warning(f"  Warning (tRNA): {os.path.basename(gb_file)}: {e}")

    return species, excel_rows, fig_entries


# ============================================================================
# R FIGURE GENERATION — automatic (follows module 13 workflow pattern)
# ============================================================================
#
# The Python pipeline exports JSON data, generates an R script as a string,
# writes it to disk, and executes it via subprocess — fully automatic.


def _serialise_entry(entry: dict) -> dict:
    """Convert a fig_entry dict to JSON-safe form."""
    return {
        'gene':           entry['gene'],
        'exons':          entry['exons'],
        'introns':        [list(t) for t in entry['introns']],
        'strand':         entry['strand'],
        'gap_sequence':   [
            [g[0], g[1]] for g in entry.get('gap_sequence', [])
        ],
        'exon_number_map': {
            f"{k[0]}_{k[1]}": v
            for k, v in entry.get('exon_number_map', {}).items()
        },
        'intron_labels':  entry.get('intron_labels', None),
    }


def export_species_json(species: str, gene_fig_entries: list,
                         trna_fig_entries: list, output_path: str) -> bool:
    """Export one species' figure data as JSON for the R script."""
    if not gene_fig_entries and not trna_fig_entries:
        logger.warning(f"  No gene structures to export for {species}")
        return False
    payload = {
        'species':       species,
        'gene_entries':  [_serialise_entry(e) for e in gene_fig_entries],
        'trna_entries':  [_serialise_entry(e) for e in trna_fig_entries],
    }
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    with open(output_path, 'w') as fh:
        json.dump(payload, fh, indent=2)
    logger.info(f"  ✓ JSON exported: {os.path.basename(output_path)}")
    return True


def export_combined_json(all_gene_figs: list, all_trna_figs: list,
                          output_path: str) -> bool:
    """Export combined (all-species) figure data as JSON for the R script."""
    if not all_gene_figs and not all_trna_figs:
        logger.warning("  No data for combined figure export.")
        return False
    payload = {
        'species':       'Combined',
        'gene_entries':  [_serialise_entry(e) for e in all_gene_figs],
        'trna_entries':  [_serialise_entry(e) for e in all_trna_figs],
    }
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    with open(output_path, 'w') as fh:
        json.dump(payload, fh, indent=2)
    logger.info(f"  ✓ Combined JSON exported: {os.path.basename(output_path)}")
    return True


# ── R package management (module 13 pattern) ─────────────────────────────────

def check_r_packages() -> tuple:
    """Check if required R packages are installed."""
    required = ['ggplot2', 'jsonlite', 'gridExtra']
    try:
        r_code = 'packages <- c({})\nmissing <- packages[!packages %in% installed.packages()[,"Package"]]\ncat(paste(missing, collapse=","))'.format(
            ", ".join([f'"{p}"' for p in required]))
        result = subprocess.run(['R', '--vanilla', '--slave', '-e', r_code],
                              capture_output=True, text=True, timeout=10)
        missing = [p.strip() for p in result.stdout.strip().split(',') if p.strip()]
        return (len(missing) == 0, missing)
    except Exception:
        return (False, required)


def install_r_packages(packages: list) -> bool:
    """Attempt to install missing R packages."""
    print(f"\n  → Installing R packages: {', '.join(packages)}")
    try:
        r_code = 'packages <- c({})\ninstall.packages(packages, repos="https://cloud.r-project.org", quiet=TRUE)'.format(
            ", ".join([f'"{p}"' for p in packages]))
        result = subprocess.run(['R', '--vanilla', '--slave', '-e', r_code],
                              capture_output=True, text=True, timeout=300)
        return result.returncode == 0
    except Exception as e:
        print(f"  ✗ Installation failed: {e}")
        return False


# ── Embedded R script ─────────────────────────────────────────────────────────

def create_gene_structure_r_script() -> str:
    """Return the full R script content for gene-structure figures."""

    # Use raw string to avoid Python escape issues with R code
    return r'''
# =============================================================================
# CGAS Module 11 - Gene Structure Figure Generator (R)
# Auto-generated by cgas_module11.py
# =============================================================================

suppressPackageStartupMessages({
  library(ggplot2)
  library(jsonlite)
  library(grid)
  library(gridExtra)
})

cat("\n========================================\n")
cat("MODULE 11: GENE STRUCTURE FIGURES (R)\n")
cat("========================================\n\n")

# Constants
COLOUR_EXON       <- "#4472C4"
COLOUR_INTRON     <- "#ED7D31"
COLOUR_TRANS      <- "#D8D8D8"
COLOUR_TRANS_EDGE <- "#888888"
COLOUR_BASELINE   <- "#BBBBBB"

EXON_WEIGHT       <- 2.0
INTRON_SCALE      <- 2.2
MIN_INTRON_WEIGHT <- 1.2
TRANS_WEIGHT      <- 1.6

LEFT  <- 0.13;  RIGHT <- 0.92;  WIDTH <- RIGHT - LEFT
BAR_H <- 0.28;  INT_H <- 0.20;  TRS_H <- 0.14

FONT_FAMILY <- "sans"

`%||%` <- function(a, b) if (is.null(a)) b else a

intron_wt <- function(len) max(MIN_INTRON_WEIGHT, log10(max(len, 1)) * INTRON_SCALE)

# =============================================================================
# CORE DRAWING FUNCTION
# =============================================================================
draw_panel <- function(fig_entries, panel_title, show_coords = FALSE) {
  # Sort entries alphabetically A-Z (top to bottom in figure)
  gene_names <- sapply(fig_entries, function(e) e$gene)
  fig_entries <- fig_entries[order(tolower(gene_names), decreasing = TRUE)]

  n <- length(fig_entries)
  top_pad <- if (show_coords) -1.5 else -0.55
  bot_pad <- if (show_coords) n - 0.3 else n - 0.55
  p <- ggplot() +
    coord_cartesian(xlim = c(0, 1), ylim = c(bot_pad, top_pad), expand = FALSE) +
    theme_void(base_family = FONT_FAMILY) +
    theme(plot.margin = margin(2, 10, 2, 5, "pt")) +
    labs(title = NULL)
  if (n == 0) return(p + annotate("text", x=0.5, y=0, label="No genes found", size=4.5))

  rects <- data.frame(xmin=numeric(), xmax=numeric(), ymin=numeric(), ymax=numeric(),
                       fill=character(), edge=character(), type=character(), stringsAsFactors=FALSE)
  texts <- data.frame(x=numeric(), y=numeric(), label=character(), colour=character(),
                       size=numeric(), fontface=character(), stringsAsFactors=FALSE)
  segs  <- data.frame(x=numeric(), xend=numeric(), y=numeric(), yend=numeric(), stringsAsFactors=FALSE)
  glabs <- data.frame(x=numeric(), y=numeric(), label=character(), stringsAsFactors=FALSE)

  for (ri in seq_along(fig_entries)) {
    e <- fig_entries[[ri]];  y <- ri - 1
    exons <- e$exons;  introns <- e$introns;  strand <- e$strand %||% "+"
    gs <- e$gap_sequence %||% list();  enm <- e$exon_number_map
    ne <- length(exons);  ni <- length(introns)
    if (ne == 0) next

    # Build render sequence
    rs <- list()
    if (length(gs) > 0) {
      for (ei in seq_len(ne)) {
        rs[[length(rs)+1]] <- list(k="exon", i=ei)
        if (ei <= length(gs)) {
          g <- gs[[ei]]
          gt <- if(is.list(g)) g[[1]] else g[1]
          gp <- if(is.list(g)) g[[2]] else g[2]
          if (is.null(gt)) gt <- "intron"
          if (gt == "trans") rs[[length(rs)+1]] <- list(k="trans", i=NA)
          else rs[[length(rs)+1]] <- list(k="intron", i=as.numeric(gp)+1)
        }
      }
    } else {
      for (ei in seq_len(ne)) {
        rs[[length(rs)+1]] <- list(k="exon", i=ei)
        if (ei <= ni) rs[[length(rs)+1]] <- list(k="intron", i=ei)
      }
    }

    # Weights & widths
    wts <- sapply(rs, function(it) {
      if (it$k=="exon") EXON_WEIGHT
      else if (it$k=="intron") { p<-it$i; if(!is.na(p)&&p<=ni) intron_wt(introns[[p]][[3]]) else MIN_INTRON_WEIGHT }
      else TRANS_WEIGHT
    })
    tw <- sum(wts); if(tw==0) tw<-1; ws <- wts/tw*WIDTH

    segs <- rbind(segs, data.frame(x=LEFT, xend=RIGHT, y=y, yend=y))
    xc <- LEFT; ec <- 0

    for (idx in seq_along(rs)) {
      it <- rs[[idx]]; w <- ws[idx]; xs <- xc; xe <- xs + w

      if (it$k == "exon") {
        ec <- ec + 1
        rects <- rbind(rects, data.frame(xmin=xs, xmax=xe, ymin=y-BAR_H, ymax=y+BAR_H,
                                          fill=COLOUR_EXON, edge="black", type="exon", stringsAsFactors=FALSE))
        enum <- ec
        if (!is.null(enm) && ec <= ne) {
          key <- paste0(exons[[ec]][[1]], "_", exons[[ec]][[2]])
          if (key %in% names(enm)) enum <- enm[[key]]
        } else if (strand == "-") enum <- ne - ec + 1

        texts <- rbind(texts, data.frame(x=(xs+xe)/2, y=y, label=paste0("E",enum),
                                          colour="white", size=4.5, fontface="bold", stringsAsFactors=FALSE))
        if (show_coords && ec <= ne) {
          texts <- rbind(texts, data.frame(x=xs, y=y+BAR_H+0.15, label=as.character(exons[[ec]][[1]]),
                                            colour="#333333", size=4.2, fontface="plain", stringsAsFactors=FALSE))
          texts <- rbind(texts, data.frame(x=xe, y=y+BAR_H+0.15, label=as.character(exons[[ec]][[2]]),
                                            colour="#333333", size=4.2, fontface="plain", stringsAsFactors=FALSE))
        }
      } else if (it$k == "intron") {
        ptr <- it$i; if(is.na(ptr)||ptr>ni) { xc<-xe; next }
        ilen <- introns[[ptr]][[3]]
        rects <- rbind(rects, data.frame(xmin=xs, xmax=xe, ymin=y-INT_H, ymax=y+INT_H,
                                          fill=COLOUR_INTRON, edge="black", type="intron", stringsAsFactors=FALSE))
        bt <- if (!is.null(e$intron_labels) && ptr <= length(e$intron_labels)) e$intron_labels[[ptr]] else paste0(ilen," bp")
        fs <- min(5.5, max(4.2, w*50))
        texts <- rbind(texts, data.frame(x=(xs+xe)/2, y=y, label=bt,
                                          colour="white", size=fs, fontface="bold", stringsAsFactors=FALSE))
      } else {
        rects <- rbind(rects, data.frame(xmin=xs, xmax=xe, ymin=y-TRS_H, ymax=y+TRS_H,
                                          fill=COLOUR_TRANS, edge=COLOUR_TRANS_EDGE, type="trans", stringsAsFactors=FALSE))
        texts <- rbind(texts, data.frame(x=(xs+xe)/2, y=y, label="trans-spliced",
                                          colour="#555555", size=3.8, fontface="italic", stringsAsFactors=FALSE))
      }
      xc <- xe
    }
    glabs <- rbind(glabs, data.frame(x=LEFT-0.01, y=y, label=e$gene, stringsAsFactors=FALSE))
  }

  # Assemble
  if (nrow(segs)>0) p <- p + geom_segment(data=segs, aes(x=x,xend=xend,y=y,yend=yend), colour=COLOUR_BASELINE, linewidth=0.3)
  tr <- rects[rects$type=="trans",]; ir <- rects[rects$type=="intron",]; er <- rects[rects$type=="exon",]
  if (nrow(tr)>0) p <- p + geom_rect(data=tr, aes(xmin=xmin,xmax=xmax,ymin=ymin,ymax=ymax), fill=COLOUR_TRANS, colour=COLOUR_TRANS_EDGE, linewidth=0.3, linetype="longdash")
  if (nrow(ir)>0) p <- p + geom_rect(data=ir, aes(xmin=xmin,xmax=xmax,ymin=ymin,ymax=ymax), fill=COLOUR_INTRON, colour="black", linewidth=0.35)
  if (nrow(er)>0) p <- p + geom_rect(data=er, aes(xmin=xmin,xmax=xmax,ymin=ymin,ymax=ymax), fill=COLOUR_EXON, colour="black", linewidth=0.35)
  if (nrow(texts)>0) p <- p + geom_text(data=texts, aes(x=x,y=y,label=label), colour=texts$colour, size=texts$size, fontface=texts$fontface, family=FONT_FAMILY)
  if (nrow(glabs)>0) p <- p + geom_text(data=glabs, aes(x=x,y=y,label=label), hjust=1, size=5.5, fontface="italic", colour="black", family=FONT_FAMILY)
  return(p)
}

# Legend
build_legend <- function() {
  ld <- data.frame(xmin=c(0,0.28,0.62), xmax=c(0.08,0.36,0.70), ymin=c(.2,.2,.2), ymax=c(.8,.8,.8),
                    fill=c(COLOUR_EXON,COLOUR_INTRON,COLOUR_TRANS), edge=c("black","black",COLOUR_TRANS_EDGE),
                    label=c("Exon (E1, E2 \u2026)","Intron (size in bp)","Trans-spliced junction"),
                    lx=c(0.095,0.375,0.715))
  ggplot() + coord_cartesian(xlim=c(-0.05,1.1), ylim=c(-0.4,1.4), expand=FALSE) + theme_void(base_family=FONT_FAMILY) +
    geom_rect(data=ld, aes(xmin=xmin,xmax=xmax,ymin=ymin,ymax=ymax), fill=ld$fill, colour=ld$edge, linewidth=0.5) +
    geom_text(data=ld, aes(x=lx,y=0.5,label=label), hjust=0, size=5.0, family=FONT_FAMILY) +
    annotate("text", x=0, y=-0.25, label="Diagrammatic \u2014 bar widths are not to genomic scale",
             hjust=0, size=3.8, colour="#555555", fontface="italic", family=FONT_FAMILY)
}

# JSON reader
read_fig <- function(path) {
  raw <- fromJSON(path, simplifyVector=FALSE)
  parse_e <- function(lst) lapply(lst, function(e) {
    ex <- lapply(e$exons, function(x) list(x[[1]],x[[2]]))
    it <- lapply(e$introns, function(x) list(x[[1]],x[[2]],x[[3]]))
    gs <- lapply(e$gap_sequence %||% list(), function(g) list(g[[1]], if(is.null(g[[2]])) NA else g[[2]]))
    nm <- if(!is.null(e$exon_number_map)){m<-list(); for(k in names(e$exon_number_map)) m[[k]]<-e$exon_number_map[[k]]; m} else NULL
    list(gene=e$gene, exons=ex, introns=it, strand=e$strand%||%"+", gap_sequence=gs, exon_number_map=nm, intron_labels=e$intron_labels)
  })
  list(species=raw$species%||%"Unknown", gene_entries=parse_e(raw$gene_entries%||%list()), trna_entries=parse_e(raw$trna_entries%||%list()))
}

# Combined entry builder
build_combined <- function(all_e) {
  gd <- list()
  for (e in all_e) {
    g <- e$gene
    if (is.null(gd[[g]])) gd[[g]] <- list(exons=e$exons, gap_sequence=e$gap_sequence, exon_number_map=e$exon_number_map, isz=list())
    for (p in seq_along(e$introns)) {
      k <- as.character(p)
      if (is.null(gd[[g]]$isz[[k]])) gd[[g]]$isz[[k]] <- c()
      gd[[g]]$isz[[k]] <- unique(c(gd[[g]]$isz[[k]], e$introns[[p]][[3]]))
    }
  }
  res <- list()
  for (g in sort(names(gd))) {
    d <- gd[[g]]; ips <- c()
    if (length(d$gap_sequence)>0) for (gap in d$gap_sequence) {
      gt <- if(is.list(gap)) gap[[1]] else gap[1]; gp <- if(is.list(gap)) gap[[2]] else gap[2]
      if (!is.null(gt) && gt=="intron" && !is.na(gp)) ips <- c(ips, as.numeric(gp)+1)
    }
    nip <- if(length(ips)>0) max(ips) else 0
    its <- list(); ils <- list()
    for (p in seq_len(nip)) {
      sz <- sort(d$isz[[as.character(p)]] %||% 500); mn <- min(sz); mx <- max(sz)
      its[[p]] <- list(0,0,round(mean(sz)))
      ils[[p]] <- if(mn==mx) paste0(mn," bp") else paste0(mn,"\u2013",mx," bp")
    }
    res[[length(res)+1]] <- list(gene=g, exons=d$exons, introns=its, strand="+",
                                  gap_sequence=d$gap_sequence, exon_number_map=d$exon_number_map, intron_labels=ils)
  }
  return(res)
}

# Save a single panel as its own figure (fixed width for manual combining)
save_single <- function(panel, height, fig_w, out_path) {
  fh <- min(max(height, 1.5), 40)
  ggsave(out_path, plot=panel, width=fig_w, height=fh, dpi=300, bg="white", limitsize=FALSE)
  pdf_p <- sub("\\.png$", ".pdf", out_path)
  ggsave(pdf_p, plot=panel, width=fig_w, height=fh, device=cairo_pdf, bg="white", limitsize=FALSE)
  cat(paste("  \u2713", basename(out_path), "\n"))
  cat(paste("  \u2713", basename(pdf_p), "\n"))
}

# =============================================================================
# MAIN
# =============================================================================
args <- commandArgs(trailingOnly = TRUE)
input_dir <- if (length(args) >= 1) args[1] else "."

cat(paste("  Reading JSON from:", input_dir, "\n\n"))
jfiles <- list.files(input_dir, pattern="\\.json$", full.names=TRUE)
if (length(jfiles)==0) { cat("  No JSON found.\n"); quit(status=1) }

# Fixed width for all figures (same size so they can be combined manually)
FIG_W <- 13

all_gf <- list(); all_tf <- list()
sf <- jfiles[!grepl("combined", basename(jfiles))]
for (jf in sf) {
  cat(paste("  Processing:", basename(jf), "\n"))
  d <- read_fig(jf)
  all_gf <- c(all_gf, d$gene_entries); all_tf <- c(all_tf, d$trna_entries)
  sn <- gsub("[^A-Za-z0-9._-]", "_", d$species)

  # Protein-coding genes — separate figure
  if (length(d$gene_entries) > 0) {
    p <- draw_panel(d$gene_entries, "Protein-coding Genes", TRUE)
    h <- length(d$gene_entries) * 1.2 + 0.8
    save_single(p, h, FIG_W, file.path(input_dir, paste0(sn, "_protein_coding_genes.png")))
  }

  # tRNA genes — separate figure
  if (length(d$trna_entries) > 0) {
    p <- draw_panel(d$trna_entries, "tRNA Genes", TRUE)
    h <- length(d$trna_entries) * 1.2 + 0.8
    save_single(p, h, FIG_W, file.path(input_dir, paste0(sn, "_tRNA_genes.png")))
  }
}

# Combined figures (across all species) — also separate
if (length(all_gf) > 0) {
  cat("\n  Building combined protein-coding figure...\n")
  ge <- build_combined(all_gf)
  p <- draw_panel(ge, "Protein-coding Genes", FALSE)
  h <- length(ge) * 1.0 + 0.8
  save_single(p, h, FIG_W, file.path(input_dir, "combined_protein_coding_genes.png"))
}

if (length(all_tf) > 0) {
  cat("\n  Building combined tRNA figure...\n")
  te <- build_combined(all_tf)
  p <- draw_panel(te, "tRNA Genes", FALSE)
  h <- length(te) * 1.0 + 0.8
  save_single(p, h, FIG_W, file.path(input_dir, "combined_tRNA_genes.png"))
}

cat("\n========================================\n")
cat("FIGURES GENERATED SUCCESSFULLY\n")
cat("========================================\n")
'''


# ── Main orchestrator ─────────────────────────────────────────────────────────

def generate_r_visualizations(figures_dir: str) -> None:
    """
    Generate R-based gene-structure figures (module 13 pattern).

    1. Check R availability
    2. Check/install required R packages
    3. Write R script to figures_dir
    4. Execute R script via subprocess
    """

    print(f"\n{'='*70}")
    print("GENERATING R FIGURES")
    print(f"{'='*70}")

    # Check R availability
    try:
        result = subprocess.run(['Rscript', '--version'],
                              capture_output=True, text=True, timeout=10)
        if result.returncode != 0:
            print("  ⚠ R not found. Skipping figure generation.")
            print("  Install R to enable automatic figure generation.")
            return
    except (subprocess.TimeoutExpired, FileNotFoundError):
        print("  ⚠ R/Rscript not found. Skipping figure generation.")
        print("  Install R from: https://www.r-project.org/")
        return

    print("  ✓ R is available")

    # Check R packages
    packages_ok, missing = check_r_packages()
    if not packages_ok:
        print(f"\n  ⚠ Missing R packages: {', '.join(missing)}")
        print("  → Attempting automatic installation...")
        if install_r_packages(missing):
            print("  ✓ R packages installed successfully")
        else:
            print("  ✗ Could not install R packages automatically")
            print("  → Please install manually in R:")
            print(f"     install.packages(c({', '.join([repr(p) for p in missing])}))")
            return
    else:
        print("  ✓ All required R packages are installed")

    # Create R script
    r_script_content = create_gene_structure_r_script()
    r_script_file    = os.path.join(figures_dir, "generate_gene_structure_plot.R")
    with open(r_script_file, 'w') as f:
        f.write(r_script_content)

    print(f"  ✓ R script created: {os.path.basename(r_script_file)}")

    # Execute R script
    print(f"\n  Executing R script...")
    try:
        result = subprocess.run(
            ['Rscript', os.path.basename(r_script_file), '.'],
            capture_output=True, text=True,
            timeout=300,
            cwd=figures_dir
        )

        if result.stdout:
            print(result.stdout)

        if result.returncode == 0:
            print(f"\n  ✓ R visualization completed successfully")

            # Verify output files
            import glob
            pdf_files = glob.glob(os.path.join(figures_dir, "*.pdf"))
            png_files = glob.glob(os.path.join(figures_dir, "*.png"))
            if pdf_files or png_files:
                print(f"\n  Generated figures:")
                for f_path in sorted(pdf_files + png_files):
                    size = os.path.getsize(f_path)
                    print(f"    ✓ {os.path.basename(f_path)} ({size:,} bytes)")
        else:
            print(f"  ✗ R script execution failed (return code: {result.returncode})")
            if result.stderr:
                print(f"\n  Error output:")
                print(result.stderr[:500])

    except subprocess.TimeoutExpired:
        print(f"  ✗ R script timed out (>300s)")
    except Exception as e:
        print(f"  ✗ Error executing R script: {e}")


# ============================================================================
# COMPARATIVE ANALYSIS FUNCTIONS  (unchanged from module 11 v2)
# ============================================================================

def analyze_intron_size_ranges(data: List[List], is_trna: bool = False) -> Dict:
    """
    Analyze intron size ranges across multiple species for each gene.

    Uses a set of (species, length) tuples to deduplicate IR regions --
    identical (species, size) pairs are counted only once.

    Parameters
    ----------
    data    : Excel rows [filename, species, gene, label, start, end, length, ...]
    is_trna : Whether this is tRNA data (informational only)

    Returns
    -------
    Dict mapping (gene, intron_label) -> statistics dict
    """
    intron_data: Dict = {}

    for record in data:
        filename, species, gene = record[0], record[1], record[2]
        intron_info = record[3:]

        for i in range(0, len(intron_info), 4):
            if i + 3 < len(intron_info):
                intron_label  = intron_info[i]
                intron_length = intron_info[i + 3]

                key = (gene, intron_label)
                if key not in intron_data:
                    intron_data[key] = set()
                intron_data[key].add((species, intron_length))

    results = {}
    for (gene, intron_label), size_set in intron_data.items():
        if not size_set:
            continue
        size_list     = list(size_set)
        lengths       = [l for _, l in size_list]
        species_names = [s for s, _ in size_list]

        min_size  = min(lengths)
        max_size  = max(lengths)
        mean_size = sum(lengths) / len(lengths)

        results[(gene, intron_label)] = {
            'min_size'     : min_size,
            'max_size'     : max_size,
            'difference'   : max_size - min_size,
            'mean_size'    : round(mean_size, 2),
            'species_count': len(set(species_names)),
            'min_species'  : species_names[lengths.index(min_size)],
            'max_species'  : species_names[lengths.index(max_size)],
            'all_sizes'    : lengths,
        }

    return results


def analyze_exon_size_ranges(data: List[List]) -> Dict:
    """
    Derive internal exon size ranges from intron flank positions.

    Only internal exons (between two consecutive introns) are sized;
    terminal exons cannot be reliably derived from intron records alone.

    Returns Dict with same shape as analyze_intron_size_ranges.
    """
    gene_intron_pos: Dict = {}

    for record in data:
        filename, species, gene = record[0], record[1], record[2]
        intron_info = record[3:]

        positions = []
        for i in range(0, len(intron_info), 4):
            if i + 3 < len(intron_info):
                positions.append((intron_info[i + 1], intron_info[i + 2]))

        key = (gene, species)
        if key not in gene_intron_pos or len(positions) > len(gene_intron_pos[key]):
            gene_intron_pos[key] = positions

    exon_data: Dict = {}
    for (gene, species), intron_positions in gene_intron_pos.items():
        if len(intron_positions) < 2:
            continue
        sorted_introns = sorted(intron_positions, key=lambda x: x[0])
        for j in range(len(sorted_introns) - 1):
            exon_start  = sorted_introns[j][1]
            exon_end    = sorted_introns[j + 1][0]
            exon_length = exon_end - exon_start
            if exon_length <= 0:
                continue
            exon_label = f'Exon {to_roman(j + 2)}'
            key = (gene, exon_label)
            if key not in exon_data:
                exon_data[key] = set()
            exon_data[key].add((species, exon_length))

    results = {}
    for (gene, exon_label), size_set in exon_data.items():
        if not size_set:
            continue
        size_list     = list(size_set)
        lengths       = [l for _, l in size_list]
        species_names = [s for s, _ in size_list]
        min_size  = min(lengths)
        max_size  = max(lengths)
        mean_size = sum(lengths) / len(lengths)
        results[(gene, exon_label)] = {
            'min_size'     : min_size,
            'max_size'     : max_size,
            'difference'   : max_size - min_size,
            'mean_size'    : round(mean_size, 2),
            'species_count': len(set(species_names)),
            'min_species'  : species_names[lengths.index(min_size)],
            'max_species'  : species_names[lengths.index(max_size)],
            'all_sizes'    : lengths,
        }

    return results


# ============================================================================
# EXCEL FORMATTING AND OUTPUT  (unchanged from module 11 v2)
# ============================================================================

def add_abbreviation_footnotes(worksheet, start_row: int) -> int:
    """Add abbreviation footnotes to Excel worksheet."""
    footnotes = [
        "Abbreviations:",
        "tRNA: Transfer RNA",
        "CDS: Coding Sequence",
        "bp: Base pairs"
    ]
    row = start_row
    for footnote in footnotes:
        cell = worksheet.cell(row, 1, footnote)
        if footnote == "Abbreviations:":
            cell.font = Font(bold=True, size=10)
        else:
            cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        row += 1
    return row


def create_comparison_sheet(workbook, sheet_name: str,
                             comparison_data: Dict, is_trna: bool = False):
    """Create a comparison sheet showing intron size ranges across species."""
    ws = workbook.create_sheet(sheet_name)

    header_font      = Font(bold=True, size=11, color='FFFFFF')
    header_fill      = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    gene_font        = Font(italic=True, size=10)
    species_font     = Font(italic=True, size=10)
    regular_font     = Font(size=10)
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment   = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        'Gene' if not is_trna else 'tRNA Gene',
        'Intron', 'Minimum Size (bp)', 'Species (Min)',
        'Maximum Size (bp)', 'Species (Max)',
        'Difference (bp)', 'Mean Size (bp)', 'No. of Species'
    ]

    for col, header in enumerate(headers, 1):
        cell           = ws.cell(1, col, header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_alignment
        cell.border    = thin_border

    sorted_keys = sorted(comparison_data.keys(), key=lambda x: (x[0], x[1]))
    row = 2
    for (gene, intron_label) in sorted_keys:
        stats    = comparison_data[(gene, intron_label)]
        data_row = [
            (gene,                   gene_font,    left_alignment),
            (intron_label,           regular_font, center_alignment),
            (stats['min_size'],      regular_font, center_alignment),
            (stats['min_species'],   species_font, center_alignment),
            (stats['max_size'],      regular_font, center_alignment),
            (stats['max_species'],   species_font, center_alignment),
            (stats['difference'],    regular_font, center_alignment),
            (stats['mean_size'],     regular_font, center_alignment),
            (stats['species_count'], regular_font, center_alignment),
        ]
        for col_idx, (val, fnt, aln) in enumerate(data_row, 1):
            cell           = ws.cell(row, col_idx, val)
            cell.font      = fnt
            cell.alignment = aln
            cell.border    = thin_border
        row += 1

    for column in ws.columns:
        max_length    = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 3, 50)

    note_row = row + 2
    ws.cell(note_row, 1, "Note:").font = Font(bold=True, size=10)
    note_row += 1
    note_text = (
        "This sheet compares intron sizes across multiple species for each gene. "
        "The difference column shows variation in intron length among species. "
        "Genes in Inverted Repeat (IR) regions are counted only once per species, "
        "even though they appear twice in the genome."
    )
    note_cell = ws.cell(note_row, 1, note_text)
    note_cell.font = Font(size=9)
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=9)

    logger.info(f"  Comparison sheet created: {sheet_name}")


def create_combined_excel_output(gene_data: List[List],
                                 trna_data: List[List],
                                 output_excel: str) -> None:
    """
    Create publication-quality Excel workbook.

    Sheets:
      1. Gene Introns
      2. tRNA Introns
      3. Gene Intron Size Comparison   (if multiple species)
      4. tRNA Intron Size Comparison   (if multiple species)
    """
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    header_font       = Font(bold=True, size=11)
    header_alignment  = Alignment(horizontal='center', vertical='center')
    species_font      = Font(bold=True, italic=True, size=11)
    species_alignment = Alignment(horizontal='center', vertical='center')
    species_fill      = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    italic_font       = Font(italic=True, size=10)
    regular_font      = Font(size=10)
    left_alignment    = Alignment(horizontal='left', vertical='center')
    center_alignment  = Alignment(horizontal='center', vertical='center')

    def write_intron_sheet(ws, data, gene_col_name):
        max_introns = max((len(row) - 3) // 4 for row in data)
        headers = [gene_col_name]
        for i in range(1, max_introns + 1):
            if max_introns == 1:
                headers += ["Intron", "Start", "End", "Length (bp)"]
            else:
                headers += [f"Intron {to_roman(i)}", "Start", "End", "Length (bp)"]

        for col, header in enumerate(headers, 1):
            cell           = ws.cell(1, col, header)
            cell.font      = header_font
            cell.alignment = header_alignment

        row = 2
        data_by_species: Dict = {}
        for record in data:
            sp       = record[1]
            gene     = record[2]
            intron_d = record[3:]
            data_by_species.setdefault(sp, []).append([gene] + intron_d)

        for sp in sorted(data_by_species):
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=len(headers))
            sc           = ws.cell(row, 1, sp)
            sc.font      = species_font
            sc.alignment = species_alignment
            sc.fill      = species_fill
            row += 1

            for rec in data_by_species[sp]:
                while len(rec) < len(headers):
                    rec.append("")
                ws.cell(row, 1, rec[0]).font      = italic_font
                ws.cell(row, 1, rec[0]).alignment = left_alignment
                for col_idx, value in enumerate(rec[1:], 2):
                    cell           = ws.cell(row, col_idx, value)
                    cell.font      = regular_font
                    cell.alignment = center_alignment
                row += 1

        for column in ws.columns:
            max_length    = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 3, 50)

        add_abbreviation_footnotes(ws, row + 2)

    if gene_data:
        write_intron_sheet(wb.create_sheet('Gene Introns'), gene_data, 'Gene')

    if trna_data:
        write_intron_sheet(wb.create_sheet('tRNA Introns'), trna_data, 'tRNA Gene')

    def unique_species(d):
        return {rec[1] for rec in d if len(rec) >= 2}

    if gene_data and len(unique_species(gene_data)) > 1:
        logger.info("  Creating gene intron comparison sheet...")
        gc = analyze_intron_size_ranges(gene_data, is_trna=False)
        if gc:
            create_comparison_sheet(wb, 'Gene Intron Size Comparison', gc, False)

    if trna_data and len(unique_species(trna_data)) > 1:
        logger.info("  Creating tRNA intron comparison sheet...")
        tc = analyze_intron_size_ranges(trna_data, is_trna=True)
        if tc:
            create_comparison_sheet(wb, 'tRNA Intron Size Comparison', tc, True)

    wb.save(output_excel)
    logger.info(f"  ✓ Excel file created: {os.path.basename(output_excel)}")


# ============================================================================
# MAIN PIPELINE
# ============================================================================

def process_all_genbank_files(input_dir: str, output_dir: str) -> Tuple[int, int]:
    """
    Process all GenBank files in the specified directory.

    For each file -- one pass only:
      1. extract_gene_introns_from_genbank  -> Excel rows + figure entries
      2. extract_tRNA_introns_from_genbank  -> Excel rows + figure entries
      3. Plot individual gene-structure PNG from the same figure entries

    After all files:
      4. Write combined Excel
      5. Write combined intron size-range PNGs
      6. Write combined exon size-range PNG (if available)

    Returns (gene_intron_count, tRNA_intron_count).
    """
    all_gene_excel: List[List] = []
    all_trna_excel: List[List] = []
    all_gene_figs:  List[dict] = []   # fig_entries from every file (for combined figure)
    all_trna_figs:  List[dict] = []

    gb_files = find_genbank_files(input_dir)
    if not gb_files:
        raise FileNotFoundError(
            f"No GenBank files found in {input_dir}\n"
            f"Expected extensions: {', '.join(VALID_EXTENSIONS)}"
        )

    logger.info(f"Found {len(gb_files)} GenBank file(s):")
    for gb_file in gb_files:
        logger.info(f"  - {os.path.basename(gb_file)}")

    figures_dir = os.path.join(output_dir, FIGURES_FOLDER)
    os.makedirs(figures_dir, exist_ok=True)

    logger.info(f"\n{'='*70}")
    logger.info("PROCESSING GENBANK FILES")
    logger.info(f"{'='*70}")

    for idx, gb_file in enumerate(gb_files, 1):
        logger.info(f"[{idx}/{len(gb_files)}] Processing: {os.path.basename(gb_file)}")

        try:
            # One pass per feature type -- Excel data and figure data are identical
            species_g, gene_rows, gene_fig = extract_gene_introns_from_genbank(gb_file)
            species_t, trna_rows, trna_fig = extract_tRNA_introns_from_genbank(gb_file)

            all_gene_excel.extend(gene_rows)
            all_trna_excel.extend(trna_rows)
            all_gene_figs.extend(gene_fig)   # collect for combined figure
            all_trna_figs.extend(trna_fig)

            logger.info(f"  Genes with introns : {len(gene_rows)}")
            logger.info(f"  tRNAs with introns : {len(trna_rows)}")

            # Export figure data as JSON (for R figure generation)
            if gene_fig or trna_fig:
                species  = species_g if species_g != 'Unknown species' else species_t
                json_path = os.path.join(
                    figures_dir, f"{safe_filename(species)}_figure_data.json"
                )
                export_species_json(species, gene_fig, trna_fig, json_path)

        except Exception as e:
            logger.warning(f"  Warning: Error processing {os.path.basename(gb_file)}: {e}")
            import traceback
            traceback.print_exc()
            continue

    if not all_gene_excel and not all_trna_excel:
        raise ValueError("No introns found in any GenBank file")

    # Excel output
    logger.info(f"\n{'='*70}")
    logger.info("GENERATING EXCEL OUTPUT")
    logger.info(f"{'='*70}")

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f"intron_data_{timestamp}.xlsx")
    create_combined_excel_output(all_gene_excel, all_trna_excel, output_file)

    # Export combined JSON + generate R figures automatically
    logger.info(f"\n{'='*70}")
    logger.info("GENERATING FIGURES (R)")
    logger.info(f"{'='*70}")

    export_combined_json(
        all_gene_figs,
        all_trna_figs,
        os.path.join(figures_dir, 'combined_figure_data.json')
    )

    generate_r_visualizations(figures_dir)

    return len(all_gene_excel), len(all_trna_excel)


# ============================================================================
# CLI ENTRY POINT
# ============================================================================

def main():
    """Main pipeline with command-line argument handling."""
    parser = argparse.ArgumentParser(
        description='CGAS Module 11: Gene and tRNA Intron Extraction and Visualization',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  %(prog)s                          # Process current directory
  %(prog)s -i genbank_files/        # Process specific directory
  %(prog)s -i data/ -o results/     # Custom input and output
        '''
    )
    parser.add_argument('-i', '--input',  type=str, default='.',
                        help='Input directory containing GenBank files (default: current)')
    parser.add_argument('-o', '--output', type=str, default=None,
                        help='Output directory (default: Module11_Intron_Analysis/)')
    args = parser.parse_args()

    input_dir  = os.path.abspath(args.input)
    output_dir = os.path.abspath(args.output) if args.output \
                 else os.path.join(input_dir, OUTPUT_FOLDER)

    print(f"\n{'='*70}")
    print("CGAS MODULE 11: INTRON EXTRACTION, ANALYSIS, AND VISUALIZATION")
    print(f"{'='*70}")
    print(f"  Input directory  : {input_dir}")
    print(f"  Output directory : {output_dir}")
    print(f"  Max intron length: {MAX_INTRON_LENGTH} bp")
    print(f"  Figures          : R (ggplot2) — automatic generation")
    print()

    os.makedirs(output_dir, exist_ok=True)

    try:
        num_gene, num_trna = process_all_genbank_files(input_dir, output_dir)

        print(f"\n{'='*70}")
        print("ANALYSIS COMPLETE")
        print(f"{'='*70}")
        print(f"  Gene introns  : {num_gene}")
        print(f"  tRNA introns  : {num_trna}")
        print(f"  Output        : {output_dir}/")
        print(f"  Figures       : {os.path.join(output_dir, FIGURES_FOLDER)}/")
        print(f"                  (per-species + combined PNG & PDF)")
        print(f"{'='*70}\n")

    except FileNotFoundError as e:
        logger.error(f"\n ERROR: {e}")
        sys.exit(1)
    except ValueError as e:
        logger.error(f"\n ERROR: {e}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"\n FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n Analysis interrupted by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\n\n FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

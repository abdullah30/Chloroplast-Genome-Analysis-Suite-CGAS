#!/usr/bin/env python3
"""
CGAS Module 12: Comprehensive Chloroplast SSR Analysis
=======================================================

This script performs comprehensive Simple Sequence Repeat (SSR) detection and 
classification in chloroplast genomes from GenBank files. SSRs are identified,
classified, and analyzed by:
1. Genomic region (LSC/SSC/IR)
2. Motif type (Mono/Di/Tri/Tetra/Penta/Hexa)
3. Genomic location (genes, tRNAs, rRNAs, introns, intergenic spacers)

Key Features:
1. Detects SSRs based on customizable repeat thresholds
2. Classifies SSRs by genomic region and motif type
3. Determines precise genomic locations (genes, introns, IGS)
4. Handles multi-part genes and trans-spliced features
5. Creates publication-quality Excel outputs with:
   - Detailed SSR catalog (all SSRs with positions)
   - Summary statistics per species
   - Individual species sheets for easy comparison

Author: Abdullah
Version: 2.0 (Module 12 - CGAS Integration)
Date: January 2026

Dependencies:
    Python: biopython, pandas, openpyxl

Usage:
    python cgas_module12.py
    python cgas_module12.py -i genbank_files/
    python cgas_module12.py -i data/ -o results/
    python cgas_module12.py -t 12,6,5,4,4,4

Output:
    Module12_SSR_Analysis/
        - all_ssrs_detailed.xlsx (comprehensive SSR catalog)
        - ssr_summary.xlsx (statistics per species)
        - ssr_by_species.xlsx (individual species sheets)
        - ssr_summary_for_r.csv (data for R visualization)
        - Figures/ (if R available):
            - SSR_Genomic_Regions.pdf and .png
            - SSR_Motif_Types.pdf and .png
            - SSR_Genomic_Locations.pdf and .png
            - excluded_categories.txt (if any categories have zero SSRs)

Notes:
    - Default thresholds: Mono=10, Di=5, Tri=4, Tetra=3, Penta=3, Hexa=3
    - All species and gene names are properly italicized
    - Handles IR regions, introns, and intergenic spacers
    - Trans-spliced genes (>10kb introns) are handled separately
    
Gene Notation:
    - gene_N: Indicates exon number (e.g., atpF_1 = atpF exon 1, rpoC1_2 = rpoC1 exon 2)
    - gene_N1-gene_N2: Intron between exons (e.g., atpF_1-atpF_2 = intron between exon 1 and 2)
    - gene1-gene2: Intergenic spacer between genes (e.g., petB-psbH)
    - gene_N-gene_M: IGS with exon numbers (e.g., petB_2-psbH_1 = IGS between petB exon 2 and psbH exon 1)
"""

import os
import sys
import argparse
import re
from pathlib import Path
from typing import Dict, List, Tuple, Set
from collections import defaultdict
from datetime import datetime

try:
    from Bio import SeqIO
    from Bio.SeqFeature import CompoundLocation
except ImportError as e:
    print(f"Error: BioPython not installed: {e}")
    print("Please install required packages using:")
    print("pip install biopython")
    sys.exit(1)

try:
    import pandas as pd
except ImportError as e:
    print(f"Error: pandas not installed: {e}")
    print("Please install required packages using:")
    print("pip install pandas")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Error: openpyxl not installed: {e}")
    print("Please install required packages using:")
    print("pip install openpyxl")
    sys.exit(1)


# ============================================================================
# CONSTANTS
# ============================================================================

OUTPUT_FOLDER = "Module12_SSR_Analysis"
MAX_INTRON_LENGTH = 10000  # Threshold for trans-splicing detection
DEFAULT_THRESHOLDS = {1: 10, 2: 5, 3: 4, 4: 3, 5: 3, 6: 3}
VALID_EXTENSIONS = ('.gb', '.gbff', '.genbank', '.gbk')


# ============================================================================
# SSR ANALYZER CLASS
# ============================================================================

class ChloroplastSSRAnalyzer:
    """Main analyzer class for SSR detection and classification"""
    
    def __init__(self, gb_folder: str, output_folder: str = OUTPUT_FOLDER):
        """
        Initialize SSR analyzer.
        
        Parameters:
        -----------
        gb_folder : str
            Path to folder containing GenBank files
        output_folder : str
            Path to output folder for results
        """
        self.gb_folder = gb_folder
        self.output_folder = output_folder
        self.thresholds = DEFAULT_THRESHOLDS.copy()
        os.makedirs(output_folder, exist_ok=True)
    
    def find_ssrs(self, sequence: str) -> List[Tuple]:
        """
        Detect SSRs in sequence based on motif thresholds.
        
        Parameters:
        -----------
        sequence : str
            DNA sequence to analyze
            
        Returns:
        --------
        List[Tuple]
            List of (motif, length, start, end, repeat_count) tuples
        """
        ssrs = []
        seq_upper = str(sequence).upper()
        seq_len = len(seq_upper)
        
        # Track positions already in SSRs to avoid overlaps
        covered = set()
        
        for motif_len in range(1, 7):  # 1-6 bp motifs
            min_repeats = self.thresholds[motif_len]
            
            # Slide through sequence
            i = 0
            while i < seq_len - motif_len * min_repeats + 1:
                # Skip if position already covered
                if i in covered:
                    i += 1
                    continue
                
                # Get potential motif
                motif = seq_upper[i:i + motif_len]
                
                # Skip if contains non-ATCG bases
                if not all(base in 'ATCG' for base in motif):
                    i += 1
                    continue
                
                # Count consecutive repeats of this motif
                repeat_count = 0
                pos = i
                while pos + motif_len <= seq_len and seq_upper[pos:pos + motif_len] == motif:
                    repeat_count += 1
                    pos += motif_len
                
                # If meets threshold, record SSR
                if repeat_count >= min_repeats:
                    ssr_start = i + 1  # 1-based
                    ssr_end = pos      # 1-based, inclusive
                    ssr_length = pos - i
                    
                    # Mark these positions as covered
                    for p in range(i, pos):
                        covered.add(p)
                    
                    ssrs.append((motif, ssr_length, ssr_start, ssr_end, repeat_count))
                    
                    # Jump past this SSR
                    i = pos
                else:
                    i += 1
        
        return sorted(ssrs, key=lambda x: x[2])
    
    def extract_regions(self, record) -> Dict:
        """
        Extract LSC, SSC, IR boundaries from GenBank record.
        
        Parameters:
        -----------
        record : Bio.SeqRecord
            GenBank sequence record
            
        Returns:
        --------
        Dict
            Dictionary mapping region names to (start, end) tuples
        """
        seq_len = len(record.seq)
        ir_features = []
        
        for feat in record.features:
            if feat.type != "repeat_region":
                continue
            
            note_text = " ".join(feat.qualifiers.get("note", []))
            rpt_type = " ".join(feat.qualifiers.get("rpt_type", []))
            combined = (note_text + " " + rpt_type).lower()
            
            if any(term in combined for term in ["inverted", "ir", "inverted repeat"]):
                try:
                    ir_start = int(feat.location.start) + 1
                    ir_end = int(feat.location.end)
                    ir_features.append((ir_start, ir_end))
                except:
                    continue
        
        if len(ir_features) >= 2:
            ir_features = sorted(ir_features, key=lambda x: x[0])
            irb_start, irb_end = ir_features[0]
            ira_start, ira_end = ir_features[1]
            
            lsc_start = 1
            lsc_end = irb_start - 1
            ssc_start = irb_end + 1
            ssc_end = ira_start - 1
            
            return {
                'LSC': (lsc_start, lsc_end),
                'SSC': (ssc_start, ssc_end),
                'IRb': (irb_start, irb_end),
                'IRa': (ira_start, ira_end)
            }
        elif len(ir_features) == 1:
            ir_start, ir_end = ir_features[0]
            lsc_start = 1
            lsc_end = ir_start - 1
            ssc_start = ir_end + 1
            ssc_end = seq_len
            
            return {
                'LSC': (lsc_start, lsc_end),
                'SSC': (ssc_start, ssc_end),
                'IR': (ir_start, ir_end)
            }
        else:
            return {'LSC': (1, seq_len)}
    
    def classify_region(self, pos: int, regions: Dict) -> str:
        """
        Classify SSR position into genomic region.
        
        Parameters:
        -----------
        pos : int
            Position to classify
        regions : Dict
            Dictionary of region boundaries
            
        Returns:
        --------
        str
            Region name (LSC, SSC, IR, or Unknown)
        """
        for region_name, (start, end) in regions.items():
            if start <= pos <= end:
                if region_name in ['IRb', 'IRa']:
                    return 'IR'
                return region_name
        return 'Unknown'
    
    def classify_motif(self, motif: str) -> str:
        """
        Classify SSR by motif length.
        
        Parameters:
        -----------
        motif : str
            SSR motif sequence
            
        Returns:
        --------
        str
            Motif type (Mono, Di, Tri, Tetra, Penta, Hexa)
        """
        motif_types = {
            1: "Mono", 2: "Di", 3: "Tri",
            4: "Tetra", 5: "Penta", 6: "Hexa"
        }
        return motif_types.get(len(motif), "Other")
    
    def get_genomic_location(self, ssr_start: int, ssr_end: int, record) -> Tuple[str, str]:
        """
        Determine precise genomic location of SSR.
        
        This function prioritizes:
        1. tRNA features
        2. rRNA features
        3. Gene features (with intron detection)
        4. Intergenic spacers
        
        Parameters:
        -----------
        ssr_start : int
            SSR start position
        ssr_end : int
            SSR end position
        record : Bio.SeqRecord
            GenBank sequence record
            
        Returns:
        --------
        Tuple[str, str]
            (location_name, location_type)
        """
        # Collect all features with their actual parts
        all_features = []
        
        for feat in record.features:
            if feat.type in ["gene", "tRNA", "rRNA", "CDS"]:
                gene_name = feat.qualifiers.get("gene", [""])[0]
                if not gene_name:
                    continue
                
                # Get overall span
                f_start = int(feat.location.start) + 1
                f_end = int(feat.location.end)
                
                # Get individual parts if split
                parts = []
                if hasattr(feat.location, 'parts'):
                    for part in feat.location.parts:
                        parts.append((int(part.start) + 1, int(part.end)))
                else:
                    parts = [(f_start, f_end)]
                
                # Calculate intron length if multi-part
                is_trans_spliced = False
                if len(parts) > 1:
                    # Check distance between parts
                    sorted_parts = sorted(parts, key=lambda x: x[0])
                    for i in range(len(sorted_parts) - 1):
                        intron_length = sorted_parts[i+1][0] - sorted_parts[i][1] - 1
                        if intron_length > MAX_INTRON_LENGTH:
                            is_trans_spliced = True
                            break
                
                # Also check for explicit trans_splicing qualifier
                if 'trans_splicing' in feat.qualifiers:
                    is_trans_spliced = True
                
                all_features.append({
                    'type': feat.type,
                    'name': gene_name,
                    'start': f_start,
                    'end': f_end,
                    'parts': parts,
                    'feature': feat,
                    'is_trans_spliced': is_trans_spliced
                })
        
        # Helper: check if SSR is in any part of a feature
        def is_in_parts(ssr_start, ssr_end, parts):
            for part_start, part_end in parts:
                if ssr_start >= part_start and ssr_end <= part_end:
                    return True
            return False
        
        # Priority 1: Check tRNAs
        trnas = [f for f in all_features if f['type'] == 'tRNA' and 
                 not f.get('is_trans_spliced', False) and
                 ssr_start >= f['start'] and ssr_end <= f['end']]
        
        if trnas:
            trna = trnas[0]
            if len(trna['parts']) > 1:
                # tRNA has intron
                if not is_in_parts(ssr_start, ssr_end, trna['parts']):
                    # SSR is in tRNA intron
                    # Check for embedded gene (like matK in trnK)
                    genes = [f for f in all_features if f['type'] == 'gene' and 
                            not f.get('is_trans_spliced', False) and
                            f['name'] != trna['name'] and
                            ssr_start >= f['start'] and ssr_end <= f['end']]
                    
                    if genes:
                        # Format: trnK-UUU_2-matK
                        return f"{trna['name']}_2-{genes[0]['name']}", "Intron"
                    else:
                        # Format: trnI-GAU_1-trnI-GAU_2
                        return f"{trna['name']}_1-{trna['name']}_2", "Intron"
            
            return trna['name'], "tRNA"
        
        # Priority 2: Check rRNAs
        rrnas = [f for f in all_features if f['type'] == 'rRNA' and 
                 ssr_start >= f['start'] and ssr_end <= f['end']]
        if rrnas:
            return rrnas[0]['name'], "rRNA"
        
        # Priority 3: Check genes (non-trans-spliced only)
        genes = [f for f in all_features if f['type'] == 'gene' and 
                 not f.get('is_trans_spliced', False) and
                 ssr_start >= f['start'] and ssr_end <= f['end']]
        
        if genes:
            gene = genes[0]
            gene_name = gene['name']
            
            # Find CDS for this gene
            cdss = [f for f in all_features if f['type'] == 'CDS' and 
                   f['name'] == gene_name and not f.get('is_trans_spliced', False)]
            
            if cdss:
                cds = cdss[0]
                
                # Check if SSR is in CDS exon
                if is_in_parts(ssr_start, ssr_end, cds['parts']):
                    # SSR is in exon - check if multi-part CDS
                    if len(cds['parts']) > 1:
                        # Find which part
                        for i, (pstart, pend) in enumerate(cds['parts'], 1):
                            if ssr_start >= pstart and ssr_end <= pend:
                                return f"{gene_name}_{i}", "Gene"
                    return gene_name, "Gene"
                
                # SSR is in gene but not in CDS = intron
                if len(cds['parts']) > 1:
                    # Multi-part CDS - check which intron
                    sorted_parts = sorted(cds['parts'], key=lambda x: x[0])
                    for i in range(len(sorted_parts) - 1):
                        intron_start = sorted_parts[i][1] + 1
                        intron_end = sorted_parts[i+1][0] - 1
                        if ssr_start >= intron_start and ssr_end <= intron_end:
                            return f"{gene_name}_{i+1}-{gene_name}_{i+2}", "Intron"
                    
                    return f"{gene_name}_1-{gene_name}_2", "Intron"
                else:
                    return gene_name, "Intron"
            else:
                # Gene has no CDS
                return gene_name, "Gene"
        
        # Priority 3b: Check if in trans-spliced CDS parts only
        for cds_feat in [f for f in all_features if f['type'] == 'CDS' and f.get('is_trans_spliced', False)]:
            if is_in_parts(ssr_start, ssr_end, cds_feat['parts']):
                # Find which part
                for i, (pstart, pend) in enumerate(cds_feat['parts'], 1):
                    if ssr_start >= pstart and ssr_end <= pend:
                        if len(cds_feat['parts']) > 1:
                            return f"{cds_feat['name']}_{i}", "Gene"
                        else:
                            return cds_feat['name'], "Gene"
        
        # Priority 4: Intergenic
        flanking = self.find_flanking_genes(ssr_start, ssr_end, all_features)
        if flanking:
            return flanking, "IGS"
        
        return "intergenic", "IGS"
    
    def find_flanking_genes(self, ssr_start: int, ssr_end: int, all_features: List[Dict]) -> str:
        """
        Find flanking genes for intergenic spacer SSRs.
        
        Parameters:
        -----------
        ssr_start : int
            SSR start position
        ssr_end : int
            SSR end position
        all_features : List[Dict]
            List of feature dictionaries
            
        Returns:
        --------
        str
            Flanking gene notation (e.g., "petB-psbH") or None
        """
        genes_before = []
        genes_after = []
        
        for feat in all_features:
            if feat['type'] not in ['gene', 'tRNA', 'rRNA']:
                continue
            
            # Skip trans-spliced genes
            if feat.get('is_trans_spliced', False):
                continue
            
            f_start = feat['start']
            f_end = feat['end']
            gene_name = feat['name']
            
            if f_end < ssr_start:
                # Find CDS to check if multi-part
                cdss = [f for f in all_features if f['type'] == 'CDS' and f['name'] == gene_name]
                if cdss and len(cdss[0]['parts']) > 1:
                    # Multi-part gene - find which part is closest
                    last_part_num = len(cdss[0]['parts'])
                    genes_before.append((f_end, f"{gene_name}_{last_part_num}"))
                else:
                    genes_before.append((f_end, gene_name))
            
            elif f_start > ssr_end:
                # Find CDS to check if multi-part
                cdss = [f for f in all_features if f['type'] == 'CDS' and f['name'] == gene_name]
                if cdss and len(cdss[0]['parts']) > 1:
                    # Multi-part gene - use first part
                    genes_after.append((f_start, f"{gene_name}_1"))
                elif feat['type'] == 'tRNA' and len(feat['parts']) > 1:
                    # Multi-part tRNA
                    genes_after.append((f_start, f"{gene_name}_2"))
                else:
                    genes_after.append((f_start, gene_name))
        
        if genes_before and genes_after:
            upstream = max(genes_before, key=lambda x: x[0])[1]
            downstream = min(genes_after, key=lambda x: x[0])[1]
            return f"{upstream}-{downstream}"
        
        return None
    
    def process_genbank_file(self, gb_file: str) -> pd.DataFrame:
        """
        Process a single GenBank file to extract SSRs.
        
        Parameters:
        -----------
        gb_file : str
            Path to GenBank file
            
        Returns:
        --------
        pd.DataFrame
            DataFrame with SSR data
        """
        try:
            record = SeqIO.read(gb_file, "genbank")
        except Exception as e:
            print(f"  ❌ ERROR reading {os.path.basename(gb_file)}: {e}")
            return None
        
        # Extract species name from filename (without extension)
        species = os.path.splitext(os.path.basename(gb_file))[0]
        
        # Get organism name from GenBank if available
        if 'organism' in record.annotations:
            species = record.annotations['organism']
        
        print(f"  Processing {species}...")
        
        regions = self.extract_regions(record)
        ssrs = self.find_ssrs(record.seq)
        
        ssr_data = []
        for motif, length, start, end, repeat_count in ssrs:
            region = self.classify_region(start, regions)
            motif_type = self.classify_motif(motif)
            location, loc_type = self.get_genomic_location(start, end, record)
            
            ssr_data.append({
                'Species': species,
                'Type': motif,
                'Motif_Type': motif_type,
                'Repeat_Count': repeat_count,
                'Length': length,
                'Start': start,
                'End': end,
                'Region': region,
                'Location': location,
                'Location_Type': loc_type
            })
        
        print(f"    ✓ Found {len(ssr_data)} SSRs")
        return pd.DataFrame(ssr_data)
    
    def add_abbreviation_footnotes(self, ws, start_row: int) -> int:
        """
        Add abbreviation footnotes to Excel worksheet.
        
        Parameters:
        -----------
        ws : openpyxl.worksheet.worksheet.Worksheet
            Worksheet to add footnotes to
        start_row : int
            Row to start adding footnotes
            
        Returns:
        --------
        int
            Next available row
        """
        from openpyxl.styles.fonts import Font as RichFont
        from openpyxl.cell.text import InlineFont
        from openpyxl.cell.rich_text import TextBlock, CellRichText
        
        # Standard footnotes (without gene examples)
        footnotes = [
            "Abbreviations:",
            "LSC: Large Single Copy region",
            "SSC: Small Single Copy region", 
            "IR: Inverted Repeat region",
            "IGS: Intergenic Spacer region",
            "Mono: Mononucleotide repeat",
            "Di: Dinucleotide repeat",
            "Tri: Trinucleotide repeat",
            "Tetra: Tetranucleotide repeat",
            "Penta: Pentanucleotide repeat",
            "Hexa: Hexanucleotide repeat",
            "tRNA: Transfer RNA",
            "rRNA: Ribosomal RNA"
        ]
        
        row = start_row
        for footnote in footnotes:
            cell = ws.cell(row, 1, footnote)
            if footnote == "Abbreviations:":
                cell.font = Font(bold=True, size=10)
            else:
                cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='left', vertical='top')
            row += 1
        
        # Empty row for spacing
        row += 1
        
        # Notes section header
        cell = ws.cell(row, 1, "Notes:")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        row += 1
        
        # Note 1: Gene_N notation with italic gene names
        cell = ws.cell(row, 1)
        note1 = CellRichText()
        note1.append(TextBlock(InlineFont(sz=9.0), "Gene_N notation: N indicates exon number (e.g., "))
        note1.append(TextBlock(InlineFont(sz=9.0, i=True), "atpF"))
        note1.append(TextBlock(InlineFont(sz=9.0), "_1 = "))
        note1.append(TextBlock(InlineFont(sz=9.0, i=True), "atpF"))
        note1.append(TextBlock(InlineFont(sz=9.0), " exon 1, "))
        note1.append(TextBlock(InlineFont(sz=9.0, i=True), "rpoC1"))
        note1.append(TextBlock(InlineFont(sz=9.0), "_2 = "))
        note1.append(TextBlock(InlineFont(sz=9.0, i=True), "rpoC1"))
        note1.append(TextBlock(InlineFont(sz=9.0), " exon 2)"))
        cell.value = note1
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        row += 1
        
        # Note 2: Gene_N1-Gene_N2 notation with italic gene names
        cell = ws.cell(row, 1)
        note2 = CellRichText()
        note2.append(TextBlock(InlineFont(sz=9.0), "Gene_N1-Gene_N2: Indicates intron between exons (e.g., "))
        note2.append(TextBlock(InlineFont(sz=9.0, i=True), "atpF"))
        note2.append(TextBlock(InlineFont(sz=9.0), "_1-"))
        note2.append(TextBlock(InlineFont(sz=9.0, i=True), "atpF"))
        note2.append(TextBlock(InlineFont(sz=9.0), "_2 = intron between exon 1 and 2)"))
        cell.value = note2
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        row += 1
        
        # Note 3: IGS notation with italic gene names
        cell = ws.cell(row, 1)
        note3 = CellRichText()
        note3.append(TextBlock(InlineFont(sz=9.0), "IGS notation: Shows flanking genes (e.g., "))
        note3.append(TextBlock(InlineFont(sz=9.0, i=True), "petB"))
        note3.append(TextBlock(InlineFont(sz=9.0), "_2-"))
        note3.append(TextBlock(InlineFont(sz=9.0, i=True), "psbH"))
        note3.append(TextBlock(InlineFont(sz=9.0), "_1 = intergenic spacer between "))
        note3.append(TextBlock(InlineFont(sz=9.0, i=True), "petB"))
        note3.append(TextBlock(InlineFont(sz=9.0), " exon 2 and "))
        note3.append(TextBlock(InlineFont(sz=9.0, i=True), "psbH"))
        note3.append(TextBlock(InlineFont(sz=9.0), " exon 1)"))
        cell.value = note3
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        row += 1
        
        return row
    
    def create_formatted_excel(self, all_data: pd.DataFrame, output_file: str):
        """
        Create publication-quality Excel file with formatting.
        
        Parameters:
        -----------
        all_data : pd.DataFrame
            Combined SSR data from all files
        output_file : str
            Path to output Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "All SSRs"
        
        # Headers
        headers = ['Type', 'Motif_Type', 'Repeat_Count', 'Length', 'Start', 'End', 'Region', 'Location', 'Location_Type']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row = 2
        
        # Sort by species, then by Start position
        all_data = all_data.sort_values(['Species', 'Start'])
        
        for species in sorted(all_data['Species'].unique()):
            species_data = all_data[all_data['Species'] == species].sort_values('Start')
            
            # Add species row (merged across all columns) - ITALIC
            ws.merge_cells(f'A{row}:I{row}')
            species_cell = ws.cell(row, 1, species)
            species_cell.font = Font(italic=True, bold=True, size=11)
            species_cell.alignment = Alignment(horizontal='center', vertical='center')
            species_cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            row += 1
            
            # Add data rows for this species
            for _, ssr_row in species_data.iterrows():
                for col_idx, col_name in enumerate(headers, 1):
                    value = ssr_row[col_name]
                    cell = ws.cell(row, col_idx, value)
                    
                    # Italicize gene names in Location column
                    if col_name == 'Location':
                        # Check if location contains gene names
                        if ssr_row['Location_Type'] in ['Gene', 'tRNA', 'rRNA', 'Intron', 'IGS']:
                            cell.font = Font(italic=True, size=10)
                        else:
                            cell.font = Font(size=10)
                    else:
                        cell.font = Font(size=10)
                    
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
        
        # Adjust column widths
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        ws.column_dimensions['H'].width = 25  # Location column wider
        
        # Add footnotes
        self.add_abbreviation_footnotes(ws, row + 2)
        
        wb.save(output_file)
        print(f"  ✓ Detailed Excel file created: {os.path.basename(output_file)}")
    
    def export_summary_for_r(self, all_data: pd.DataFrame, output_folder: str):
        """
        Export summary data to CSV files for R visualization.
        
        Parameters:
        -----------
        all_data : pd.DataFrame
            Combined SSR data
        output_folder : str
            Path to output folder
            
        Returns:
        --------
        Tuple[pd.DataFrame, str]
            Summary DataFrame and CSV file path
        """
        summary_data = []
        
        for species in sorted(all_data['Species'].unique()):
            species_data = all_data[all_data['Species'] == species]
            
            region_counts = species_data['Region'].value_counts().to_dict()
            motif_counts = species_data['Motif_Type'].value_counts().to_dict()
            loc_counts = species_data['Location_Type'].value_counts().to_dict()
            
            summary = {
                'Species': species,
                'Total_SSRs': len(species_data),
                'LSC': region_counts.get('LSC', 0),
                'SSC': region_counts.get('SSC', 0),
                'IR': region_counts.get('IR', 0),
                'Mono': motif_counts.get('Mono', 0),
                'Di': motif_counts.get('Di', 0),
                'Tri': motif_counts.get('Tri', 0),
                'Tetra': motif_counts.get('Tetra', 0),
                'Penta': motif_counts.get('Penta', 0),
                'Hexa': motif_counts.get('Hexa', 0),
                'Gene': loc_counts.get('Gene', 0),
                'tRNA': loc_counts.get('tRNA', 0),
                'rRNA': loc_counts.get('rRNA', 0),
                'Intron': loc_counts.get('Intron', 0),
                'IGS': loc_counts.get('IGS', 0)
            }
            summary_data.append(summary)
        
        summary_df = pd.DataFrame(summary_data)
        
        # Export to CSV for R
        csv_file = os.path.join(output_folder, "ssr_summary_for_r.csv")
        summary_df.to_csv(csv_file, index=False)
        print(f"  ✓ CSV file for R created: {os.path.basename(csv_file)}")
        
        return summary_df, csv_file
    
    def create_summary_tables(self, all_data: pd.DataFrame) -> Workbook:
        """
        Create summary statistics tables.
        
        Parameters:
        -----------
        all_data : pd.DataFrame
            Combined SSR data
            
        Returns:
        --------
        Workbook
            Excel workbook with summary data
        """
        summary_data = []
        
        for species in sorted(all_data['Species'].unique()):
            species_data = all_data[all_data['Species'] == species]
            
            region_counts = species_data['Region'].value_counts().to_dict()
            motif_counts = species_data['Motif_Type'].value_counts().to_dict()
            loc_counts = species_data['Location_Type'].value_counts().to_dict()
            
            summary = {
                'Species': species,
                'Total_SSRs': len(species_data),
                'LSC': region_counts.get('LSC', 0),
                'SSC': region_counts.get('SSC', 0),
                'IR': region_counts.get('IR', 0),
                'Mono': motif_counts.get('Mono', 0),
                'Di': motif_counts.get('Di', 0),
                'Tri': motif_counts.get('Tri', 0),
                'Tetra': motif_counts.get('Tetra', 0),
                'Penta': motif_counts.get('Penta', 0),
                'Hexa': motif_counts.get('Hexa', 0),
                'Gene': loc_counts.get('Gene', 0),
                'tRNA': loc_counts.get('tRNA', 0),
                'rRNA': loc_counts.get('rRNA', 0),
                'Intron': loc_counts.get('Intron', 0),
                'IGS': loc_counts.get('IGS', 0)
            }
            summary_data.append(summary)
        
        summary_df = pd.DataFrame(summary_data)
        
        # Create formatted Excel with italic species names
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        
        # Write headers
        headers = list(summary_df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Write data with italic species names
        for row_idx, (idx, row_data) in enumerate(summary_df.iterrows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row_idx, col_idx, value)
                # Italicize species column
                if col_idx == 1:
                    cell.font = Font(italic=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Add footnotes
        last_data_row = len(summary_df) + 1
        self.add_abbreviation_footnotes(ws, last_data_row + 2)
        
        return wb
    
    def create_individual_species_sheets(self, all_data: pd.DataFrame, output_file: str):
        """
        Create Excel file with individual sheets per species.
        
        Parameters:
        -----------
        all_data : pd.DataFrame
            Combined SSR data
        output_file : str
            Path to output file
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        headers = ['Type', 'Motif_Type', 'Repeat_Count', 'Length', 'Start', 'End', 'Region', 'Location', 'Location_Type']
        
        for species in sorted(all_data['Species'].unique()):
            species_data = all_data[all_data['Species'] == species].sort_values('Start')
            
            # Create sheet (truncate species name if too long)
            sheet_name = species[:31] if len(species) > 31 else species
            ws = wb.create_sheet(sheet_name)
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row_idx, ssr_row in enumerate(species_data.iterrows(), 2):
                _, data = ssr_row
                for col_idx, col_name in enumerate(headers, 1):
                    value = data[col_name]
                    cell = ws.cell(row_idx, col_idx, value)
                    
                    # Italicize gene names in Location column
                    if col_name == 'Location':
                        if data['Location_Type'] in ['Gene', 'tRNA', 'rRNA', 'Intron', 'IGS']:
                            cell.font = Font(italic=True, size=10)
                        else:
                            cell.font = Font(size=10)
                    else:
                        cell.font = Font(size=10)
                    
                    cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            ws.column_dimensions['H'].width = 25
            
            # Add footnotes
            last_row = len(species_data) + 1
            self.add_abbreviation_footnotes(ws, last_row + 2)
        
        wb.save(output_file)
        print(f"  ✓ Individual species file created: {os.path.basename(output_file)}")
    
    def run_analysis(self):
        """Main analysis pipeline execution."""
        # Find GenBank files
        gb_files = [
            os.path.join(self.gb_folder, f)
            for f in os.listdir(self.gb_folder)
            if f.endswith(VALID_EXTENSIONS)
        ]
        
        if not gb_files:
            print("\n❌ ERROR: No GenBank files found!")
            print(f"Expected extensions: {', '.join(VALID_EXTENSIONS)}")
            return
        
        print(f"\nFound {len(gb_files)} GenBank file(s)")
        
        all_ssr_data = []
        
        print(f"\n{'='*70}")
        print("PROCESSING GENBANK FILES")
        print(f"{'='*70}")
        
        for gb_file in sorted(gb_files):
            df = self.process_genbank_file(gb_file)
            if df is not None and not df.empty:
                all_ssr_data.append(df)
        
        if not all_ssr_data:
            print("\n❌ No SSRs detected in any files!")
            return
        
        combined_df = pd.concat(all_ssr_data, ignore_index=True)
        
        print(f"\n{'='*70}")
        print("GENERATING OUTPUT FILES")
        print(f"{'='*70}")
        
        # Export summary data for R and create Excel summary
        summary_df, csv_file = self.export_summary_for_r(combined_df, self.output_folder)
        
        # Create formatted Excel with merged species cells
        detailed_file = os.path.join(self.output_folder, "all_ssrs_detailed.xlsx")
        self.create_formatted_excel(combined_df, detailed_file)
        
        # Create summary table
        summary_file = os.path.join(self.output_folder, "ssr_summary.xlsx")
        summary_wb = self.create_summary_tables(combined_df)
        summary_wb.save(summary_file)
        print(f"  ✓ Summary table created: {os.path.basename(summary_file)}")
        
        # Create individual species sheets
        species_file = os.path.join(self.output_folder, "ssr_by_species.xlsx")
        self.create_individual_species_sheets(combined_df, species_file)
        
        # Generate R visualizations
        generate_r_visualizations(csv_file, self.output_folder, summary_df)
        
        print(f"\n{'='*70}")
        print("ANALYSIS COMPLETE")
        print(f"{'='*70}")
        print(f"✓ Total SSRs detected: {len(combined_df)}")
        print(f"✓ Species analyzed: {len(combined_df['Species'].unique())}")
        print(f"✓ Output files saved in: {self.output_folder}/")
        print(f"{'='*70}\n")


# ============================================================================
# R VISUALIZATION FUNCTIONS
# ============================================================================

def generate_r_visualizations(csv_file: str, output_folder: str, summary_df: pd.DataFrame):
    """Generate R-based visualizations for SSR data."""
    import subprocess
    
    print(f"\n{'='*70}")
    print("GENERATING R VISUALIZATIONS")
    print(f"{'='*70}")
    
    # Check R availability
    try:
        result = subprocess.run(['Rscript', '--version'],
                              capture_output=True,
                              text=True,
                              timeout=10)
        if result.returncode != 0:
            print("  ⚠ R not found. Skipping visualization.")
            return
    except (subprocess.TimeoutExpired, FileNotFoundError):
        print("  ⚠ R not found. Skipping visualization.")
        return
    
    # Check zero columns
    zero_columns = []
    location_cols = ['Gene', 'tRNA', 'rRNA', 'Intron', 'IGS']
    for col in location_cols:
        if col in summary_df.columns and summary_df[col].sum() == 0:
            zero_columns.append(col)
    
    # Create note about excluded columns
    if zero_columns:
        figures_dir = os.path.join(output_folder, "Figures")
        os.makedirs(figures_dir, exist_ok=True)
        note_file = os.path.join(figures_dir, "excluded_categories.txt")
        with open(note_file, 'w') as f:
            f.write("SSR Location Categories Excluded from Visualization\n")
            f.write("=" * 60 + "\n\n")
            f.write("The following categories were excluded from Figure 3\n")
            f.write("because all species had zero SSRs:\n\n")
            for col in zero_columns:
                f.write(f"  - {col}\n")
        print(f"  ✓ Exclusion note created")
    
    # Create R script
    r_script_content = create_r_visualization_script(
        csv_file, 
        len(summary_df), 
        zero_columns
    )
    
    r_script_file = os.path.join(output_folder, "generate_ssr_figures.R")
    with open(r_script_file, 'w') as f:
        f.write(r_script_content)
    
    print(f"  ✓ R script created")
    
    # Execute R script
    print(f"\n  Executing R script...")
    try:
        result = subprocess.run(['Rscript', os.path.basename(r_script_file)],
                              capture_output=True,
                              text=True,
                              timeout=120,
                              cwd=output_folder)
        
        if result.stdout:
            print(result.stdout)
        
        if result.returncode == 0:
            print(f"\n  ✓ Visualization completed")
        else:
            print(f"  ✗ R script failed (code: {result.returncode})")
            if result.stderr:
                print(result.stderr[:500])
    except Exception as e:
        print(f"  ✗ Error: {e}")


def create_r_visualization_script(csv_file: str, num_species: int, zero_columns: List[str]) -> str:
    """
    Create R script for SSR visualizations with formula-based dynamic sizing.
    
    Parameters:
    -----------
    csv_file : str
        Path to CSV file
    num_species : int
        Number of species
    zero_columns : List[str]
        List of location columns with all zeros
        
    Returns:
    --------
    str
        R script content with formula-based dynamic sizing
    """
    
    # Determine which location columns to include
    all_location_cols = ['Gene', 'tRNA', 'rRNA', 'Intron', 'IGS']
    included_location_cols = [col for col in all_location_cols if col not in zero_columns]
    
    script = f'''
# CGAS Module 12: SSR Visualization Script with Formula-Based Dynamic Sizing
# Generated automatically

suppressPackageStartupMessages({{
  library(ggplot2)
  library(dplyr)
  library(tidyr)
  library(reshape2)
}})

cat("\\n========================================\\n")
cat("CGAS MODULE 12: SSR VISUALIZATION\\n")
cat("========================================\\n\\n")

# Read data
cat("Reading SSR summary data from CSV...\\n")
data <- read.csv("{os.path.basename(csv_file)}", check.names=FALSE)

if (nrow(data) == 0) {{
  stop("ERROR: CSV file is empty - no data to plot")
}}

num_species <- nrow(data)
cat(paste("  ✓ Loaded:", nrow(data), "rows\\n"))
cat(paste("  ✓ Number of species:", num_species, "\\n\\n"))

# Create figures directory
if (!dir.exists("Figures")) {{
  dir.create("Figures")
}}

# Dynamic sizing formula - publication quality
# Optimized for 2-30 species range
if (num_species <= 3) {{
  plot_width <- 4 + num_species * 1.2
}} else if (num_species <= 5) {{
  plot_width <- 7.6 + (num_species - 3) * 0.9
}} else if (num_species <= 10) {{
  plot_width <- 9.4 + (num_species - 5) * 0.6
}} else if (num_species <= 20) {{
  plot_width <- 12.4 + (num_species - 10) * 0.36
}} else {{
  plot_width <- 16
}}

plot_width <- min(18, max(5, plot_width))
plot_height <- 6

# Bar width - tighter for fewer species
if (num_species <= 5) {{
  bar_width_proportion <- 0.7
}} else if (num_species <= 10) {{
  bar_width_proportion <- 0.75
}} else {{
  bar_width_proportion <- 0.8
}}

# Font sizes
if (num_species <= 5) {{
  base_font_size <- 12
  axis_text_size <- 10
  title_size <- 14
  legend_text_size <- 10
}} else if (num_species <= 10) {{
  base_font_size <- 11
  axis_text_size <- 9
  title_size <- 13
  legend_text_size <- 9
}} else if (num_species <= 20) {{
  base_font_size <- 9
  axis_text_size <- 7
  title_size <- 11
  legend_text_size <- 8
}} else {{
  base_font_size <- 8
  axis_text_size <- 6
  title_size <- 10
  legend_text_size <- 7
}}

cat(paste("  ✓ Plot dimensions:", round(plot_width, 1), "x", plot_height, "inches\\n"))
cat(paste("  ✓ Bar width:", bar_width_proportion, "\\n"))
cat(paste("  ✓ Fonts: base=", base_font_size, ", axis=", axis_text_size, "\\n\\n"))

# ============================================================================
# FIGURE 1: SSR DISTRIBUTION BY GENOMIC REGIONS (LSC, SSC, IR)
# ============================================================================

cat("Generating Figure 1: SSR Distribution by Genomic Regions...\\n")

# Prepare data for stacked bar plot
region_data <- data %>%
  select(Species, LSC, SSC, IR) %>%
  pivot_longer(cols = c(LSC, SSC, IR), names_to = "Region", values_to = "Count")

# Order regions
region_data$Region <- factor(region_data$Region, levels = c("LSC", "SSC", "IR"))

# Define colors for regions
region_colors <- c(
  "LSC" = "#4472C4",    # Blue
  "SSC" = "#ED7D31",    # Orange
  "IR" = "#A5A5A5"      # Gray
)

# Create stacked bar plot with formula-based dynamic sizing
p1 <- ggplot(region_data, aes(x = Species, y = Count, fill = Region)) +
  geom_bar(stat = "identity", position = "stack", color = "black", linewidth = 0.3, width = bar_width_proportion) +
  scale_fill_manual(values = region_colors, name = "Genomic Region") +
  labs(x = "", y = "SSR Count", title = "SSR Distribution by Genomic Regions") +
  theme_minimal(base_size = base_font_size) +
  theme(
    axis.text.x = element_text(angle = 90, hjust = 1, vjust = 0.5, size = axis_text_size, face = "italic"),
    axis.text.y = element_text(size = axis_text_size),
    axis.title.y = element_text(size = base_font_size, face = "plain", margin = margin(r = 10)),
    plot.title = element_text(size = title_size, face = "plain", hjust = 0.5, margin = margin(b = 8)),
    legend.position = "bottom",
    legend.title = element_text(size = base_font_size, face = "plain"),
    legend.text = element_text(size = legend_text_size),
    panel.grid.major.x = element_blank(),
    panel.grid.minor = element_blank(),
    panel.grid.major.y = element_line(color = "gray90"),
    panel.border = element_rect(color = "black", fill = NA, linewidth = 0.5),
    plot.margin = margin(t = 5, r = 5, b = 10, l = 5)
  ) +
  scale_y_continuous(expand = expansion(mult = c(0, 0.05))) +
  scale_x_discrete(expand = expansion(mult = c(0.05, 0.05)))

# Save with formula-based dynamic dimensions
pdf("Figures/SSR_Genomic_Regions.pdf", width = plot_width, height = plot_height)
print(p1)
dev.off()

ggsave("Figures/SSR_Genomic_Regions.png", plot = p1, width = plot_width, height = plot_height, dpi = 600)

cat("  ✓ Figure 1 saved\\n\\n")

# ============================================================================
# FIGURE 2: SSR DISTRIBUTION BY MOTIF TYPE
# ============================================================================

cat("Generating Figure 2: SSR Distribution by Motif Type...\\n")

# Prepare data for stacked bar plot
motif_data <- data %>%
  select(Species, Mono, Di, Tri, Tetra, Penta, Hexa) %>%
  pivot_longer(cols = c(Mono, Di, Tri, Tetra, Penta, Hexa), names_to = "Motif", values_to = "Count")

# Order motifs
motif_data$Motif <- factor(motif_data$Motif, levels = c("Mono", "Di", "Tri", "Tetra", "Penta", "Hexa"))

# Define colors for motifs
motif_colors <- c(
  "Mono" = "#4472C4",    # Blue
  "Di" = "#ED7D31",      # Orange
  "Tri" = "#A5A5A5",     # Gray
  "Tetra" = "#FFC000",   # Yellow
  "Penta" = "#70AD47",   # Green
  "Hexa" = "#5B9BD5"     # Light Blue
)

# Create stacked bar plot with formula-based dynamic sizing
p2 <- ggplot(motif_data, aes(x = Species, y = Count, fill = Motif)) +
  geom_bar(stat = "identity", position = "stack", color = "black", linewidth = 0.3, width = bar_width_proportion) +
  scale_fill_manual(values = motif_colors, name = "Motif Type") +
  labs(x = "", y = "SSR Count", title = "SSR Distribution by Motif Type") +
  theme_minimal(base_size = base_font_size) +
  theme(
    axis.text.x = element_text(angle = 90, hjust = 1, vjust = 0.5, size = axis_text_size, face = "italic"),
    axis.text.y = element_text(size = axis_text_size),
    axis.title.y = element_text(size = base_font_size, face = "plain", margin = margin(r = 10)),
    plot.title = element_text(size = title_size, face = "plain", hjust = 0.5, margin = margin(b = 8)),
    legend.position = "bottom",
    legend.title = element_text(size = base_font_size, face = "plain"),
    legend.text = element_text(size = legend_text_size),
    panel.grid.major.x = element_blank(),
    panel.grid.minor = element_blank(),
    panel.grid.major.y = element_line(color = "gray90"),
    panel.border = element_rect(color = "black", fill = NA, linewidth = 0.5),
    plot.margin = margin(t = 5, r = 5, b = 10, l = 5)
  ) +
  scale_y_continuous(expand = expansion(mult = c(0, 0.05))) +
  scale_x_discrete(expand = expansion(mult = c(0.05, 0.05)))

# Save with formula-based dynamic dimensions
pdf("Figures/SSR_Motif_Types.pdf", width = plot_width, height = plot_height)
print(p2)
dev.off()

ggsave("Figures/SSR_Motif_Types.png", plot = p2, width = plot_width, height = plot_height, dpi = 600)

cat("  ✓ Figure 2 saved\\n\\n")

# ============================================================================
# FIGURE 3: SSR DISTRIBUTION BY GENOMIC LOCATION
# ============================================================================

cat("Generating Figure 3: SSR Distribution by Genomic Location...\\n")

'''

    # Add location plot code with conditional column selection
    if included_location_cols:
        cols_str = ", ".join(included_location_cols)
        cols_list = ", ".join([f'"{col}"' for col in included_location_cols])
        
        script += f'''
# Prepare data for stacked bar plot (excluding zero columns)
location_data <- data %>%
  select(Species, {cols_str}) %>%
  pivot_longer(cols = c({cols_list}), names_to = "Location", values_to = "Count")

# Order locations
location_data$Location <- factor(location_data$Location, levels = c({cols_list}))

# Define colors for locations
location_colors <- c(
  "Gene" = "#4472C4",    # Blue
  "tRNA" = "#ED7D31",    # Orange
  "rRNA" = "#A5A5A5",    # Gray
  "Intron" = "#FFC000",  # Yellow
  "IGS" = "#70AD47"      # Green
)

# Create stacked bar plot with formula-based dynamic sizing
p3 <- ggplot(location_data, aes(x = Species, y = Count, fill = Location)) +
  geom_bar(stat = "identity", position = "stack", color = "black", linewidth = 0.3, width = bar_width_proportion) +
  scale_fill_manual(values = location_colors, name = "Genomic Location") +
  labs(x = "", y = "SSR Count", title = "SSR Distribution by Genomic Location") +
  theme_minimal(base_size = base_font_size) +
  theme(
    axis.text.x = element_text(angle = 90, hjust = 1, vjust = 0.5, size = axis_text_size, face = "italic"),
    axis.text.y = element_text(size = axis_text_size),
    axis.title.y = element_text(size = base_font_size, face = "plain", margin = margin(r = 10)),
    plot.title = element_text(size = title_size, face = "plain", hjust = 0.5, margin = margin(b = 8)),
    legend.position = "bottom",
    legend.title = element_text(size = base_font_size, face = "plain"),
    legend.text = element_text(size = legend_text_size),
    panel.grid.major.x = element_blank(),
    panel.grid.minor = element_blank(),
    panel.grid.major.y = element_line(color = "gray90"),
    panel.border = element_rect(color = "black", fill = NA, linewidth = 0.5),
    plot.margin = margin(t = 5, r = 5, b = 10, l = 5)
  ) +
  scale_y_continuous(expand = expansion(mult = c(0, 0.05))) +
  scale_x_discrete(expand = expansion(mult = c(0.05, 0.05)))

# Save with formula-based dynamic dimensions
pdf("Figures/SSR_Genomic_Locations.pdf", width = plot_width, height = plot_height)
print(p3)
dev.off()

ggsave("Figures/SSR_Genomic_Locations.png", plot = p3, width = plot_width, height = plot_height, dpi = 600)

cat("  ✓ Figure 3 saved\\n\\n")
'''
    else:
        script += f'''
# All location categories have zero values, skipping Figure 3
cat("  ⚠ All location categories have zero SSRs, skipping Figure 3\\n\\n")
'''
    
    script += f'''
cat("\\n========================================\\n")
cat("VISUALIZATION COMPLETE\\n")
cat("========================================\\n")
cat(paste("  ✓ Generated figures for", num_species, "species\\n"))
cat(paste("  ✓ Plot dimensions:", round(plot_width, 1), "x", round(plot_height, 1), "inches\\n"))
cat(paste("  ✓ Bar width proportion:", bar_width_proportion, "\\n"))
cat(paste("  ✓ Font size:", round(base_font_size, 1), "pt\\n"))
cat("\\n========================================\\n")
'''

    return script


# ============================================================================
# MAIN EXECUTION BLOCK
# ============================================================================

def parse_thresholds(threshold_str: str) -> Dict[int, int]:
    """
    Parse threshold string into dictionary.
    
    Parameters:
    -----------
    threshold_str : str
        Comma-separated thresholds (e.g., "12,6,5,4,4,4")
        
    Returns:
    --------
    Dict[int, int]
        Dictionary mapping motif length to threshold
    """
    try:
        thresholds = [int(x.strip()) for x in threshold_str.split(',')]
        if len(thresholds) != 6:
            raise ValueError("Must provide exactly 6 thresholds")
        return {i+1: thresholds[i] for i in range(6)}
    except ValueError as e:
        print(f"Error parsing thresholds: {e}")
        print("Using default thresholds.")
        return DEFAULT_THRESHOLDS

def main():
    """Main function to handle command-line arguments and run analysis."""
    parser = argparse.ArgumentParser(
        description="CGAS Module 12: Comprehensive Chloroplast SSR Analysis",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python cgas_module12.py
  python cgas_module12.py -i genbank_files/
  python cgas_module12.py -i data/ -o results/
  python cgas_module12.py -t 12,6,5,4,4,4
        """
    )
    
    parser.add_argument(
        '-i', '--input',
        default='.',
        help='Input folder with GenBank files (default: current directory)'
    )
    
    parser.add_argument(
        '-o', '--output',
        default=OUTPUT_FOLDER,
        help=f'Output folder name (default: {OUTPUT_FOLDER})'
    )
    
    parser.add_argument(
        '-t', '--thresholds',
        type=str,
        help='Comma-separated thresholds for mono-to-hexa repeats (e.g., "12,6,5,4,4,4")'
    )
    
    args = parser.parse_args()
    
    # Validate input folder
    if not os.path.isdir(args.input):
        print(f"\n❌ ERROR: Input folder not found: {args.input}")
        sys.exit(1)
    
    # Create analyzer
    analyzer = ChloroplastSSRAnalyzer(args.input, args.output)
    
    # Set custom thresholds if provided
    if args.thresholds:
        analyzer.thresholds = parse_thresholds(args.thresholds)
        print(f"\nUsing custom thresholds: {analyzer.thresholds}")
    else:
        print(f"\nUsing default thresholds: {analyzer.thresholds}")
    
    # Run analysis
    analyzer.run_analysis()

if __name__ == "__main__":
    main()
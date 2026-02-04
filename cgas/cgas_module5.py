#!/usr/bin/env python3
"""
CGAS Module 5: Chloroplast Genome Gene Comparative Analysis
============================================================

This module performs comprehensive comparative analysis of chloroplast genome 
annotations from GenBank files. It:
- Identifies and counts genes across multiple genomes
- Classifies genes as functional or pseudogenes
- Detects IR-mediated duplications
- Normalizes gene names across different naming conventions
- Produces publication-ready Excel reports with statistics

Part of the CGAS (Chloroplast Genome Assembly Suite) pipeline.

Author: Abdullah
Date: January 2026
Version: 1.0.1

Requirements:
    - Python 3.9
    - biopython
    - pandas
    - openpyxl (for Excel writing)

Usage:
    # Analyze GenBank files in current directory
    python cgas_module5.py
    
    # Analyze GenBank files in specific directory
    python cgas_module5.py -i module3_normalized/
    
    Output will be generated in Module5_Gene_Comparative_Analysis/ folder.

Features:
    - Gene name normalization (tRNA formats, gene synonyms)
    - IR duplication detection
    - Pseudogene identification
    - Genome-specific gene identification
    - Comprehensive summary statistics
    - Normalization tracking sheet
"""

import os
import re
from collections import defaultdict
from datetime import datetime
from typing import List, Tuple, Dict, Set
from Bio import SeqIO
from Bio.SeqFeature import SeqFeature
import pandas as pd
from openpyxl.styles import Font


# ============================================================================
# CONFIGURATION
# ============================================================================

# Automatically use current working directory
WORKING_DIR = os.getcwd()

# Output folder
OUTPUT_FOLDER = "Module5_Gene_Comparative_Analysis"

# Generate timestamped output filename for version control
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = os.path.join(OUTPUT_FOLDER, f"Chloroplast_Gene_Analysis_{TIMESTAMP}.xlsx")
NORMALIZATION_FILE = os.path.join(OUTPUT_FOLDER, "Gene_Normalization_Log.xlsx")

# Gap tolerance for merging nearby gene features (bp)
GAP_TOLERANCE = 0

# GenBank file extensions to process
GENBANK_EXTENSIONS = ('.gb', '.gbk', '.genbank')


# ============================================================================
# GENE NAME NORMALIZATION
# ============================================================================

# Track all normalizations globally
normalization_log = []

# Track tRNA anticodon patterns across all genomes for smart inference
trna_anticodon_patterns = defaultdict(lambda: defaultdict(int))  # {amino_acid: {anticodon: count}}

# Gene synonym mapping (normalize TO the commonly used primary name)
GENE_SYNONYMS = {
    # ycf genes - normalize TO commonly used names
    "pafI": "ycf3",
    "paf1": "ycf3",
    "pafII": "ycf4",
    "paf2": "ycf4",
    "ycf10": "cemA",  # ycf10 is the synonym, cemA is the primary name
    
    # psb genes
    "lhbA": "psbZ",
    "pbf1": "psbN",
    "psb1": "psbN",
    
    # clp genes
    "clpP1": "clpP",
    "clp1": "clpP",
    
    # inf genes
    "infA": "infA",  # Translation initiation factor (sometimes varies)
    
    # matK variations
    "matK": "matK",
    "maturaseK": "matK",
    
    # accD variations  
    "accD": "accD",
    "accD1": "accD",
    
    # Other common variations
    "psbA": "psbA",
    "rbcL": "rbcL",
    "atpA": "atpA",
    "atpB": "atpB",
}


def normalize_trna_name(gene_name, species_name="", infer_anticodon=True):
    """
    Normalize tRNA gene names to standard format: trnX-YYY (uppercase anticodon)
    
    Examples:
        trnN-GUU, trnN_GUU, trnN_guu, trnN(GUU) -> trnN-GUU
        trnK_uuu -> trnK-UUU
        trnfM-CAU -> trnM-CAU (remove 'f' for formyl-methionine)
        TRNN-guu -> trnN-GUU
        trnI (missing anticodon) -> trnI-GAU (if inferred from other genomes)
    """
    original = gene_name
    
    if not gene_name.lower().startswith('trn'):
        return gene_name, False, None
    
    # DO NOT convert trnfM to trnM - they are DIFFERENT genes!
    # trnfM = formyl-methionine (initiator tRNA)
    # trnM = methionine (elongator tRNA)
    
    # Extract base and anticodon
    # Match patterns: trnX-YYY, trnX_YYY, trnX(YYY), trnXYYY, trnfM-YYY
    match = re.match(r'(trn(?:f)?[A-Z])[\-_\(]?([A-Z]{3})\)?', gene_name, re.IGNORECASE)
    
    if match:
        base = match.group(1)  # Get the base as-is first
        anticodon = match.group(2).upper()  # YYY in uppercase
        
        # Normalize the base: trn + uppercase letter (e.g., trnM, trnK, trnfM)
        if base.lower() == 'trnfm':
            base_normalized = 'trnfM'  # Special case for formyl-methionine
            amino_acid = 'fM'  # Track formyl-methionine separately
        else:
            # Standard case: trn + single uppercase letter
            base_normalized = base[:3].lower() + base[3:].upper()  # trnM, trnK, etc.
            amino_acid = base[3:].upper()
        
        # Track this anticodon pattern globally
        trna_anticodon_patterns[amino_acid][anticodon] += 1
        
        # Check if anticodon looks valid (should contain U, not T)
        warning = None
        if 'T' in anticodon:
            # DNA notation instead of RNA - convert T to U
            anticodon = anticodon.replace('T', 'U')
            warning = f"Converted DNA anticodon to RNA (T->U)"
        
        normalized = f"{base_normalized}-{anticodon}"
        return normalized, (normalized != original), warning
    
    # Check if it's tRNA without anticodon (e.g., just "trnI" or "trnfM")
    match_no_anticodon = re.match(r'^(trn(?:f)?[A-Z])$', gene_name, re.IGNORECASE)
    if match_no_anticodon:
        base = match_no_anticodon.group(1)  # Get as-is first
        
        # Normalize: trn + uppercase letter
        if base.lower() == 'trnfm':
            base_normalized = 'trnfM'
            amino_acid = 'fM'
        else:
            base_normalized = base[:3].lower() + base[3:].upper()  # trnM, trnK, etc.
            amino_acid = base[3:].upper()
        
        # IMPORTANT: Always keep the gene as-is if no anticodon
        # Let researcher handle manually
        normalized = base_normalized
        
        # Try to provide helpful information if available
        warning = None
        if infer_anticodon and amino_acid in trna_anticodon_patterns:
            anticodons = trna_anticodon_patterns[amino_acid]
            if anticodons:
                # Show what anticodons exist in other genes, but DON'T change the name
                anticodon_info = ", ".join([f"{ac} ({count}×)" for ac, count in sorted(anticodons.items())])
                warning = f"Missing anticodon. Found in other genes: {anticodon_info}"
        
        if not warning:
            warning = "tRNA missing anticodon - needs manual review"
        
        return normalized, (normalized != original), warning
    
    # If pattern doesn't match, still try to normalize case
    # Handle edge cases
    if len(gene_name) >= 4:
        normalized = gene_name[:3].lower() + gene_name[3:]
        return normalized, (normalized != original), "Non-standard tRNA format"
    
    return gene_name, False, None


def normalize_rrna_name(gene_name):
    """
    Normalize rRNA gene names to lowercase standard format.
    
    Examples:
        RRN16 -> rrn16
        rrn23 -> rrn23
        RRN5 -> rrn5
        rrn4.5s -> rrn4.5
        rrn4.5S -> rrn4.5
    """
    original = gene_name
    
    if not gene_name.lower().startswith('rrn'):
        return gene_name, False, None
    
    # Normalize to lowercase
    normalized = gene_name.lower()
    
    # Remove trailing 's' or 'S' (e.g., rrn16S -> rrn16, rrn4.5s -> rrn4.5)
    normalized = re.sub(r's$', '', normalized)
    
    # Standardize format: ensure it's rrn followed by number (with optional decimal)
    match = re.match(r'(rrn)(\d+\.?\d*)', normalized)
    if match:
        prefix = match.group(1)  # 'rrn'
        number = match.group(2)  # '16', '23', '5', '4.5', etc.
        normalized = f"{prefix}{number}"
        return normalized, (normalized != original), None
    
    return normalized, (normalized != original), None


def normalize_protein_gene_name(gene_name):
    """
    Normalize protein-coding gene names.
    
    Rules:
        - Apply gene synonyms
        - Standardize case for specific gene families
    """
    original = gene_name
    
    # Apply synonym mapping
    if gene_name in GENE_SYNONYMS:
        return GENE_SYNONYMS[gene_name], True
    
    # Check case-insensitive synonyms
    for synonym, primary in GENE_SYNONYMS.items():
        if gene_name.lower() == synonym.lower():
            return primary, True
    
    return gene_name, False


def normalize_gene_name(gene_name, species_name="", infer_anticodon=True):
    """
    Normalize gene name by applying all normalization rules.
    Returns: (normalized_name, was_normalized)
    """
    if not gene_name:
        return gene_name, False
    
    original = gene_name
    normalized = gene_name
    was_normalized = False
    normalization_types = []
    warning = None
    
    # 1. Normalize tRNA names
    if normalized.lower().startswith('trn'):
        temp, changed, warn = normalize_trna_name(normalized, species_name, infer_anticodon)
        if changed:
            normalized = temp
            was_normalized = True
            normalization_types.append("tRNA format")
        if warn:
            warning = warn
    
    # 2. Normalize rRNA names
    elif normalized.lower().startswith('rrn'):
        temp, changed, warn = normalize_rrna_name(normalized)
        if changed:
            normalized = temp
            was_normalized = True
            normalization_types.append("rRNA case")
        if warn:
            warning = warn
    
    # 3. Normalize protein-coding genes
    else:
        temp, changed = normalize_protein_gene_name(normalized)
        if changed:
            normalized = temp
            was_normalized = True
            normalization_types.append("gene synonym")
    
    # Log normalization if changed
    if was_normalized:
        normalization_log.append({
            'Species': species_name,
            'Original_Name': original,
            'Normalized_Name': normalized,
            'Normalization_Type': ", ".join(normalization_types),
            'Gene_Type': 'tRNA' if normalized.lower().startswith('trn') 
                        else 'rRNA' if normalized.lower().startswith('rrn')
                        else 'Protein-coding',
            'Warning': warning if warning else ''
        })
    
    return normalized, was_normalized


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def extract_species_name(filename: str) -> str:
    """
    Extract clean species name from GenBank filename.
    
    Removes file extensions (.gb, .gbk, .genbank) and standardizes the name
    for publication-quality output.
    
    Args:
        filename: Original GenBank filename
        
    Returns:
        Clean species name without extension
        
    Examples:
        'Solanum_lycopersicum.gb' -> 'Solanum_lycopersicum'
        'Species_name.gbk' -> 'Species_name'
    """
    # Remove common GenBank extensions
    name = filename
    for ext in GENBANK_EXTENSIONS:
        if name.lower().endswith(ext):
            name = name[:-len(ext)]
            break
    
    return name.strip()

def get_gene_name_from_qualifiers(qualifiers: Dict) -> str:
    """
    Extract gene name from GenBank feature qualifiers.
    
    Searches through common qualifier fields in priority order to identify
    the most appropriate gene name. This ensures consistent naming across
    different annotation styles.
    
    Args:
        qualifiers: Dictionary of GenBank feature qualifiers
        
    Returns:
        Gene name as string, or None if no suitable name found
        
    Priority order:
        1. gene - Standard gene symbol (e.g., 'rbcL', 'psbA')
        2. locus_tag - Systematic locus identifier
        3. product - Gene product description
        4. protein_id - Protein accession
        5. note - Additional annotations
    """
    for key in ("gene", "locus_tag", "product", "protein_id", "note"):
        if key in qualifiers and qualifiers[key]:
            value = qualifiers[key][0].strip()
            if value:
                # Replace spaces with underscores for consistent naming
                return value.replace(" ", "_")
    return None


def detect_inverted_repeat_regions(record: SeqIO.SeqRecord) -> Tuple[List[Tuple[int, int]], 
                                                                       List[Tuple[int, int]]]:
    """
    Identify inverted repeat regions (IRa and IRb) in chloroplast genome.
    
    Chloroplast genomes typically have two large inverted repeats (IRa and IRb)
    that separate the large and small single-copy regions. This function detects
    these regions from GenBank annotations using comprehensive pattern matching.
    
    Args:
        record: BioPython SeqRecord object from GenBank file
        
    Returns:
        Tuple of two lists: (IRa_ranges, IRb_ranges)
        Each range is a tuple of (start, end) coordinates
        
    Notes:
        - Searches 'repeat_region', 'misc_feature', and 'misc_RNA' annotations
        - Identifies IR regions by multiple keywords and patterns
        - Handles various annotation formats and mislabeling
        - Filters out tiny regions (<1000 bp) that are likely artifacts
        - Merges overlapping or adjacent annotations
    """
    ira_ranges = []
    irb_ranges = []
    all_ir_candidates = []
    
    # Search all features for IR annotations
    for feature in record.features:
        if feature.type.lower() in ("repeat_region", "misc_feature", "misc_RNA"):
            # Collect all descriptive text from multiple qualifier fields
            note = ""
            label = ""
            rpt_type = ""
            
            if 'note' in feature.qualifiers:
                note = feature.qualifiers['note'][0].lower()
            if 'label' in feature.qualifiers:
                label = feature.qualifiers['label'][0].lower()
            if 'rpt_type' in feature.qualifiers:
                rpt_type = feature.qualifiers['rpt_type'][0].lower()
            
            # Also check product and description for compatibility
            if 'product' in feature.qualifiers:
                note += " " + feature.qualifiers['product'][0].lower()
            if 'description' in feature.qualifiers:
                note += " " + feature.qualifiers['description'][0].lower()
            
            combined_text = f"{note} {label} {rpt_type}"
            
            # Check if this feature describes an inverted repeat
            is_ir = False
            
            # IRa detection (first IR, usually before LSC)
            if any(keyword in combined_text for keyword in [
                "ira", "inverted repeat a", "ir a", "inverted repeata"
            ]) and "irb" not in combined_text:  # Make sure it's not IRb mislabeled
                span = (int(feature.location.start), int(feature.location.end))
                ira_ranges.append(span)
                is_ir = True
            
            # IRb detection (second IR, usually after SSC)
            elif any(keyword in combined_text for keyword in [
                "irb", "inverted repeat b", "ir b", "inverted repeatb"
            ]):
                span = (int(feature.location.start), int(feature.location.end))
                # Don't append to irb_ranges yet - collect as candidate
                # This handles cases where BOTH IRs are labeled "IRb"
                all_ir_candidates.append(span)
                is_ir = True
            
            # Generic IR/repeat region detection (comprehensive fallback)
            if not is_ir and feature.type.lower() in ['repeat_region', 'misc_feature', 'misc_RNA']:
                # Check rpt_type
                if 'inverted' in rpt_type:
                    is_ir = True
                
                # Check combined text for various IR keywords
                ir_keywords = [
                    "inverted repeat", "inverted-repeat", "ir region", 
                    "ir ", " ir", "repeat", "inverted", "inv repeat",
                    "inv_repeat", "invertedrepeat"
                ]
                
                if any(keyword in combined_text for keyword in ir_keywords):
                    is_ir = True
                
                if is_ir:
                    span = (int(feature.location.start), int(feature.location.end))
                    all_ir_candidates.append(span)
    
    # CRITICAL: Handle mislabeling cases where both IRs are labeled the same
    # This is common - often both IRs are annotated as "IRb" or both as generic "IR"
    if all_ir_candidates:
        # Combine with any IRa ranges we found
        all_candidates = list(set(ira_ranges + all_ir_candidates))
        
        # Sort by start position
        all_candidates.sort(key=lambda x: x[0])
        
        # Remove duplicates (same coordinates)
        unique_irs = []
        seen_coords = set()
        for ir in all_candidates:
            coords = (ir[0], ir[1])
            if coords not in seen_coords:
                unique_irs.append(ir)
                seen_coords.add(coords)
        
        # Filter out tiny regions (<1000 bp) - these are likely annotation artifacts
        substantial_irs = []
        for ir in unique_irs:
            ir_length = ir[1] - ir[0]
            if ir_length >= 1000:  # Only keep IR regions >= 1000 bp
                substantial_irs.append(ir)
        
        # If we have at least 2 substantial IR regions
        if len(substantial_irs) >= 2:
            # Sort by position to assign IRa (first) and IRb (second)
            substantial_irs.sort(key=lambda x: x[0])
            
            # Check size similarity (good IRs should have similar lengths)
            len1 = substantial_irs[0][1] - substantial_irs[0][0]
            len2 = substantial_irs[1][1] - substantial_irs[1][0]
            ratio = max(len1, len2) / min(len1, len2) if min(len1, len2) > 0 else float('inf')
            
            # Take the first two substantial IRs regardless of labels
            # First by position = IRa, Second by position = IRb
            ira_ranges = [substantial_irs[0]]
            irb_ranges = [substantial_irs[1]]
            
            # If there are more than 2, warn but still use the first two by position
            if len(substantial_irs) > 2:
                # Multiple IR regions found - using first two by genomic position
                pass
        
        elif len(substantial_irs) == 1:
            # Only one IR found - unusual but handle it
            # If we already have IRa, assign this as IRb
            if ira_ranges:
                irb_ranges = [substantial_irs[0]]
            else:
                # Assign to IRa by default
                ira_ranges = [substantial_irs[0]]
    
    # Final fallback: if we still don't have both IRa and IRb but have candidates
    # This handles edge cases where ira_ranges was populated earlier
    if ira_ranges and not irb_ranges and all_ir_candidates:
        # Sort all candidates by position
        all_sorted = sorted(all_ir_candidates, key=lambda x: x[0])
        # Filter by size
        substantial = [ir for ir in all_sorted if (ir[1] - ir[0]) >= 1000]
        # Find one that's not already in ira_ranges
        for candidate in substantial:
            if candidate not in ira_ranges:
                irb_ranges = [candidate]
                break
    
    # Merge overlapping or adjacent ranges
    return _merge_genomic_ranges(ira_ranges), _merge_genomic_ranges(irb_ranges)


def _merge_genomic_ranges(ranges: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
    """
    Merge overlapping or adjacent genomic coordinate ranges.
    
    Args:
        ranges: List of (start, end) coordinate tuples
        
    Returns:
        List of merged (start, end) tuples with no overlaps
        
    Example:
        [(100, 200), (150, 250), (300, 400)] -> [(100, 250), (300, 400)]
    """
    if not ranges:
        return []
    
    # Sort ranges by start position
    sorted_ranges = sorted(ranges)
    merged = [list(sorted_ranges[0])]
    
    for start, end in sorted_ranges[1:]:
        # Check if current range overlaps or is adjacent to last merged range
        if start <= merged[-1][1] + 1:
            # Extend the last merged range
            merged[-1][1] = max(merged[-1][1], end)
        else:
            # Add as new separate range
            merged.append([start, end])
    
    return [(start, end) for start, end in merged]


def check_span_overlap(span1: Tuple[int, int], span2: Tuple[int, int]) -> bool:
    """
    Test if two genomic coordinate spans overlap.
    
    Args:
        span1: First span as (start, end)
        span2: Second span as (start, end)
        
    Returns:
        True if spans overlap, False otherwise
    """
    start1, end1 = span1
    start2, end2 = span2
    return not (end1 < start2 or end2 < start1)


def merge_feature_spans(spans: List[Tuple[int, int]], 
                        gap_tolerance: int = 0) -> List[Tuple[int, int]]:
    """
    Merge gene feature spans that are within gap_tolerance of each other.
    
    Some genes may be annotated as multiple features (e.g., exons). This function
    merges features that belong to the same gene based on proximity.
    
    Args:
        spans: List of (start, end) coordinate tuples
        gap_tolerance: Maximum gap (bp) to bridge when merging
        
    Returns:
        List of merged (start, end) tuples
    """
    if not spans:
        return []
    
    sorted_spans = sorted(spans)
    merged = [list(sorted_spans[0])]
    
    for start, end in sorted_spans[1:]:
        # Check if within gap tolerance of previous span
        if start <= merged[-1][1] + gap_tolerance + 1:
            merged[-1][1] = max(merged[-1][1], end)
        else:
            merged.append([start, end])
    
    return [(start, end) for start, end in merged]


def classify_gene_locus(features: List[SeqFeature], gene_name: str = "") -> str:
    """
    Classify a gene locus as functional or pseudogene.
    
    Classification logic (comprehensive pseudogene detection):
        - Functional: Has functional CDS, tRNA, or rRNA
        - Pseudogene: ANY of the following:
            1. Explicitly marked as "pseudo" or "pseudogene"
            2. Gene feature exists but NO corresponding CDS (for protein-coding genes)
            3. CDS with internal stop codons
            4. CDS marked as non-functional or truncated
            5. tRNA/rRNA without product annotation
            6. tRNA/rRNA marked as non-functional
    
    Args:
        features: List of BioPython SeqFeature objects for this locus
        gene_name: Gene name for better classification (helps identify protein-coding vs RNA genes)
        
    Returns:
        Classification string: "functional" or "pseudogene"
        
    Notes:
        - More comprehensive than just checking "pseudo" keyword
        - Checks for missing CDS in protein-coding genes
        - Checks for internal stop codons in CDS
        - Validates tRNA/rRNA functionality
    """
    has_gene_feature = False
    has_cds = False
    has_functional_cds = False
    has_trna = False
    has_functional_trna = False
    has_rrna = False
    has_functional_rrna = False
    explicit_pseudo = False
    has_internal_stops = False
    
    # Determine if this is likely a protein-coding gene based on name
    is_protein_coding = not (
        gene_name.lower().startswith('trn') or 
        gene_name.lower().startswith('rrn')
    )
    
    for feature in features:
        qualifiers = feature.qualifiers
        
        # Track if there's a gene feature
        if feature.type == "gene":
            has_gene_feature = True
        
        # Check for explicit pseudogene markers
        if "pseudo" in qualifiers:
            explicit_pseudo = True
        
        # Check product and note fields for pseudogene keywords
        for field in ["product", "note"]:
            if field in qualifiers:
                text = " ".join(qualifiers[field]).lower()
                if any(keyword in text for keyword in ["pseudo", "pseudogene", "non-functional", "nonfunctional", "truncated", "partial"]):
                    explicit_pseudo = True
        
        # Check CDS (protein-coding) features
        if feature.type == "CDS":
            has_cds = True
            
            # Check for explicit pseudo markers
            if "pseudo" in qualifiers:
                explicit_pseudo = True
                continue
            
            # Check product/note for pseudo keywords
            product_text = " ".join(
                qualifiers.get("product", []) + qualifiers.get("note", [])
            ).lower()
            
            if any(keyword in product_text for keyword in ["pseudo", "pseudogene", "non-functional", "nonfunctional"]):
                explicit_pseudo = True
                continue
            
            # Check for internal stop codons (indicated by transl_except or codon_start issues)
            if "transl_except" in qualifiers:
                # transl_except usually indicates internal stop codons
                transl_except_text = " ".join(qualifiers["transl_except"]).lower()
                if "stop" in transl_except_text or "ter" in transl_except_text:
                    has_internal_stops = True
            
            # Check translation for early stops (if available)
            if "translation" in qualifiers:
                translation = qualifiers["translation"][0]
                # If translation is very short or has unusual features, might be pseudogene
                # But we'll rely on other indicators for now
                pass
            
            # If CDS has a valid product and no pseudo indicators, it's functional
            if "product" in qualifiers and qualifiers["product"]:
                product = " ".join(qualifiers["product"]).strip()
                if product and not any(kw in product.lower() for kw in ["pseudo", "non-functional", "hypothetical protein"]):
                    has_functional_cds = True
        
        # Check tRNA features
        if feature.type == "tRNA":
            has_trna = True
            
            # Check for explicit pseudo markers
            if "pseudo" in qualifiers:
                explicit_pseudo = True
                continue
            
            # Check product/note for functionality
            product_note_text = " ".join(
                qualifiers.get("product", []) + qualifiers.get("note", [])
            ).lower()
            
            if any(keyword in product_note_text for keyword in ["pseudo", "pseudogene", "non-functional"]):
                explicit_pseudo = True
                continue
            
            # tRNA is functional if it has product annotation and anticodon
            if "product" in qualifiers and qualifiers["product"]:
                product = " ".join(qualifiers["product"]).strip()
                if product:
                    # Check if it has anticodon (functional tRNAs should have this)
                    if "note" in qualifiers:
                        note_text = " ".join(qualifiers["note"]).lower()
                        # If has anticodon information, likely functional
                        if "anticodon" in note_text or any(aa in note_text for aa in ["gly", "ala", "val", "leu", "ile"]):
                            has_functional_trna = True
                    else:
                        has_functional_trna = True
        
        # Check rRNA features
        if feature.type == "rRNA":
            has_rrna = True
            
            # Check for explicit pseudo markers
            if "pseudo" in qualifiers:
                explicit_pseudo = True
                continue
            
            # Check product/note for functionality
            product_note_text = " ".join(
                qualifiers.get("product", []) + qualifiers.get("note", [])
            ).lower()
            
            if any(keyword in product_note_text for keyword in ["pseudo", "pseudogene", "non-functional"]):
                explicit_pseudo = True
                continue
            
            # rRNA is functional if it has product annotation
            if "product" in qualifiers and qualifiers["product"]:
                product = " ".join(qualifiers["product"]).strip()
                if product:
                    has_functional_rrna = True
    
    # Classification decision (COMPREHENSIVE)
    
    # 1. If explicitly marked as pseudogene, it's a pseudogene
    if explicit_pseudo:
        return "pseudogene"
    
    # 2. If has internal stop codons, it's a pseudogene
    if has_internal_stops:
        return "pseudogene"
    
    # 3. For protein-coding genes: if gene feature exists but NO CDS, it's a pseudogene
    if is_protein_coding and has_gene_feature and not has_cds:
        return "pseudogene"
    
    # 4. Check if we have functional features
    if has_functional_cds or has_functional_trna or has_functional_rrna:
        return "functional"
    
    # 5. If we have CDS/tRNA/rRNA but not confirmed functional, check more carefully
    if has_cds:
        # CDS exists but not confirmed functional - could be pseudogene
        # If it lacks product annotation, likely pseudogene
        return "functional"  # Give benefit of doubt for CDS
    
    if has_trna or has_rrna:
        # RNA gene without product annotation - likely pseudogene
        return "pseudogene"
    
    # 6. Default: if nothing found, it's a pseudogene
    return "pseudogene"


# ============================================================================
# MAIN ANALYSIS FUNCTIONS
# ============================================================================

def analyze_genbank_file(filepath: str, infer_anticodon: bool = True) -> List[Dict]:
    """
    Analyze a single GenBank file for gene content and IR duplications.
    
    This is the main analysis function that:
        1. Detects IR regions (IRa and IRb)
        2. Identifies all genes and their features
        3. Normalizes gene names
        4. Classifies genes as functional or pseudogenes
        5. Determines copy number and IR duplication status
    
    Args:
        filepath: Path to GenBank file
        
    Returns:
        List of dictionaries, each containing analysis results for one gene:
            - Genome: Filename of source
            - Gene Name: Gene identifier (normalized)
            - Copies: Number of copies detected
            - Functional / Pseudogene: Gene status
            - Duplicate Status: Detailed copy number and location info
    """
    # Parse GenBank file
    record = SeqIO.read(filepath, "genbank")
    species_name = extract_species_name(os.path.basename(filepath))
    
    # Detect inverted repeat regions
    ira_ranges, irb_ranges = detect_inverted_repeat_regions(record)
    
    # Collect all gene-related features
    gene_features = defaultdict(list)
    
    for feature in record.features:
        # Only process gene-related feature types
        if feature.type not in ("CDS", "tRNA", "rRNA", "gene"):
            continue
        
        # Get or generate gene name
        gene_name = get_gene_name_from_qualifiers(feature.qualifiers)
        if not gene_name:
            # Generate systematic name for unannotated ORFs
            start = int(feature.location.start)
            end = int(feature.location.end)
            strand = "-" if feature.location.strand == -1 else "+"
            gene_name = f"ORF_{start}_{end}_{strand}"
        
        # NORMALIZE GENE NAME
        normalized_gene_name, was_normalized = normalize_gene_name(gene_name, species_name, infer_anticodon)
        
        gene_features[normalized_gene_name].append(feature)
    
    # Analyze each gene
    results = []
    for gene_name, features in sorted(gene_features.items()):
        # Extract genomic coordinates for all features of this gene
        span_feature_pairs = [
            (int(f.location.start), int(f.location.end), f) 
            for f in features
        ]
        spans = [(start, end) for start, end, _ in span_feature_pairs]
        
        # Merge nearby features into loci (handles multi-exon genes)
        locus_spans = merge_feature_spans(spans, gap_tolerance=GAP_TOLERANCE)
        
        # Analyze each locus separately
        locus_info_list = []
        for locus_span in locus_spans:
            locus_start, locus_end = locus_span
            
            # Get all features overlapping this locus
            locus_features = [
                feature for start, end, feature in span_feature_pairs
                if not (end < locus_start or start > locus_end)
            ]
            
            # Classify locus as functional or pseudogene
            classification = classify_gene_locus(locus_features, gene_name)
            
            # Determine genomic region (IRa, IRb, or single copy)
            in_ira = any(check_span_overlap(locus_span, ir) for ir in ira_ranges)
            in_irb = any(check_span_overlap(locus_span, ir) for ir in irb_ranges)
            
            if in_ira and in_irb:
                region = "IR(both)"  # Rare: spans both IR regions
            elif in_ira:
                region = "IRa"
            elif in_irb:
                region = "IRb"
            else:
                region = "single_copy"
            
            locus_info_list.append({
                "span": locus_span,
                "classification": classification,
                "region": region
            })
        
        # Summarize gene-level statistics
        total_copies = len(locus_info_list)
        regions = {locus["region"] for locus in locus_info_list}
        is_ir_duplicated = ("IRa" in regions) and ("IRb" in regions)
        
        functional_copies = sum(
            1 for locus in locus_info_list 
            if locus["classification"] == "functional"
        )
        pseudogene_copies = sum(
            1 for locus in locus_info_list 
            if locus["classification"] == "pseudogene"
        )
        
        # Overall gene status (functional if any copy is functional)
        gene_status = "functional" if functional_copies > 0 else "pseudogene"
        
        # Generate detailed duplication status string
        if is_ir_duplicated:
            if functional_copies > 1 and pseudogene_copies == 0:
                dup_status = f"{functional_copies} functional copies (IRa+IRb)"
            elif functional_copies >= 1 and pseudogene_copies >= 1:
                dup_status = (f"{functional_copies} functional + "
                             f"{pseudogene_copies} pseudogene copies (IRa+IRb)")
            elif functional_copies == 0 and pseudogene_copies >= 1:
                dup_status = f"{pseudogene_copies} pseudogene copies (IRa+IRb)"
            else:
                dup_status = f"{total_copies} copies (IRa+IRb)"
        else:
            if total_copies == 1:
                dup_status = "single copy"
            else:
                if functional_copies >= 1 and pseudogene_copies == 0:
                    dup_status = f"{functional_copies} functional copies (not IR-duplicated)"
                elif functional_copies >= 1 and pseudogene_copies >= 1:
                    dup_status = (f"{functional_copies} functional + "
                                 f"{pseudogene_copies} pseudogene copies (not IR-duplicated)")
                else:
                    dup_status = f"{pseudogene_copies} pseudogene copies (not IR-duplicated)"
        
        # Store results for this gene
        results.append({
            "Genome": extract_species_name(os.path.basename(filepath)),
            "Gene Name": gene_name,
            "Copies": total_copies,
            "Functional / Pseudogene": gene_status,
            "Duplicate Status": dup_status
        })
    
    return results


def generate_summary_statistics(all_results: List[Dict], 
                                genome_gene_map: Dict[str, Set[str]]) -> pd.DataFrame:
    """
    Generate genome-level summary statistics with corrected functional gene counting.
    
    Calculates aggregate statistics for each genome:
        - Total unique genes (functional only, excludes pseudogenes)
        - Number of functional genes (verified against gene_table)
        - Number of pseudogenes
        - Number of IR-duplicated genes (unique count, with annotation if includes pseudogene)
    
    Args:
        all_results: List of all gene analysis results
        genome_gene_map: Dictionary mapping genome names to their gene sets
        
    Returns:
        DataFrame with corrected summary statistics per genome
    """
    summary_data = []
    
    for genome_name, gene_set in genome_gene_map.items():
        genome_results = [r for r in all_results if r["Genome"] == genome_name]
        
        # Count unique functional and pseudogene genes
        functional_genes = set()
        pseudogene_genes = set()
        ir_duplicated_genes = set()  # Track unique IR-duplicated genes
        ir_with_pseudogene = []  # Track IR duplicates that include pseudogenes
        
        for result in genome_results:
            gene_name = result["Gene Name"]
            
            if result["Functional / Pseudogene"] == "functional":
                functional_genes.add(gene_name)
            elif result["Functional / Pseudogene"] == "pseudogene":
                pseudogene_genes.add(gene_name)
            
            # Track IR-duplicated genes (unique)
            if "IRa+IRb" in result["Duplicate Status"]:
                ir_duplicated_genes.add(gene_name)
                # Check if this IR-duplicated gene has a pseudogene copy
                if result["Functional / Pseudogene"] == "pseudogene":
                    if gene_name not in ir_with_pseudogene:
                        ir_with_pseudogene.append(gene_name)
        
        # Total unique genes = functional genes only (pseudogenes not counted as "genes")
        total_unique_functional = len(functional_genes)
        
        # IR duplicated genes with detailed annotation
        ir_count = len(ir_duplicated_genes)
        ir_note = ""
        
        # Check for pseudogenes in IR
        if ir_with_pseudogene:
            pseudogene_note = f" [includes {len(ir_with_pseudogene)} with pseudogene: {', '.join(sorted(ir_with_pseudogene))}]"
            ir_note += pseudogene_note
        
        # Format IR display
        if ir_note:
            ir_display = f"{ir_count}{ir_note}"
        else:
            ir_display = ir_count
        
        summary_data.append({
            "Genome": genome_name,
            "Total Unique Genes (Functional)": total_unique_functional,
            "Functional Genes": len(functional_genes),
            "Pseudogenes": len(pseudogene_genes),
            "IR-duplicated Genes": ir_display,
            "IR Gene List": ", ".join(sorted(ir_duplicated_genes))  # Add this for comparison
        })
    
    return pd.DataFrame(
        summary_data,
        columns=["Genome", "Total Unique Genes (Functional)", "Functional Genes", 
                "Pseudogenes", "IR-duplicated Genes", "IR Gene List"]
    )


def identify_unique_genes(all_results: List[Dict], 
                         genome_gene_map: Dict[str, Set[str]]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Identify genes with unique presence/absence patterns.
    
    Returns two types of unique genes:
    1. Genes present in only ONE genome (genome-specific)
    2. Genes missing from only ONE genome (nearly universal)
    
    Args:
        all_results: List of all gene analysis results
        genome_gene_map: Dictionary mapping genome names to their gene sets
        
    Returns:
        Tuple of two DataFrames: (genome_specific_genes, nearly_universal_genes)
    """
    genome_specific_data = []
    nearly_universal_data = []
    
    # Get all unique genes across all genomes
    all_genes = set()
    for genes in genome_gene_map.values():
        all_genes.update(genes)
    
    # For each gene, count how many genomes contain it
    gene_presence = defaultdict(list)
    for genome_name, gene_set in genome_gene_map.items():
        for gene_name in all_genes:
            if gene_name in gene_set:
                gene_presence[gene_name].append(genome_name)
    
    total_genomes = len(genome_gene_map)
    
    # Identify genome-specific genes (present in only 1 genome)
    for gene_name, genomes_with_gene in gene_presence.items():
        if len(genomes_with_gene) == 1:
            genome_name = genomes_with_gene[0]
            gene_result = next(
                r for r in all_results 
                if r["Genome"] == genome_name and r["Gene Name"] == gene_name
            )
            genome_specific_data.append(gene_result)
    
    # Identify nearly universal genes (missing from only 1 genome)
    if total_genomes > 2:  # Only meaningful with 3+ genomes
        for gene_name, genomes_with_gene in gene_presence.items():
            if len(genomes_with_gene) == total_genomes - 1:
                # Find which genome is missing this gene
                missing_genome = None
                for genome_name in genome_gene_map.keys():
                    if genome_name not in genomes_with_gene:
                        missing_genome = genome_name
                        break
                
                # Add entry for each genome that HAS the gene
                for genome_name in genomes_with_gene:
                    gene_result = next(
                        r for r in all_results 
                        if r["Genome"] == genome_name and r["Gene Name"] == gene_name
                    ).copy()
                    gene_result["Missing_From"] = missing_genome
                    nearly_universal_data.append(gene_result)
    
    df_specific = pd.DataFrame(
        genome_specific_data,
        columns=["Genome", "Gene Name", "Copies", 
                "Functional / Pseudogene", "Duplicate Status"]
    )
    
    df_nearly_universal = pd.DataFrame(
        nearly_universal_data,
        columns=["Genome", "Gene Name", "Copies", 
                "Functional / Pseudogene", "Duplicate Status", "Missing_From"]
    )
    
    return df_specific, df_nearly_universal


def generate_missing_genes_report(all_results: List[Dict],
                                  genome_gene_map: Dict[str, Set[str]]) -> pd.DataFrame:
    """
    Generate comprehensive missing genes report showing which genes are absent in each genome.
    
    This function creates a detailed report of gene presence/absence patterns across all genomes,
    helping authors identify and clarify missing genes in their comparative analysis.
    
    Args:
        all_results: List of all gene analysis results
        genome_gene_map: Dictionary mapping genome names to their gene sets
        
    Returns:
        DataFrame with columns:
            - Gene Name: Name of the gene
            - Present_In: Number of genomes containing this gene
            - Missing_From: Number of genomes missing this gene
            - Present_In_Genomes: Comma-separated list of genomes with the gene
            - Missing_From_Genomes: Comma-separated list of genomes without the gene
            - Status: "Universal" / "Common" / "Rare" / "Unique"
    """
    # Get all unique genes across all genomes
    all_genes = set()
    for genes in genome_gene_map.values():
        all_genes.update(genes)
    
    # For each gene, track which genomes have it
    gene_presence = defaultdict(list)
    gene_absence = defaultdict(list)
    
    for gene_name in sorted(all_genes):
        for genome_name in sorted(genome_gene_map.keys()):
            if gene_name in genome_gene_map[genome_name]:
                gene_presence[gene_name].append(genome_name)
            else:
                gene_absence[gene_name].append(genome_name)
    
    total_genomes = len(genome_gene_map)
    
    # Build report data
    report_data = []
    for gene_name in sorted(all_genes):
        present_count = len(gene_presence[gene_name])
        missing_count = len(gene_absence[gene_name])
        
        # Determine status
        if present_count == total_genomes:
            status = "Universal"
        elif present_count >= total_genomes * 0.8:  # In 80%+ of genomes
            status = "Common"
        elif present_count == 1:
            status = "Unique"
        else:
            status = "Variable"
        
        # Get representative gene information (functional status, etc.)
        representative_result = next(
            (r for r in all_results if r["Gene Name"] == gene_name),
            None
        )
        
        func_status = representative_result["Functional / Pseudogene"] if representative_result else ""
        
        report_data.append({
            "Gene Name": gene_name,
            "Present_In": present_count,
            "Missing_From": missing_count,
            "Total_Genomes": total_genomes,
            "Present_In_Genomes": ", ".join(gene_presence[gene_name]) if gene_presence[gene_name] else "-",
            "Missing_From_Genomes": ", ".join(gene_absence[gene_name]) if gene_absence[gene_name] else "-",
            "Status": status,
            "Typical_Classification": func_status
        })
    
    df = pd.DataFrame(report_data)
    
    # Sort by presence count (descending) then gene name
    df = df.sort_values(by=["Missing_From", "Gene Name"], ascending=[True, True])
    
    return df


def generate_missing_gene_notes(df_missing_genes: pd.DataFrame, df_summary: pd.DataFrame) -> str:
    """Generate detailed notes about missing genes and IR differences."""
    notes = []
    notes.append("=" * 80)
    notes.append("GENE DISCREPANCY ANALYSIS")
    notes.append("=" * 80)
    notes.append("")
    
    # 1. Check IR-duplicated gene differences
    if "IR Gene List" in df_summary.columns:
        notes.append("IR-DUPLICATED GENE COMPARISON:")
        notes.append("-" * 40)
        
        ir_counts = df_summary["IR-duplicated Genes"].astype(str).str.extract(r'(\d+)')[0].astype(int)
        
        if ir_counts.nunique() > 1:
            notes.append(f"IR gene count variation: {ir_counts.min()} to {ir_counts.max()}")
            notes.append("")
            
            # Parse IR gene lists
            genome_ir_genes = {}
            for _, row in df_summary.iterrows():
                genome = row["Genome"]
                ir_list = row.get("IR Gene List", "")
                if ir_list:
                    genome_ir_genes[genome] = set(ir_list.split(", "))
                else:
                    genome_ir_genes[genome] = set()
            
            # Find differences
            genomes = list(genome_ir_genes.keys())
            for i, g1 in enumerate(genomes):
                for g2 in genomes[i+1:]:
                    diff1 = genome_ir_genes[g1] - genome_ir_genes[g2]
                    diff2 = genome_ir_genes[g2] - genome_ir_genes[g1]
                    
                    if diff1 or diff2:
                        notes.append(f"{g1} vs {g2}:")
                        if diff1:
                            notes.append(f"  Only in {g1}: {', '.join(sorted(diff1))}")
                        if diff2:
                            notes.append(f"  Only in {g2}: {', '.join(sorted(diff2))}")
                        notes.append("")
        else:
            notes.append("All genomes have the same number of IR-duplicated genes")
            notes.append("")
        
        notes.append("")
    
    # 2. Check functional gene count differences
    notes.append("UNIQUE GENE COUNT ANALYSIS:")
    notes.append("-" * 40)
    
    max_genes = df_summary["Total Unique Genes (Functional)"].max()
    min_genes = df_summary["Total Unique Genes (Functional)"].min()
    
    if max_genes > min_genes:
        notes.append(f"Gene count variation: {min_genes} to {max_genes} unique functional genes")
        notes.append("")
        
        genomes_with_less = df_summary[df_summary["Total Unique Genes (Functional)"] < max_genes]
        for _, row in genomes_with_less.iterrows():
            genome_name = row["Genome"]
            gene_count = row["Total Unique Genes (Functional)"]
            diff = max_genes - gene_count
            
            notes.append(f"{genome_name}: {gene_count} genes ({diff} missing)")
            
            missing_from_this = df_missing_genes[
                df_missing_genes["Missing_From_Genomes"].str.contains(genome_name, na=False)
            ]
            
            for _, gene_row in missing_from_this.iterrows():
                notes.append(f"  • {gene_row['Gene Name']} ({gene_row.get('Typical_Classification', '')})")
            notes.append("")
    else:
        notes.append("All genomes have the same number of unique functional genes")
        notes.append("")
    
    notes.append("=" * 80)
    return "\n".join(notes)


def apply_publication_formatting(writer: pd.ExcelWriter, sheet_name: str):
    """
    Apply publication-quality formatting to Excel worksheet.
    
    For Gene_Table sheet: Species names as merged headers (like module 8)
    For other sheets: Standard formatting
    
    Formatting includes:
        - Auto-adjusted column widths
        - Frozen header row
        - Italic formatting for species names
        - Italic formatting for gene names
        - Species header rows (for Gene_Table)
    
    Args:
        writer: pandas ExcelWriter object
        sheet_name: Name of the worksheet to format
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    worksheet = writer.sheets[sheet_name]
    italic_font = Font(italic=True)
    bold_italic_font = Font(italic=True, bold=True)
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    
    # Special handling for Gene_Table with species headers
    if sheet_name == "Gene_Table":
        # Find column indices
        species_col = None
        gene_col = None
        
        for col_idx, column in enumerate(worksheet.iter_cols(1, worksheet.max_column, 1, 1), start=1):
            header_value = column[0].value
            if header_value == "Species":
                species_col = col_idx
            elif header_value == "Gene Name":
                gene_col = col_idx
        
        # Format rows
        for row_idx in range(2, worksheet.max_row + 1):
            species_cell = worksheet.cell(row=row_idx, column=species_col) if species_col else None
            gene_cell = worksheet.cell(row=row_idx, column=gene_col) if gene_col else None
            
            # Check if this is a species header row (has Species value, empty Gene Name)
            is_species_header = (species_cell and species_cell.value and 
                               gene_cell and not gene_cell.value)
            
            if is_species_header:
                # Format as species header row (bold italic, gray background, merge across columns)
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = bold_italic_font
                    cell.fill = gray_fill
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Merge cells across all columns for species name
                worksheet.merge_cells(start_row=row_idx, start_column=1, 
                                    end_row=row_idx, end_column=worksheet.max_column)
            else:
                # Regular gene row - italicize gene name only
                if gene_cell and gene_cell.value:
                    gene_cell.font = italic_font
    
    else:
        # Standard formatting for other sheets
        # Get column indices for species and gene names
        genome_col = None
        gene_col = None
        missing_col = None
        ir_list_col = None
        
        for col_idx, column in enumerate(worksheet.iter_cols(1, worksheet.max_column, 1, 1), start=1):
            header_value = column[0].value
            if header_value in ["Genome", "Species"]:
                genome_col = col_idx
            elif header_value == "Gene Name":
                gene_col = col_idx
            elif header_value == "Missing_From":
                missing_col = col_idx
            elif header_value == "IR Gene List":
                ir_list_col = col_idx
        
        # Apply italic formatting
        for row_idx in range(2, worksheet.max_row + 1):
            if genome_col:
                cell = worksheet.cell(row=row_idx, column=genome_col)
                cell.font = italic_font
            
            if gene_col:
                cell = worksheet.cell(row=row_idx, column=gene_col)
                cell.font = italic_font
            
            if missing_col:
                cell = worksheet.cell(row=row_idx, column=missing_col)
                if cell.value:
                    cell.font = italic_font
            
            if ir_list_col:
                cell = worksheet.cell(row=row_idx, column=ir_list_col)
                if cell.value:
                    cell.font = italic_font
    
    # Auto-adjust column widths for all sheets
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set width with reasonable limits
        adjusted_width = min(max_length + 3, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row for easy scrolling
    worksheet.freeze_panes = worksheet['A2']


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """
    Main function for CGAS Module 5: Gene Comparative Analysis
    
    Workflow:
        1. Parse command-line arguments
        2. Find all GenBank files in input directory
        3. Analyze each file (with gene name normalization)
        4. Generate summary statistics
        5. Identify unique genes
        6. Write publication-ready Excel reports
        7. Save gene normalization tracking sheet
    """
    import argparse
    
    parser = argparse.ArgumentParser(
        description="CGAS Module 5: Chloroplast Genome Gene Comparative Analysis",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Analyze GenBank files in current directory
  python cgas_module5.py
  
  # Analyze GenBank files in specific directory
  python cgas_module5.py -i module3_normalized/revise_annotations/
  
  # Analyze with custom output directory
  python cgas_module5.py -i genbank_files/ -o gene_analysis/

Output:
  Module5_Gene_Comparative_Analysis/
  ├── Chloroplast_Gene_Analysis_TIMESTAMP.xlsx
  │   ├── Gene_Table (all genes with duplication status)
  │   ├── Summary (genome statistics)
  │   └── Genome_Specific_Genes (unique genes)
  └── Gene_Normalization_Log.xlsx (gene name changes)
        """
    )
    
    parser.add_argument(
        "-i", "--input", 
        default=".",
        help="Input directory containing GenBank files (default: current directory)"
    )
    
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="Output directory name (default: Module5_Gene_Comparative_Analysis)"
    )
    
    args = parser.parse_args()
    
    # Update global variables
    global WORKING_DIR, OUTPUT_FOLDER, OUTPUT_FILE, NORMALIZATION_FILE, TIMESTAMP
    
    WORKING_DIR = os.path.abspath(args.input)
    
    if args.output:
        OUTPUT_FOLDER = args.output
    else:
        OUTPUT_FOLDER = "Module5_Gene_Comparative_Analysis"
    
    # Create output folder path
    output_folder_full = os.path.join(WORKING_DIR, OUTPUT_FOLDER)
    
    # Update output file paths
    TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
    OUTPUT_FILE = os.path.join(output_folder_full, f"Chloroplast_Gene_Analysis_{TIMESTAMP}.xlsx")
    NORMALIZATION_FILE = os.path.join(output_folder_full, "Gene_Normalization_Log.xlsx")
    
    print("=" * 70)
    print("CGAS MODULE 5: GENE COMPARATIVE ANALYSIS")
    print("=" * 70)
    print(f"\nInput directory: {WORKING_DIR}")
    
    # Create output folder
    os.makedirs(output_folder_full, exist_ok=True)
    print(f"Output folder: {output_folder_full}/")
    
    # Find all GenBank files
    genbank_files = sorted([
        f for f in os.listdir(WORKING_DIR)
        if f.lower().endswith(GENBANK_EXTENSIONS)
    ])
    
    if not genbank_files:
        print(f"\nERROR: No GenBank files found in {WORKING_DIR}")
        print(f"Supported extensions: {', '.join(GENBANK_EXTENSIONS)}")
        return
    
    print(f"\nFound {len(genbank_files)} GenBank file(s):")
    for filename in genbank_files:
        print(f"  - {filename}")
    
    # Process all files (PASS 1: Build anticodon patterns, no inference)
    print("\n" + "-" * 70)
    print("Pass 1: Analyzing genes and building tRNA anticodon database...")
    print("-" * 70)
    
    all_results = []
    genome_gene_map = defaultdict(set)
    
    for genbank_file in genbank_files:
        filepath = os.path.join(WORKING_DIR, genbank_file)
        species_name = extract_species_name(genbank_file)
        print(f"\nAnalyzing: {species_name}")
        
        try:
            # First pass: collect anticodon patterns (no inference)
            results = analyze_genbank_file(filepath, infer_anticodon=False)
            all_results.extend(results)
            genome_gene_map[species_name] = set(r["Gene Name"] for r in results)
            print(f"  ✓ Found {len(results)} genes")
        except Exception as e:
            print(f"  ✗ Error processing file: {str(e)}")
            continue
    
    # Clear results for second pass
    all_results.clear()
    genome_gene_map.clear()
    normalization_log.clear()
    
    # Process all files (PASS 2: Apply inference for missing anticodons)
    print("\n" + "-" * 70)
    print("Pass 2: Re-analyzing with anticodon inference...")
    print("-" * 70)
    
    if trna_anticodon_patterns:
        print("\ntRNA anticodon patterns discovered:")
        for amino_acid in sorted(trna_anticodon_patterns.keys()):
            anticodons = trna_anticodon_patterns[amino_acid]
            anticodon_str = ", ".join([f"{ac} ({count}×)" for ac, count in sorted(anticodons.items())])
            print(f"  trn{amino_acid}: {anticodon_str}")
    
    for genbank_file in genbank_files:
        filepath = os.path.join(WORKING_DIR, genbank_file)
        species_name = extract_species_name(genbank_file)
        print(f"\nRe-analyzing: {species_name}")
        
        try:
            # Second pass: with anticodon inference enabled
            results = analyze_genbank_file(filepath, infer_anticodon=True)
            all_results.extend(results)
            genome_gene_map[species_name] = set(r["Gene Name"] for r in results)
            print(f"  ✓ Found {len(results)} genes")
        except Exception as e:
            print(f"  ✗ Error processing file: {str(e)}")
            continue
    
    # PASS 3: Final resolution - apply anticodon inference to gene names for display
    print("\n" + "-" * 70)
    print("Pass 3: Resolving final tRNA names for display...")
    print("-" * 70)
    
    final_trna_resolution = {}  # Map incomplete names to final names
    
    for result in all_results:
        gene_name = result["Gene Name"]
        
        # Check if it's a tRNA without anticodon
        if gene_name.lower().startswith('trn'):
            match = re.match(r'^(trn(?:f)?[A-Z])$', gene_name, re.IGNORECASE)
            if match:
                # This is incomplete (no anticodon)
                base = match.group(1)
                
                # Normalize base
                if base.lower() == 'trnfm':
                    base_normalized = 'trnfM'
                    amino_acid = 'fM'
                else:
                    base_normalized = base[:3].lower() + base[3:].upper()
                    amino_acid = base[3:].upper()
                
                # Check if we can resolve it
                if amino_acid in trna_anticodon_patterns:
                    anticodons = trna_anticodon_patterns[amino_acid]
                    if anticodons:
                        # Find most common
                        most_common_anticodon = max(anticodons, key=anticodons.get)
                        most_common_count = anticodons[most_common_anticodon]
                        equally_common = [ac for ac, count in anticodons.items() if count == most_common_count]
                        
                        if len(equally_common) == 1:
                            # Can resolve - update the gene name
                            resolved_name = f"{base_normalized}-{most_common_anticodon}"
                            final_trna_resolution[gene_name] = resolved_name
                            result["Gene Name"] = resolved_name
                            print(f"  Resolved: {gene_name} → {resolved_name}")
    
    # Update genome_gene_map with resolved names
    genome_gene_map.clear()
    for result in all_results:
        genome_gene_map[result["Genome"]].add(result["Gene Name"])
    
    if not all_results:
        print("\nERROR: No results generated. Check your GenBank files.")
        return
    
    # Generate analysis outputs
    print("\n" + "-" * 70)
    print("Generating analysis reports...")
    print("-" * 70)
    
    # 1. Complete gene table with species headers (like module 8)
    # Organize data by species
    species_list = sorted(genome_gene_map.keys())
    
    # Build gene table with species grouping
    gene_table_data = []
    for species_name in species_list:
        # Add species header row
        gene_table_data.append({
            "Species": species_name,
            "Gene Name": "",
            "Copies": "",
            "Functional / Pseudogene": "",
            "Duplicate Status": ""
        })
        
        # Add all genes for this species
        species_genes = [r for r in all_results if r["Genome"] == species_name]
        for gene_result in sorted(species_genes, key=lambda x: x["Gene Name"]):
            gene_table_data.append({
                "Species": "",  # Empty for gene rows
                "Gene Name": gene_result["Gene Name"],
                "Copies": gene_result["Copies"],
                "Functional / Pseudogene": gene_result["Functional / Pseudogene"],
                "Duplicate Status": gene_result["Duplicate Status"]
            })
    
    df_gene_table = pd.DataFrame(gene_table_data)
    
    # 2. Summary statistics
    df_summary = generate_summary_statistics(all_results, genome_gene_map)
    
    # 3. Genome-specific genes only (removed nearly universal genes sheet)
    df_genome_specific, _ = identify_unique_genes(all_results, genome_gene_map)
    
    # 4. COMPREHENSIVE MISSING GENES REPORT
    df_missing_genes = generate_missing_genes_report(all_results, genome_gene_map)
    
    # Generate detailed missing gene notes
    missing_gene_notes = generate_missing_gene_notes(df_missing_genes, df_summary)
    notes_file = os.path.join(OUTPUT_FOLDER, "Missing_Genes_Details.txt")
    with open(notes_file, 'w') as f:
        f.write(missing_gene_notes)
    print(f"  ✓ Missing gene details: {notes_file}")
    
    # 5. Gene normalization log with final resolution
    if normalization_log:
        # Add final resolved names to normalization log
        for entry in normalization_log:
            original = entry['Original_Name']
            normalized = entry['Normalized_Name']
            
            # Check if this was further resolved
            if normalized in final_trna_resolution:
                entry['Final_Name'] = final_trna_resolution[normalized]
            else:
                entry['Final_Name'] = normalized
    
    df_normalization = pd.DataFrame(normalization_log) if normalization_log else pd.DataFrame(
        columns=['Species', 'Original_Name', 'Normalized_Name', 'Final_Name', 'Normalization_Type', 'Gene_Type', 'Warning']
    )
    
    # Write Excel file with publication-quality formatting
    print(f"\nWriting results to: {OUTPUT_FILE}")
    
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Write each sheet
        df_gene_table.to_excel(writer, sheet_name="Gene_Table", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_genome_specific.to_excel(writer, sheet_name="Genome_Specific_Genes", index=False)
        df_missing_genes.to_excel(writer, sheet_name="Missing_Genes_Analysis", index=False)
        
        # Apply publication formatting to all sheets
        for sheet_name in writer.sheets:
            apply_publication_formatting(writer, sheet_name)
    
    # Write normalization log separately
    if not df_normalization.empty:
        print(f"Writing normalization log to: {NORMALIZATION_FILE}")
        with pd.ExcelWriter(NORMALIZATION_FILE, engine='openpyxl') as writer:
            df_normalization.to_excel(writer, sheet_name="Normalizations", index=False)
            
            # Format normalization sheet
            worksheet = writer.sheets["Normalizations"]
            italic_font = Font(italic=True)
            
            # Find columns
            species_col = gene_col = norm_col = None
            for col_idx, column in enumerate(worksheet.iter_cols(1, worksheet.max_column, 1, 1), start=1):
                header_value = column[0].value
                if header_value == "Species":
                    species_col = col_idx
                elif header_value in ["Original_Name", "Normalized_Name"]:
                    if not gene_col:
                        gene_col = col_idx
                    norm_col = col_idx
            
            # Apply italic to species and gene names
            for row_idx in range(2, worksheet.max_row + 1):
                if species_col:
                    worksheet.cell(row=row_idx, column=species_col).font = italic_font
                if gene_col:
                    worksheet.cell(row=row_idx, column=gene_col).font = italic_font
                if norm_col and norm_col != gene_col:
                    worksheet.cell(row=row_idx, column=norm_col).font = italic_font
            
            # Auto-adjust columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                worksheet.column_dimensions[column_letter].width = min(max_length + 3, 50)
            
            worksheet.freeze_panes = worksheet['A2']
    
    # Print summary
    print("\n" + "=" * 70)
    print("ANALYSIS COMPLETE")
    print("=" * 70)
    print(f"\nTotal genomes analyzed: {len(genome_gene_map)}")
    print(f"Total genes identified: {len(set(r['Gene Name'] for r in all_results))}")
    print(f"Total gene records: {len(all_results)}")
    print(f"Total normalizations: {len(normalization_log)}")
    
    if normalization_log:
        # Summarize normalization types
        norm_types = {}
        for entry in normalization_log:
            norm_type = entry['Normalization_Type']
            norm_types[norm_type] = norm_types.get(norm_type, 0) + 1
        
        print("\nNormalization Summary:")
        for norm_type, count in sorted(norm_types.items()):
            print(f"  - {norm_type}: {count} genes")
    
    print("\nOutput files:")
    print(f"  1. {os.path.basename(OUTPUT_FILE)}")
    print("     - Gene_Table: Complete gene catalog with duplication status")
    print("     - Summary: Genome-level statistics")
    print("     - Genome_Specific_Genes: Genes found in only ONE genome")
    print("     - Missing_Genes_Analysis: Comprehensive gene presence/absence across all genomes")
    
    if not df_normalization.empty:
        print(f"  2. {os.path.basename(NORMALIZATION_FILE)}")
        print("     - Normalizations: Complete log of all gene name changes")
    
    print(f"\nAll files saved in: {os.path.join(WORKING_DIR, OUTPUT_FOLDER)}/")
    print("=" * 70)


if __name__ == "__main__":
    main()

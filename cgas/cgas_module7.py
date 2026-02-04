#!/usr/bin/env python3
"""
CGAS Module 7: Comparative Chloroplast Genome Analysis
========================================

This script performs comprehensive comparative analysis of chloroplast genomes by:
1. Extracting genome structure information (LSC, SSC, IR regions)
2. Calculating genome lengths for each region
3. Computing GC content for regions and gene types
4. Generating publication-quality comparative table

Author: Abdullah
Version: 1.0.1
Date: January 2026

Dependencies:
    - biopython
    - pandas
    - openpyxl

Usage:
    Place GenBank files (.gb, .gbf, .gbk) in the working directory and run:
    python cgas_module7.py

Output:
    Module7_Comparative_Analysis/
        - Comparative_Genome_Analysis.xlsx
"""

import os
import datetime
import pandas as pd
from Bio import SeqIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONSTANTS
# ============================================================================

OUTPUT_FOLDER = "Module7_Comparative_Analysis"
OUTPUT_FILENAME = "Comparative_Genome_Analysis.xlsx"


# ============================================================================
# ANALYSIS FUNCTIONS
# ============================================================================

# Function to calculate GC content with rounding to two decimal places
def calculate_gc(sequence):
    total_bases = len(sequence)
    gc_count = sequence.count('G') + sequence.count('C')
    return round((gc_count / total_bases) * 100, 2) if total_bases > 0 else 0


def detect_structural_features(record):
    """
    Detect LSC, SSC, IRa, and IRb regions from GenBank features.
    
    Handles various annotation formats:
    - misc_feature with note containing "LSC", "SSC", "IRa", "IRb"
    - repeat_region features
    - inverted_repeat features
    - Automatic inference when only partial annotations exist
    - Handles mislabeled IRs (e.g., both labeled "IRb")
    
    Returns:
        tuple: (features_dict, warnings_list)
        - features_dict: dict with 'lsc', 'ssc', 'ira', 'irb' features (or None if not found)
        - warnings_list: list of warning messages for this genome
    """
    features = {
        'lsc': None,
        'ssc': None,
        'ira': None,
        'irb': None
    }
    
    warnings = []
    
    # Track all repeat regions for comprehensive detection
    all_ir_candidates = []
    
    for feature in record.features:
        feature_type = feature.type.lower()
        
        # Check feature qualifiers for region identification
        note = ""
        label = ""
        rpt_type = ""
        
        if 'note' in feature.qualifiers:
            note = feature.qualifiers['note'][0].lower()
        if 'label' in feature.qualifiers:
            label = feature.qualifiers['label'][0].lower()
        if 'rpt_type' in feature.qualifiers:
            rpt_type = feature.qualifiers['rpt_type'][0].lower()
        
        combined_text = f"{note} {label} {rpt_type}"
        
        # LSC detection
        if any(keyword in combined_text for keyword in [
            "large single copy", "lsc", "large single-copy"
        ]):
            if not features['lsc']:  # Take first match
                features['lsc'] = feature
                print(f"    ✓ LSC detected: {feature.location.start}-{feature.location.end}")
        
        # SSC detection
        elif any(keyword in combined_text for keyword in [
            "small single copy", "ssc", "small single-copy"
        ]):
            if not features['ssc']:
                features['ssc'] = feature
                print(f"    ✓ SSC detected: {feature.location.start}-{feature.location.end}")
        
        # IRa detection (first IR, usually before LSC)
        elif any(keyword in combined_text for keyword in [
            "ira", "inverted repeat a", "ir a", "inverted repeata"
        ]) and "irb" not in combined_text:  # Make sure it's not IRb mislabeled
            if not features['ira']:
                features['ira'] = feature
                print(f"    ✓ IRa detected: {feature.location.start}-{feature.location.end}")
        
        # IRb detection (second IR, usually after SSC)
        elif any(keyword in combined_text for keyword in [
            "irb", "inverted repeat b", "ir b", "inverted repeatb"
        ]):
            # Even if labeled IRb, we'll sort them later if there are two
            if not features['irb']:
                features['irb'] = feature
                print(f"    ✓ IRb detected: {feature.location.start}-{feature.location.end}")
            else:
                # Found a second "IRb" - this is likely a mislabeling issue
                all_ir_candidates.append(features['irb'])
                all_ir_candidates.append(feature)
                features['irb'] = None  # Will reassign below
        
        # Generic IR/repeat region detection (comprehensive)
        elif feature_type in ['repeat_region', 'misc_feature', 'misc_RNA']:
            # Check for inverted repeat indicators
            is_ir = False
            
            # Check rpt_type
            if 'inverted' in rpt_type:
                is_ir = True
            
            # Check combined text for various IR keywords
            ir_keywords = [
                "inverted repeat", "inverted-repeat", "ir region", 
                "ir ", " ir", "repeat", "inverted", "inv repeat",
                "inv_repeat", "invertedrepeat"
            ]
            
            if any(keyword in combined_text for keyword in ir_keywords):
                is_ir = True
            
            if is_ir:
                all_ir_candidates.append(feature)
    
    # If we have IR candidates but IRa/IRb not properly assigned
    if all_ir_candidates and (not features['ira'] or not features['irb']):
        # Remove any already assigned IRs from candidates
        if features['ira'] and features['ira'] not in all_ir_candidates:
            all_ir_candidates.append(features['ira'])
        if features['irb'] and features['irb'] not in all_ir_candidates:
            all_ir_candidates.append(features['irb'])
        
        # Sort by start position
        all_ir_candidates.sort(key=lambda f: f.location.start)
        
        # Remove duplicates (same coordinates)
        unique_irs = []
        seen_coords = set()
        for ir in all_ir_candidates:
            coords = (ir.location.start, ir.location.end)
            if coords not in seen_coords:
                unique_irs.append(ir)
                seen_coords.add(coords)
        
        # CRITICAL FIX: Filter out tiny regions and validate IR sizes
        # Strategy:
        # 1. Remove regions < 1000 bp (annotation artifacts)
        # 2. Check if remaining IRs have reasonable sizes
        # 3. Prefer IRs with similar lengths
        # 4. Warn if IR sizes are unusual
        
        substantial_irs = []
        tiny_regions = []
        
        for ir in unique_irs:
            ir_length = ir.location.end - ir.location.start
            if ir_length >= 1000:  # Only keep IR regions >= 1000 bp
                substantial_irs.append(ir)
            else:
                tiny_regions.append((ir.location.start+1, ir.location.end, ir_length))
        
        # Report tiny regions
        for start, end, length in tiny_regions:
            print(f"    → Ignoring tiny IR-like region: {start}-{end} ({length} bp)")
            warnings.append(f"Tiny IR region ignored: {start}-{end} ({length} bp)")
        
        # If we have at least 2 substantial IR regions after filtering
        if len(substantial_irs) >= 2:
            # Sort by length (descending) to analyze sizes
            substantial_irs_by_size = sorted(substantial_irs, key=lambda f: f.location.end - f.location.start, reverse=True)
            
            # Get the two largest
            largest_ir_len = substantial_irs_by_size[0].location.end - substantial_irs_by_size[0].location.start
            second_largest_len = substantial_irs_by_size[1].location.end - substantial_irs_by_size[1].location.start
            
            # VALIDATION: Check if IR sizes are reasonable
            size_ratio = largest_ir_len / second_largest_len if second_largest_len > 0 else 0
            
            # Case 1: IRs have similar sizes (within 2x) - NORMAL
            if size_ratio <= 2.0:
                selected_irs = substantial_irs_by_size[:2]
                print(f"    ✓ Two IRs with similar sizes detected ({largest_ir_len:,} bp and {second_largest_len:,} bp)")
            
            # Case 2: One IR is much larger than the other (> 2x but < 4x) - SUSPICIOUS
            elif size_ratio <= 4.0:
                selected_irs = substantial_irs_by_size[:2]
                warning_msg = f"IR size discrepancy: IRa={largest_ir_len:,} bp, IRb={second_largest_len:,} bp (ratio: {size_ratio:.1f}x)"
                print(f"    ⚠ Warning: {warning_msg}")
                warnings.append(warning_msg)
            
            # Case 3: One IR is MUCH larger (> 4x) - LIKELY WRONG ANNOTATION
            else:
                warning_msg = f"Unusual IR sizes: largest={largest_ir_len:,} bp, second={second_largest_len:,} bp (ratio: {size_ratio:.1f}x)"
                print(f"    ⚠ WARNING: {warning_msg}")
                warnings.append(warning_msg)
                
                # Check if third IR exists and is more similar in size to the largest
                if len(substantial_irs_by_size) >= 3:
                    third_largest_len = substantial_irs_by_size[2].location.end - substantial_irs_by_size[2].location.start
                    ratio_with_third = largest_ir_len / third_largest_len if third_largest_len > 0 else 0
                    
                    # If third IR is more similar to first than second is
                    if ratio_with_third < size_ratio:
                        print(f"    → Found better IR match: using largest ({largest_ir_len:,} bp) and third ({third_largest_len:,} bp)")
                        print(f"    → Skipping second IR ({second_largest_len:,} bp) as it's too different")
                        selected_irs = [substantial_irs_by_size[0], substantial_irs_by_size[2]]
                        warnings.append(f"Used 1st and 3rd largest IRs instead of 1st and 2nd (better size match)")
                    else:
                        selected_irs = substantial_irs_by_size[:2]
                        warnings.append(f"Proceeding with 2 largest IRs despite size difference")
                else:
                    selected_irs = substantial_irs_by_size[:2]
                    warnings.append(f"Only 2 substantial IRs found, using both despite size difference")
            
            # Sort selected IRs by position to assign IRa (first) and IRb (second)
            selected_irs.sort(key=lambda f: f.location.start)
            
            features['ira'] = selected_irs[0]
            features['irb'] = selected_irs[1]
            
            ira_len = features['ira'].location.end - features['ira'].location.start
            irb_len = features['irb'].location.end - features['irb'].location.start
            
            if len(substantial_irs) == 2:
                print(f"    ✓ IRa assigned: {features['ira'].location.start+1}-{features['ira'].location.end} ({ira_len:,} bp)")
                print(f"    ✓ IRb assigned: {features['irb'].location.start+1}-{features['irb'].location.end} ({irb_len:,} bp)")
            else:
                print(f"    ✓ IRa and IRb inferred from {len(substantial_irs)} repeat regions")
                print(f"      IRa: {features['ira'].location.start+1}-{features['ira'].location.end} ({ira_len:,} bp)")
                print(f"      IRb: {features['irb'].location.start+1}-{features['irb'].location.end} ({irb_len:,} bp)")
                if len(substantial_irs) > 2:
                    print(f"    ⚠ Note: Found {len(substantial_irs)} substantial IR regions (>= 1kb), selected 2 best matches")
                    warnings.append(f"Multiple IR regions detected ({len(substantial_irs)} regions >= 1kb)")
        
        elif len(substantial_irs) == 1:
            warning_msg = "Only one substantial IR region found (>= 1kb)"
            print(f"    ⚠ Warning: {warning_msg}")
            warnings.append(warning_msg)
        else:
            warning_msg = f"No substantial IR regions found (>= 1kb), all {len(unique_irs)} IR-like regions were < 1kb"
            print(f"    ⚠ Warning: {warning_msg}")
            warnings.append(warning_msg)
    
    # COMPREHENSIVE INTELLIGENT INFERENCE
    # Handle all possible partial annotation scenarios
    
    has_ira = features['ira'] is not None
    has_irb = features['irb'] is not None
    has_lsc = features['lsc'] is not None
    has_ssc = features['ssc'] is not None
    
    # Scenario 1: Only IRs annotated → Infer LSC/SSC
    if has_ira and has_irb and not has_lsc and not has_ssc:
        print(f"    → Both IRs found, inferring LSC/SSC regions...")
        inferred = infer_lsc_ssc_from_irs(record, features['ira'], features['irb'])
        if inferred:
            features['lsc'] = inferred['lsc']
            features['ssc'] = inferred['ssc']
            warnings.extend(inferred.get('warnings', []))
    
    # Scenario 2: Only LSC/SSC annotated → Infer IRs
    elif has_lsc and has_ssc and not has_ira and not has_irb:
        print(f"    → LSC and SSC found, inferring IR regions...")
        inferred = infer_irs_from_lsc_ssc(record, features['lsc'], features['ssc'])
        if inferred:
            features['ira'] = inferred['ira']
            features['irb'] = inferred['irb']
            warnings.extend(inferred.get('warnings', []))
    
    # Scenario 3: IRs + SSC (missing LSC) → Infer LSC
    elif has_ira and has_irb and has_ssc and not has_lsc:
        print(f"    → IRs and SSC found, inferring LSC...")
        # LSC is the region not covered by IRs and SSC
        inferred_lsc = infer_missing_region(record, features['ira'], features['irb'], features['ssc'], 'LSC')
        if inferred_lsc:
            features['lsc'] = inferred_lsc['feature']
            warnings.extend(inferred_lsc.get('warnings', []))
    
    # Scenario 4: IRs + LSC (missing SSC) → Infer SSC
    elif has_ira and has_irb and has_lsc and not has_ssc:
        print(f"    → IRs and LSC found, inferring SSC...")
        inferred_ssc = infer_missing_region(record, features['ira'], features['irb'], features['lsc'], 'SSC')
        if inferred_ssc:
            features['ssc'] = inferred_ssc['feature']
            warnings.extend(inferred_ssc.get('warnings', []))
    
    # Scenario 5: LSC + one IR (missing SSC + one IR) → Infer both
    elif has_lsc and (has_ira or has_irb) and not has_ssc:
        known_ir = features['ira'] if has_ira else features['irb']
        ir_name = 'IRa' if has_ira else 'IRb'
        print(f"    → LSC and {ir_name} found, inferring SSC and other IR...")
        # This is complex - would need the other IR and SSC
        warning_msg = f"Only LSC and one IR annotated - insufficient information to infer remaining regions"
        print(f"    ⚠ Warning: {warning_msg}")
        warnings.append(warning_msg)
    
    # Scenario 6: SSC + one IR (missing LSC + one IR) → Infer both
    elif has_ssc and (has_ira or has_irb) and not has_lsc:
        known_ir = features['ira'] if has_ira else features['irb']
        ir_name = 'IRa' if has_ira else 'IRb'
        print(f"    → SSC and {ir_name} found, inferring LSC and other IR...")
        warning_msg = f"Only SSC and one IR annotated - insufficient information to infer remaining regions"
        print(f"    ⚠ Warning: {warning_msg}")
        warnings.append(warning_msg)
    
    # Scenario 7: Only one IR found (no LSC/SSC)
    elif (has_ira or has_irb) and not has_lsc and not has_ssc:
        ir_name = 'IRa' if has_ira else 'IRb'
        warning_msg = f"Only {ir_name} annotated - insufficient information to infer other regions"
        print(f"    ⚠ Warning: {warning_msg}")
        warnings.append(warning_msg)
    
    # Scenario 8: Only LSC or only SSC found
    elif (has_lsc or has_ssc) and not has_ira and not has_irb:
        region_name = 'LSC' if has_lsc else 'SSC'
        warning_msg = f"Only {region_name} annotated - insufficient information to infer other regions"
        print(f"    ⚠ Warning: {warning_msg}")
        warnings.append(warning_msg)
    
    return features, warnings


def infer_missing_region(record, ira_feature, irb_feature, known_region, missing_type):
    """
    Infer missing LSC or SSC when we have both IRs and one of LSC/SSC.
    
    Logic:
    - Genome = IRa + IRb + LSC + SSC
    - If we have IRs + SSC → LSC = Total - IRa - IRb - SSC
    - If we have IRs + LSC → SSC = Total - IRa - IRb - LSC
    
    Args:
        record: BioPython SeqRecord
        ira_feature: IRa feature
        irb_feature: IRb feature
        known_region: The known region (LSC or SSC)
        missing_type: 'LSC' or 'SSC'
    
    Returns:
        dict with 'feature' and 'warnings'
    """
    from Bio.SeqFeature import SeqFeature, FeatureLocation
    
    genome_length = len(record.seq)
    warnings = []
    
    # Get all known region boundaries
    ira_start = int(ira_feature.location.start)
    ira_end = int(ira_feature.location.end)
    irb_start = int(irb_feature.location.start)
    irb_end = int(irb_feature.location.end)
    known_start = int(known_region.location.start)
    known_end = int(known_region.location.end)
    
    # Sort all regions by position
    all_regions = [
        (ira_start, ira_end, 'IRa'),
        (irb_start, irb_end, 'IRb'),
        (known_start, known_end, 'LSC' if missing_type == 'SSC' else 'SSC')
    ]
    all_regions.sort(key=lambda x: x[0])
    
    print(f"    → Known regions:")
    for start, end, name in all_regions:
        print(f"      {name}: {start+1}-{end} ({end-start:,} bp)")
    
    # Find gaps between regions
    gaps = []
    
    # Gap before first region (if any)
    if all_regions[0][0] > 0:
        gaps.append((0, all_regions[0][0], all_regions[0][0]))
    
    # Gaps between consecutive regions
    for i in range(len(all_regions) - 1):
        current_end = all_regions[i][1]
        next_start = all_regions[i+1][0]
        if next_start > current_end:
            gaps.append((current_end, next_start, next_start - current_end))
    
    # Gap after last region (if any)
    if all_regions[-1][1] < genome_length:
        gaps.append((all_regions[-1][1], genome_length, genome_length - all_regions[-1][1]))
    
    if not gaps:
        warning_msg = f"No gaps found to assign as {missing_type}"
        print(f"    ✗ {warning_msg}")
        warnings.append(warning_msg)
        return {'feature': None, 'warnings': warnings}
    
    # The gap should be the missing region
    if len(gaps) == 1:
        missing_start, missing_end, missing_length = gaps[0]
        print(f"    ✓ {missing_type} inferred: {missing_start+1}-{missing_end} ({missing_length:,} bp)")
    else:
        # Multiple gaps - take the largest
        gaps.sort(key=lambda x: x[2], reverse=True)
        missing_start, missing_end, missing_length = gaps[0]
        warning_msg = f"Multiple gaps found, using largest as {missing_type}"
        print(f"    ⚠ {warning_msg}")
        print(f"    ✓ {missing_type} inferred: {missing_start+1}-{missing_end} ({missing_length:,} bp)")
        warnings.append(warning_msg)
    
    # Create feature
    missing_feature = SeqFeature(
        FeatureLocation(missing_start, missing_end),
        type="misc_feature",
        qualifiers={"note": [f"{missing_type.lower()} - inferred from IRs and {'SSC' if missing_type == 'LSC' else 'LSC'}"]}
    )
    
    return {'feature': missing_feature, 'warnings': warnings}


def infer_lsc_ssc_from_irs(record, ira_feature, irb_feature):
    """
    Infer LSC and SSC regions when only IRa and IRb are annotated.
    
    Logic:
    - Identifies all gaps between/around the two IRs
    - Larger gap = LSC
    - Smaller gap = SSC
    
    Handles cases where:
    - IRs start at position 0 (standard)
    - IRs start mid-genome (non-standard annotation)
    
    Args:
        record: BioPython SeqRecord
        ira_feature: IRa feature
        irb_feature: IRb feature
    
    Returns:
        dict with inferred 'lsc' and 'ssc' features, or None if inference fails
    """
    from Bio.SeqFeature import SeqFeature, FeatureLocation
    
    genome_length = len(record.seq)
    
    # Get IR boundaries
    ira_start = int(ira_feature.location.start)
    ira_end = int(ira_feature.location.end)
    irb_start = int(irb_feature.location.start)
    irb_end = int(irb_feature.location.end)
    
    print(f"    → IR regions found:")
    print(f"      IR1: {ira_start+1}-{ira_end} ({ira_end - ira_start:,} bp)")
    print(f"      IR2: {irb_start+1}-{irb_end} ({irb_end - irb_start:,} bp)")
    
    # Sort IRs by start position to identify them correctly
    if ira_start < irb_start:
        first_ir_start, first_ir_end = ira_start, ira_end
        second_ir_start, second_ir_end = irb_start, irb_end
    else:
        first_ir_start, first_ir_end = irb_start, irb_end
        second_ir_start, second_ir_end = ira_start, ira_end
    
    # Identify all potential regions (gaps between/around IRs)
    regions = []
    
    # Region 1: Before first IR (if IR doesn't start at 0)
    if first_ir_start > 0:
        region1_start = 0
        region1_end = first_ir_start
        region1_length = region1_end - region1_start
        regions.append({
            'start': region1_start,
            'end': region1_end,
            'length': region1_length,
            'name': f'Before first IR (0-{first_ir_start})'
        })
        print(f"    → Region 1: 1-{first_ir_start} ({region1_length:,} bp)")
    
    # Region 2: Between the two IRs
    region2_start = first_ir_end
    region2_end = second_ir_start
    region2_length = region2_end - region2_start
    if region2_length > 0:
        regions.append({
            'start': region2_start,
            'end': region2_end,
            'length': region2_length,
            'name': f'Between IRs ({first_ir_end}-{second_ir_start})'
        })
        print(f"    → Region 2: {first_ir_end+1}-{second_ir_start} ({region2_length:,} bp)")
    
    # Region 3: After second IR (if IR doesn't end at genome end)
    if second_ir_end < genome_length:
        region3_start = second_ir_end
        region3_end = genome_length
        region3_length = region3_end - region3_start
        regions.append({
            'start': region3_start,
            'end': region3_end,
            'length': region3_length,
            'name': f'After second IR ({second_ir_end}-{genome_length})'
        })
        print(f"    → Region 3: {second_ir_end+1}-{genome_length} ({region3_length:,} bp)")
    
    # If we don't have exactly 2 regions, something is unusual
    if len(regions) != 2:
        print(f"    ⚠ Warning: Found {len(regions)} potential LSC/SSC regions (expected 2)")
        print(f"    → This may indicate unusual genome structure or annotation errors")
    
    # Sort regions by length (descending)
    regions.sort(key=lambda x: x['length'], reverse=True)
    
    print(f"    → Regions sorted by size:")
    for i, r in enumerate(regions):
        print(f"      {i+1}. Start={r['start']+1}, End={r['end']}, Length={r['length']:,} bp")
    
    # Largest region = LSC, smallest = SSC
    if len(regions) >= 2:
        lsc_region = regions[0]
        ssc_region = regions[1]
        print(f"    → LSC will use: start={lsc_region['start']}, end={lsc_region['end']}")
        print(f"    → SSC will use: start={ssc_region['start']}, end={ssc_region['end']}")
    elif len(regions) == 1:
        # Only one gap region found - unusual
        print(f"    ⚠ Warning: Only one gap region found, treating as LSC")
        lsc_region = regions[0]
        ssc_region = {'start': 0, 'end': 0, 'length': 0, 'name': 'Not found'}
    else:
        print(f"    ✗ Error: No gap regions found between IRs")
        return None
    
    # Create LSC feature
    if lsc_region['length'] > 0:
        print(f"    DEBUG: Creating LSC with start={lsc_region['start']}, end={lsc_region['end']}")
        lsc_feature = SeqFeature(
            FeatureLocation(lsc_region['start'], lsc_region['end']),
            type="misc_feature",
            qualifiers={"note": [f"large single copy (LSC) - inferred from IR gaps"]}
        )
        print(f"    ✓ LSC inferred: {lsc_region['start']+1}-{lsc_region['end']} ({lsc_region['length']:,} bp)")
    else:
        lsc_feature = None
        print(f"    ⚠ LSC not found")
    
    # Create SSC feature
    if ssc_region['length'] > 0:
        print(f"    DEBUG: Creating SSC with start={ssc_region['start']}, end={ssc_region['end']}")
        ssc_feature = SeqFeature(
            FeatureLocation(ssc_region['start'], ssc_region['end']),
            type="misc_feature",
            qualifiers={"note": [f"small single copy (SSC) - inferred from IR gaps"]}
        )
        print(f"    ✓ SSC inferred: {ssc_region['start']+1}-{ssc_region['end']} ({ssc_region['length']:,} bp)")
    else:
        ssc_feature = None
        print(f"    ⚠ SSC not found")
    
    # Validation
    total_calculated = (first_ir_end - first_ir_start) + (second_ir_end - second_ir_start)
    if lsc_feature:
        total_calculated += lsc_region['length']
    if ssc_feature:
        total_calculated += ssc_region['length']
    
    warnings = []
    diff = abs(genome_length - total_calculated)
    if diff > 100:
        warning_msg = f"Calculated total ({total_calculated:,} bp) differs from genome ({genome_length:,} bp) by {diff:,} bp"
        print(f"    ⚠ Warning: {warning_msg}")
        warnings.append(warning_msg)
    
    return {
        'lsc': lsc_feature,
        'ssc': ssc_feature,
        'warnings': warnings
    }


def infer_irs_from_lsc_ssc(record, lsc_feature, ssc_feature):
    """
    Infer IRa and IRb regions when only LSC and SSC are annotated.
    
    Logic:
    - IRa = region before LSC (or after SSC in circular genome)
    - IRb = region after SSC (or before LSC in circular genome)
    - Both IRs should have similar length
    
    Args:
        record: BioPython SeqRecord
        lsc_feature: LSC feature
        ssc_feature: SSC feature
    
    Returns:
        dict with inferred 'ira' and 'irb' features, or None if inference fails
    """
    from Bio.SeqFeature import SeqFeature, FeatureLocation
    
    genome_length = len(record.seq)
    
    lsc_start = lsc_feature.location.start
    lsc_end = lsc_feature.location.end
    ssc_start = ssc_feature.location.start
    ssc_end = ssc_feature.location.end
    
    # Determine genome structure order
    # Typical: IRa - LSC - IRb - SSC - (back to IRa)
    # But could be: IRa - SSC - IRb - LSC - (back to IRa)
    
    if lsc_start < ssc_start:
        # LSC comes before SSC: IRa - LSC - IRb - SSC - IRa
        # IRb is between LSC and SSC
        irb_start = lsc_end
        irb_end = ssc_start
        
        # IRa wraps around (after SSC to before LSC)
        ira_start = ssc_end
        ira_end = genome_length if lsc_start == 0 else lsc_start
        
    else:
        # SSC comes before LSC: IRa - SSC - IRb - LSC - IRa
        # IRb is between SSC and LSC
        irb_start = ssc_end
        irb_end = lsc_start
        
        # IRa wraps around (after LSC to before SSC)
        ira_start = lsc_end
        ira_end = genome_length if ssc_start == 0 else ssc_start
    
    # Calculate IR lengths
    irb_length = irb_end - irb_start
    if ira_end > ira_start:
        ira_length = ira_end - ira_start
    else:
        ira_length = (genome_length - ira_start) + ira_end
    
    # Create inferred features
    ira_feature = SeqFeature(
        FeatureLocation(ira_start, ira_end if ira_end > ira_start else genome_length),
        type="misc_feature",
        qualifiers={"note": ["inverted repeat A (IRa) - inferred"]}
    )
    
    irb_feature = SeqFeature(
        FeatureLocation(irb_start, irb_end),
        type="misc_feature",
        qualifiers={"note": ["inverted repeat B (IRb) - inferred"]}
    )
    
    print(f"    ✓ IRa inferred: {ira_start}-{ira_end if ira_end > ira_start else genome_length} (length: {ira_length:,} bp)")
    print(f"    ✓ IRb inferred: {irb_start}-{irb_end} (length: {irb_length:,} bp)")
    
    # Validate that IRs have similar length (should be nearly identical)
    length_diff = abs(ira_length - irb_length)
    if length_diff > 100:
        print(f"    ⚠ Warning: Inferred IR lengths differ by {length_diff} bp (IRa: {ira_length}, IRb: {irb_length})")
    
    return {
        'ira': ira_feature,
        'irb': irb_feature
    }


def calculate_region_stats(record, features):
    """
    Calculate lengths and GC content for LSC, SSC, and IR regions.
    
    Args:
        record: BioPython SeqRecord
        features: dict from detect_structural_features()
    
    Returns:
        dict with lengths and GC percentages
    """
    genome_length = len(record.seq)
    
    stats = {
        'lsc_length': 0,
        'ssc_length': 0,
        'ir_length': 0,
        'lsc_gc': 0,
        'ssc_gc': 0,
        'ir_gc': 0
    }
    
    # LSC stats
    if features['lsc']:
        lsc_seq = record.seq[features['lsc'].location.start:features['lsc'].location.end]
        stats['lsc_length'] = len(lsc_seq)
        stats['lsc_gc'] = calculate_gc(lsc_seq)
    
    # SSC stats
    if features['ssc']:
        ssc_seq = record.seq[features['ssc'].location.start:features['ssc'].location.end]
        stats['ssc_length'] = len(ssc_seq)
        stats['ssc_gc'] = calculate_gc(ssc_seq)
    
    # IR stats - use IRa or IRb (should be identical)
    if features['ira']:
        ira_seq = record.seq[features['ira'].location.start:features['ira'].location.end]
        stats['ir_length'] = len(ira_seq)
        stats['ir_gc'] = calculate_gc(ira_seq)
        print(f"    → Using IRa for IR statistics (length: {stats['ir_length']} bp)")
    elif features['irb']:
        irb_seq = record.seq[features['irb'].location.start:features['irb'].location.end]
        stats['ir_length'] = len(irb_seq)
        stats['ir_gc'] = calculate_gc(irb_seq)
        print(f"    → Using IRb for IR statistics (length: {stats['ir_length']} bp)")
    
    # Fallback: Calculate IR by subtraction if not directly annotated
    if stats['ir_length'] == 0 and stats['lsc_length'] > 0 and stats['ssc_length'] > 0:
        stats['ir_length'] = (genome_length - stats['lsc_length'] - stats['ssc_length']) // 2
        print(f"    → IR length calculated by subtraction: {stats['ir_length']} bp")
        
        # Calculate IR GC from remaining sequence
        if features['lsc'] and features['ssc']:
            ir_seq = (
                record.seq[:features['lsc'].location.start] +
                record.seq[features['lsc'].location.end:features['ssc'].location.start] +
                record.seq[features['ssc'].location.end:]
            )
            stats['ir_gc'] = calculate_gc(ir_seq)
    
    # Validation
    total_calculated = stats['lsc_length'] + stats['ssc_length'] + (stats['ir_length'] * 2)
    if total_calculated > 0:
        diff = abs(genome_length - total_calculated)
        if diff > 100:  # Allow small differences due to gaps/overlaps
            print(f"    ⚠ Warning: Calculated total ({total_calculated:,}) differs from genome length ({genome_length:,}) by {diff} bp")
    
    return stats

# ============================================================================
# MAIN ANALYSIS
# ============================================================================


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main():
    import argparse
    
    parser = argparse.ArgumentParser(
        description="CGAS Module 7: Comparative Chloroplast Genome Analysis"
    )
    parser.add_argument("-i", "--input", default=".", help="Input directory")
    parser.add_argument("-o", "--output", default=None, help="Output directory name")
    args = parser.parse_args()
    
    global OUTPUT_FOLDER
    folder_path = os.path.abspath(args.input)
    if args.output:
        OUTPUT_FOLDER = args.output
    output_folder_full = os.path.join(folder_path, OUTPUT_FOLDER)
    
    print(f"\n{'='*70}")
    print("CGAS MODULE 7: COMPARATIVE GENOME ANALYSIS")
    print(f"{'='*70}")

    # Initialize list to store results and track errors
    results = []
    failed_files = []
    all_warnings = {}  # Dict: filename -> list of warnings
    success_count = 0

    # Use current working directory
    folder_path = folder_path
    print(f"\nInput directory: {folder_path}")

    # Create output folder
    output_folder_full = os.path.join(folder_path, OUTPUT_FOLDER)
    os.makedirs(output_folder_full, exist_ok=True)
    print(f"Output folder: {output_folder_full}/")

    # Count GenBank files
    gb_files = [f for f in os.listdir(folder_path) if f.endswith(('.gb', '.gbf', '.gbk', '.genbank'))]
    
    if not gb_files:
        print(f"\n⚠ No GenBank files found in {folder_path}")
        print("Supported extensions: .gb, .gbf, .gbk, .genbank")
        return
    
    print(f"\nFound {len(gb_files)} GenBank file(s)")
    print(f"\n{'='*70}")
    print("ANALYZING GENOMES")
    print(f"{'='*70}")

    # Loop through GenBank files with error handling
    for filename in gb_files:
        try:
            print(f"\nProcessing: {filename}")
            record_path = os.path.join(folder_path, filename)
            
            # Read GenBank file
            try:
                record = SeqIO.read(record_path, 'genbank')
            except Exception as e:
                print(f"  ✗ Error reading GenBank file: {str(e)}")
                failed_files.append((filename, f"Failed to read GenBank file: {str(e)}"))
                continue
        
            # Basic genome information
            try:
                species = record.annotations.get('organism', 'Unknown')
                genome_length = len(record.seq)
                accession_number = record.annotations.get('accessions', ['N/A'])[0]
            except Exception as e:
                print(f"  ✗ Error extracting metadata: {str(e)}")
                failed_files.append((filename, f"Failed to extract metadata: {str(e)}"))
                continue
        
            print(f"  Species: {species}")
            print(f"  Genome length: {genome_length:,} bp")
            print(f"  Detecting structural features...")
            
            # Detect LSC, SSC, IRa, IRb regions
            try:
                features, genome_warnings = detect_structural_features(record)
                if genome_warnings:
                    all_warnings[filename] = genome_warnings
            except Exception as e:
                print(f"  ✗ Error detecting structural features: {str(e)}")
                failed_files.append((filename, f"Failed to detect features: {str(e)}"))
                continue
            
            # Calculate region statistics
            try:
                region_stats = calculate_region_stats(record, features)
            except Exception as e:
                print(f"  ✗ Error calculating region statistics: {str(e)}")
                failed_files.append((filename, f"Failed to calculate stats: {str(e)}"))
                continue
            
            # Extract values
            lsc_length = region_stats['lsc_length']
            ssc_length = region_stats['ssc_length']
            ir_length = region_stats['ir_length']
            lsc_gc = region_stats['lsc_gc']
            ssc_gc = region_stats['ssc_gc']
            ir_gc = region_stats['ir_gc']
            
            # Calculate total genome GC
            total_gc = calculate_gc(record.seq)
    
            # Extract sequences for tRNA, rRNA, CDS
            try:
                tRNA_seq = rRNA_seq = CDS_seq = ""

                for feature in record.features:
                    parts = []
                    if hasattr(feature.location, "parts"):
                        parts = feature.location.parts
                    else:
                        parts = [feature.location]

                    if feature.type == 'CDS':
                        for part in parts:
                            CDS_seq += part.extract(record.seq)
                    elif feature.type == 'tRNA':
                        for part in parts:
                            tRNA_seq += part.extract(record.seq)
                    elif feature.type == 'rRNA':
                        for part in parts:
                            rRNA_seq += part.extract(record.seq)

                # GC content for gene groups
                tRNA_gc = calculate_gc(tRNA_seq)
                rRNA_gc = calculate_gc(rRNA_seq)
                CDS_gc = calculate_gc(CDS_seq)
            except Exception as e:
                print(f"  ⚠ Warning: Error calculating gene-type GC content: {str(e)}")
                print(f"  → Using 0 for tRNA/rRNA/CDS GC values")
                tRNA_gc = rRNA_gc = CDS_gc = 0
    
            # Store results as NUMBERS (not formatted strings)
            results.append([
                species, genome_length, lsc_length, ssc_length, ir_length, 
                total_gc, lsc_gc, ssc_gc, ir_gc, 
                tRNA_gc, rRNA_gc, CDS_gc, accession_number
            ])
    
            print(f"  ✓ Analysis complete")
            success_count += 1
            
        except Exception as e:
            # Catch any unexpected errors
            print(f"  ✗ Unexpected error processing {filename}: {str(e)}")
            failed_files.append((filename, f"Unexpected error: {str(e)}"))
            continue

    # Summary of processing
    print(f"\n{'='*70}")
    print("PROCESSING SUMMARY")
    print(f"{'='*70}")
    print(f"Total files: {len(gb_files)}")
    print(f"Successfully processed: {success_count}")
    print(f"Failed: {len(failed_files)}")
    
    if failed_files:
        print(f"\nFailed files:")
        for filename, error in failed_files:
            print(f"  ✗ {filename}: {error}")
    
    if success_count == 0:
        print(f"\n⚠ ERROR: No genomes were successfully analyzed!")
        print(f"Cannot generate comparative table.")
        return

    print(f"\n{'='*70}")
    print("CREATING COMPARATIVE TABLE")
    print(f"{'='*70}")

    # Create output table with new column structure
    df = pd.DataFrame(results, columns=[
        "Species", "Complete", "LSC", "SSC", "IR",
        "Complete_GC", "LSC_GC", "SSC_GC", "IR_GC",
        "tRNA", "rRNA", "CDS", "Accession Number"
    ])

    # Save to Excel in output folder
    output_file = os.path.join(output_folder_full, OUTPUT_FILENAME)
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Apply publication-quality formatting
    wb = load_workbook(output_file)
    ws = wb.active

    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    subheader_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_font = Font(name='Arial', size=10)
    species_font = Font(name='Arial', size=10, italic=True)
    cell_alignment = Alignment(horizontal='center', vertical='center')
    species_alignment = Alignment(horizontal='left', vertical='center')

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Insert a new row at the top for main headers
    ws.insert_rows(1)

    # Merge cells for main headers
    ws.merge_cells('A1:A2')  # Species
    ws.merge_cells('B1:E1')  # Genome Length (bp)
    ws.merge_cells('F1:I1')  # GC (%)
    ws.merge_cells('J1:L1')  # GC (%)
    ws.merge_cells('M1:M2')  # Accession Number

    # Set main headers (row 1)
    ws['A1'] = 'Species'
    ws['B1'] = 'Genome Length (bp)'
    ws['F1'] = 'GC (%)'
    ws['J1'] = 'GC (%)'
    ws['M1'] = 'Accession Number'

    # Set subheaders (row 2)
    ws['B2'] = 'Complete'
    ws['C2'] = 'LSC'
    ws['D2'] = 'SSC'
    ws['E2'] = 'IR'
    ws['F2'] = 'Complete'
    ws['G2'] = 'LSC'
    ws['H2'] = 'SSC'
    ws['I2'] = 'IR'
    ws['J2'] = 'tRNA'
    ws['K2'] = 'rRNA'
    ws['L2'] = 'CDS'

    # Format all header cells
    for row in [1, 2]:
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font if row == 1 else subheader_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    # Format data cells (starting from row 3)
    for row_num in range(3, len(df) + 3):
        for col_num in range(1, 14):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
        
            # Species column - italic and left-aligned
            if col_num == 1:
                cell.font = species_font
                cell.alignment = species_alignment
            # Genome length columns (B-E: Complete, LSC, SSC, IR)
            elif col_num in [2, 3, 4, 5]:
                cell.font = cell_font
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0'
            # GC percentage columns (F-L)
            elif col_num in [6, 7, 8, 9, 10, 11, 12]:
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.number_format = '0.00'
            # Accession number
            else:
                cell.font = cell_font
                cell.alignment = cell_alignment

    # Adjust column widths
    column_widths = {
        'A': 35,  # Species
        'B': 12,  # Complete
        'C': 12,  # LSC
        'D': 12,  # SSC
        'E': 12,  # IR
        'F': 12,  # Complete GC
        'G': 12,  # LSC GC
        'H': 12,  # SSC GC
        'I': 12,  # IR GC
        'J': 12,  # tRNA GC
        'K': 12,  # rRNA GC
        'L': 12,  # CDS GC
        'M': 16   # Accession Number
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Set row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25

    # Freeze the header rows
    ws.freeze_panes = 'A3'

    # Add footnote
    footnote_row = len(df) + 4
    ws.merge_cells(f'A{footnote_row}:M{footnote_row}')
    footnote_cell = ws.cell(row=footnote_row, column=1)
    footnote_cell.value = ('GC: guanine-cytosine content; LSC: large single-copy region; SSC: small single-copy region; '
                           'IR: inverted repeat region; Complete: complete chloroplast genome; '
                           'tRNA: transfer RNA; rRNA: ribosomal RNA; CDS: protein-coding gene.')
    footnote_cell.font = Font(name='Arial', size=9)
    footnote_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[footnote_row].height = 30

    # Save formatted workbook
    wb.save(output_file)

    # Create warnings log file if there are any warnings
    if all_warnings:
        warnings_file = os.path.join(output_folder_full, "Analysis_Warnings.txt")
        with open(warnings_file, 'w') as f:
            f.write("="*80 + "\n")
            f.write("CGAS MODULE 7 - STRUCTURAL ANNOTATION WARNINGS\n")
            f.write("="*80 + "\n\n")
            f.write(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total genomes with warnings: {len(all_warnings)}\n\n")
            f.write("="*80 + "\n\n")
            
            for filename, warnings in all_warnings.items():
                f.write(f"File: {filename}\n")
                f.write("-" * 80 + "\n")
                for i, warning in enumerate(warnings, 1):
                    f.write(f"  {i}. {warning}\n")
                f.write("\n")
            
            f.write("="*80 + "\n")
            f.write("EXPLANATION OF WARNINGS\n")
            f.write("="*80 + "\n\n")
            f.write("• Tiny IR regions (< 1kb): Likely annotation artifacts, safely ignored\n")
            f.write("• IR size discrepancy (2-4x ratio): One IR may be incorrectly annotated\n")
            f.write("• Unusual IR sizes (> 4x ratio): Strong indication of annotation errors\n")
            f.write("• Multiple IR regions: More than 2 IR-like features found, best match selected\n")
            f.write("• Calculated total differs: Sum of regions doesn't match genome length\n")
            f.write("  (may indicate gaps, overlaps, or missing annotations)\n\n")
            f.write("RECOMMENDATIONS:\n")
            f.write("• Review genomes with warnings using a genome viewer (e.g., Geneious, Artemis)\n")
            f.write("• Verify IR boundaries match the actual inverted repeat sequences\n")
            f.write("• Consider re-annotation if warnings indicate serious issues\n")
            f.write("• For publication, manually verify structural regions for warned genomes\n\n")
        
        print(f"  ✓ Warnings log: Analysis_Warnings.txt ({len(all_warnings)} genomes)")

    print(f"\n{'='*70}")
    print("CGAS MODULE 7 COMPLETE")
    print(f"{'='*70}")
    print(f"\n📊 RESULTS:")
    print(f"  ✓ Successfully analyzed: {success_count}/{len(gb_files)} genomes")
    if failed_files:
        print(f"  ✗ Failed to analyze: {len(failed_files)} genomes")
    if all_warnings:
        print(f"  ⚠ Genomes with warnings: {len(all_warnings)}")
    print(f"\n📁 OUTPUT:")
    print(f"  ✓ Publication-quality table: {OUTPUT_FILENAME}")
    print(f"  ✓ Format: Two-level headers with grouped columns")
    print(f"  ✓ Location: {output_folder_full}/")
    if all_warnings:
        print(f"  ⚠ Warnings log: Analysis_Warnings.txt")
    
    if failed_files:
        print(f"\n⚠ FAILED FILES:")
        print(f"  The following files could not be analyzed:")
        for filename, error in failed_files:
            print(f"    • {filename}")
        print(f"  See error messages above for details.")
    
    if all_warnings:
        print(f"\n⚠ ANNOTATION WARNINGS:")
        print(f"  {len(all_warnings)} genome(s) have structural annotation issues")
        print(f"  See Analysis_Warnings.txt for details and recommendations")
    
    print(f"\n{'='*70}\n")

if __name__ == "__main__":
    main()

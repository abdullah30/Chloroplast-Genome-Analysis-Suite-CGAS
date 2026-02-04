#!/usr/bin/env python3
"""
CGAS Module 8: Codon Usage Analysis (RSCU) with R Visualization
================================================================

This script performs comprehensive codon usage analysis on chloroplast genomes by:
1. Extracting coding sequences (CDS) from GenBank files
2. Calculating Relative Synonymous Codon Usage (RSCU) values
3. Calculating codon counts for all codons
4. Generating individual RSCU reports for each genome
5. Creating merged comparative analysis across all genomes
6. Generating publication-quality R-based visualizations

Author: Abdullah
Version: 1.0.1
Date: January 2026

Dependencies:
    Python: biopython, pandas, openpyxl
    R: ggplot2, pheatmap, RColorBrewer, reshape2

Usage:
    python cgas_module8.py
    python cgas_module8.py -i genbank_files/
    python cgas_module8.py -i genbank_files/ -o custom_output/
    python cgas_module8.py --no-figures  # Skip R visualization

Output:
    Module8_Codon_Usage_Analysis/
        - Individual RSCU files: [genome_name]_RSCU.xlsx
        - Merged analysis: Complete_Codon_Usage_Analysis.xlsx
        - Figures/ (if R available):
            - RSCU_Heatmap.pdf
            - Codon_Usage_Comparison.pdf
"""

import os
import sys
import argparse
import subprocess
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from collections import Counter, defaultdict

try:
    from Bio import SeqIO
    import pandas as pd
except ImportError as e:
    print(f"Error: Required package not installed: {e}")
    print("Please install required packages using:")
    print("pip install biopython pandas openpyxl --break-system-packages")
    sys.exit(1)


# ============================================================================
# CONSTANTS
# ============================================================================

OUTPUT_FOLDER = "Module8_Codon_Usage_Analysis"
FIGURES_FOLDER = "Figures"

# Standard genetic code: mapping of amino acids to their synonymous codons
SYNONYMOUS_CODONS = {
    'Ala': ['GCT', 'GCC', 'GCA', 'GCG'],
    'Arg': ['CGT', 'CGC', 'CGA', 'CGG', 'AGA', 'AGG'],
    'Asn': ['AAT', 'AAC'],
    'Asp': ['GAT', 'GAC'],
    'Cys': ['TGT', 'TGC'],
    'Gln': ['CAA', 'CAG'],
    'Glu': ['GAA', 'GAG'],
    'Gly': ['GGT', 'GGC', 'GGA', 'GGG'],
    'His': ['CAT', 'CAC'],
    'Ile': ['ATT', 'ATC', 'ATA'],
    'Leu': ['TTA', 'TTG', 'CTT', 'CTC', 'CTA', 'CTG'],
    'Lys': ['AAA', 'AAG'],
    'Met': ['ATG'],
    'Phe': ['TTT', 'TTC'],
    'Pro': ['CCT', 'CCC', 'CCA', 'CCG'],
    'Ser': ['TCT', 'TCC', 'TCA', 'TCG', 'AGT', 'AGC'],
    'Thr': ['ACT', 'ACC', 'ACA', 'ACG'],
    'Trp': ['TGG'],
    'Tyr': ['TAT', 'TAC'],
    'Val': ['GTT', 'GTC', 'GTA', 'GTG'],
    'Stop': ['TAA', 'TAG', 'TGA']
}

# Full amino acid names
AMINO_ACID_NAMES = {
    'Ala': 'Alanine', 'Arg': 'Arginine', 'Asn': 'Asparagine',
    'Asp': 'Aspartate', 'Cys': 'Cysteine', 'Gln': 'Glutamine',
    'Glu': 'Glutamate', 'Gly': 'Glycine', 'His': 'Histidine',
    'Ile': 'Isoleucine', 'Leu': 'Leucine', 'Lys': 'Lysine',
    'Met': 'Methionine', 'Phe': 'Phenylalanine', 'Pro': 'Proline',
    'Ser': 'Serine', 'Thr': 'Threonine', 'Trp': 'Tryptophan',
    'Tyr': 'Tyrosine', 'Val': 'Valine', 'Stop': 'Stop'
}

AA_ORDER = ['Ala', 'Arg', 'Asn', 'Asp', 'Cys', 'Gln', 'Glu', 'Gly', 'His', 
            'Ile', 'Leu', 'Lys', 'Met', 'Phe', 'Pro', 'Ser', 'Thr', 'Trp', 
            'Tyr', 'Val', 'Stop']


# ============================================================================
# CODON EXTRACTION AND COUNTING
# ============================================================================

def extract_coding_sequences(genbank_file: str) -> List[str]:
    """Extract all coding sequences (CDS) from a GenBank file."""
    try:
        genome_record = SeqIO.read(genbank_file, "genbank")
    except FileNotFoundError:
        raise FileNotFoundError(f"GenBank file not found: {genbank_file}")
    except Exception as e:
        raise ValueError(f"Error parsing GenBank file {genbank_file}: {e}")
    
    coding_sequences = []
    cds_count = 0
    
    for feature in genome_record.features:
        if feature.type == "CDS":
            cds_count += 1
            seq = str(feature.extract(genome_record.seq))
            seq = seq.replace("\n", "").replace(" ", "").upper()
            coding_sequences.append(seq)
    
    print(f"  - Extracted {cds_count} coding sequences")
    return coding_sequences


def count_codons(coding_sequences: List[str]) -> Tuple[Counter, List[str]]:
    """Count the occurrence of each codon across all coding sequences."""
    valid_nucleotides = set('ATCG')
    all_cds_sequence = "".join(coding_sequences)
    
    codon_counts = Counter()
    unusual_codons = []
    
    for i in range(0, len(all_cds_sequence), 3):
        codon = all_cds_sequence[i:i+3]
        
        if len(codon) == 3:
            codon_counts[codon] += 1
            
            codon_nucleotides = set(codon)
            if not codon_nucleotides.issubset(valid_nucleotides):
                if codon not in unusual_codons:
                    unusual_codons.append(codon)
    
    total_codons = sum(codon_counts.values())
    print(f"  - Total codons counted: {total_codons:,}")
    
    if unusual_codons:
        print(f"  ⚠ WARNING: Found {len(unusual_codons)} unusual codon(s): {', '.join(unusual_codons[:10])}")
        if len(unusual_codons) > 10:
            print(f"           (and {len(unusual_codons) - 10} more...)")
    
    return codon_counts, unusual_codons


# ============================================================================
# RSCU CALCULATION
# ============================================================================

def calculate_rscu(codon_counts: Counter, 
                   synonymous_codons: Dict[str, List[str]],
                   amino_acid_names: Dict[str, str]) -> pd.DataFrame:
    """
    Calculate RSCU values for all codons.
    
    RSCU = (Observed frequency of codon) / (Expected frequency if all synonymous codons used equally)
    RSCU = 1.0 means no bias
    RSCU > 1.0 means codon is used more than expected
    RSCU < 1.0 means codon is used less than expected
    """
    rscu_data = []
    
    for aa_code in AA_ORDER:
        codons = synonymous_codons[aa_code]
        amino_acid_name = amino_acid_names[aa_code]
        
        # Get counts for all synonymous codons
        codon_count_list = [codon_counts.get(codon, 0) for codon in codons]
        total_count = sum(codon_count_list)
        
        # Number of synonymous codons for this amino acid
        n_synonymous = len(codons)
        
        # Calculate RSCU for each codon
        for codon, count in zip(codons, codon_count_list):
            if total_count == 0:
                rscu = 0.0
            else:
                # RSCU formula
                expected_freq = count / total_count
                uniform_freq = 1.0 / n_synonymous
                rscu = expected_freq / uniform_freq if uniform_freq > 0 else 0.0
            
            rscu_data.append({
                'AA': aa_code,
                'Amino_Acid': amino_acid_name,
                'Codon': codon,
                'Count': count,
                'RSCU': rscu
            })
    
    return pd.DataFrame(rscu_data)


# ============================================================================
# FILE PROCESSING
# ============================================================================

def find_genbank_files(directory: str) -> List[str]:
    """Find all GenBank files in the specified directory."""
    extensions = ['.gb', '.gbf', '.gbk', '.genbank']
    genbank_files = []
    
    for ext in extensions:
        genbank_files.extend(Path(directory).glob(f'*{ext}'))
    
    return [str(f) for f in sorted(genbank_files)]


def get_genome_name(genbank_file: str) -> str:
    """Extract genome/species name from GenBank file."""
    try:
        record = SeqIO.read(genbank_file, "genbank")
        organism = record.annotations.get('organism', None)
        
        if organism:
            return organism.replace(' ', '_')
        else:
            return Path(genbank_file).stem
    except:
        return Path(genbank_file).stem


def process_single_genbank_file(genbank_file: str, output_folder: str) -> Tuple[Optional[str], str, pd.DataFrame]:
    """Process a single GenBank file and generate RSCU analysis."""
    filename = os.path.basename(genbank_file)
    genome_name = get_genome_name(genbank_file)
    
    print(f"\nProcessing: {filename}")
    print(f"  Genome: {genome_name}")
    
    try:
        # Extract CDS and count codons
        coding_sequences = extract_coding_sequences(genbank_file)
        
        if not coding_sequences:
            print(f"  ⚠ WARNING: No coding sequences found!")
            return None, genome_name, pd.DataFrame()
        
        codon_counts, unusual_codons = count_codons(coding_sequences)
        
        # Calculate RSCU
        rscu_df = calculate_rscu(codon_counts, SYNONYMOUS_CODONS, AMINO_ACID_NAMES)
        
        # Prepare output for individual file
        output_data = rscu_df.copy()
        
        # Show amino acid name only on first row of each group
        prev_aa = None
        for idx, row in output_data.iterrows():
            if row['AA'] == prev_aa:
                output_data.at[idx, 'AA'] = ''
                output_data.at[idx, 'Amino_Acid'] = ''
            prev_aa = row['AA'] if row['AA'] else prev_aa
        
        # Save individual file
        output_file = os.path.join(output_folder, f"{genome_name}_RSCU.xlsx")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            output_data.to_excel(writer, sheet_name='RSCU_Analysis', index=False)
            
            # Format the worksheet
            workbook = writer.book
            worksheet = writer.sheets['RSCU_Analysis']
            
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # Header formatting
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data formatting
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                for col_idx, cell in enumerate(row):
                    cell.border = thin_border
                    
                    if col_idx <= 2:  # AA, Amino_Acid, Codon columns
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if col_idx == 1 and cell.value:  # Amino_Acid
                            cell.font = Font(bold=True, size=10)
                    elif col_idx == 3:  # Count column
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:  # RSCU column
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if cell.value is not None:
                            cell.number_format = '0.0000'
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 5   # AA
            worksheet.column_dimensions['B'].width = 15  # Amino_Acid
            worksheet.column_dimensions['C'].width = 8   # Codon
            worksheet.column_dimensions['D'].width = 10  # Count
            worksheet.column_dimensions['E'].width = 10  # RSCU
            
            worksheet.freeze_panes = 'A2'
        
        print(f"  ✓ RSCU analysis saved: {os.path.basename(output_file)}")
        return output_file, genome_name, rscu_df
        
    except Exception as e:
        print(f"  ✗ Error: {str(e)}")
        return None, genome_name, pd.DataFrame()


def merge_rscu_files(rscu_files: List[str], output_file: str):
    """Merge multiple RSCU files into comparative analysis."""
    print(f"\n{'='*70}")
    print("CREATING MERGED COMPARATIVE ANALYSIS")
    print(f"{'='*70}")
    
    all_data = []
    genome_names = []
    
    for rscu_file in rscu_files:
        try:
            df = pd.read_excel(rscu_file, sheet_name='RSCU_Analysis')
            
            # Get genome name from filename
            genome_name = os.path.basename(rscu_file).replace('_RSCU.xlsx', '')
            genome_names.append(genome_name)
            
            # Restore full dataframe for merging
            full_df = []
            current_aa = None
            current_aa_name = None
            
            for idx, row in df.iterrows():
                if row['AA'] and str(row['AA']).strip():
                    current_aa = row['AA']
                    current_aa_name = row['Amino_Acid']
                
                full_df.append({
                    'AA': current_aa,
                    'Amino_Acid': current_aa_name,
                    'Codon': row['Codon'],
                    'RSCU': row['RSCU']
                })
            
            all_data.append(pd.DataFrame(full_df))
            print(f"  - Loaded: {genome_name}")
            
        except Exception as e:
            print(f"  ✗ Error loading {os.path.basename(rscu_file)}: {e}")
            continue
    
    if not all_data:
        print("  ✗ No RSCU files could be loaded for merging")
        return
    
    # Create merged structure
    base_data = all_data[0][['AA', 'Amino_Acid', 'Codon']].copy()
    
    # Show amino acid name only on first occurrence
    display_amino_acid = []
    for idx, row in base_data.iterrows():
        if row['AA']:
            display_amino_acid.append(row['Amino_Acid'])
        else:
            display_amino_acid.append('')
    
    # Build final merged dataframe
    merged_data = pd.DataFrame()
    merged_data['Amino_Acid'] = display_amino_acid
    merged_data['Codon'] = base_data['Codon']
    
    # Add RSCU columns from each genome
    for genome_name, data in zip(genome_names, all_data):
        merged_data[genome_name] = data['RSCU'].values
    
    # Save merged file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        merged_data.to_excel(writer, sheet_name='Merged_RSCU', index=False)
        
        # Format worksheet
        workbook = writer.book
        worksheet = writer.sheets['Merged_RSCU']
        
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Make species name headers italic
        for col_idx in range(3, len(genome_names) + 3):
            header_cell = worksheet.cell(row=1, column=col_idx)
            header_cell.font = Font(bold=True, italic=True, color='FFFFFF', size=11)
        
        # Data formatting
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Track amino acid groups for merging
        merge_groups = []
        current_aa = None
        start_row = None
        
        # Format data rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            amino_acid_value = row[0].value
            
            # Track amino acid groups
            if amino_acid_value and str(amino_acid_value).strip():
                if current_aa is not None and start_row is not None:
                    end_row = row_idx - 1
                    if end_row > start_row:
                        merge_groups.append((start_row, end_row))
                
                current_aa = amino_acid_value
                start_row = row_idx
            
            # Format cells
            for col_idx, cell in enumerate(row):
                cell.border = thin_border
                
                if col_idx == 0:  # Amino_Acid
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if cell.value and str(cell.value).strip():
                        cell.font = Font(bold=True, size=10)
                elif col_idx == 1:  # Codon
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:  # RSCU values
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if cell.value is not None:
                        try:
                            cell.number_format = '0.0000'
                        except:
                            pass
        
        # Merge last group
        if current_aa is not None and start_row is not None:
            end_row = worksheet.max_row
            if end_row > start_row:
                merge_groups.append((start_row, end_row))
        
        # Merge cells for amino acids
        for start_row, end_row in merge_groups:
            worksheet.merge_cells(f'A{start_row}:A{end_row}')
            merged_cell = worksheet[f'A{start_row}']
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            merged_cell.font = Font(bold=True, size=10)
            merged_cell.border = thin_border
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 15  # Amino_Acid
        worksheet.column_dimensions['B'].width = 8   # Codon
        
        for col_idx in range(3, len(genome_names) + 3):
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[col_letter].width = 14
        
        worksheet.freeze_panes = 'C2'
    
    print(f"\n  ✓ Merged analysis saved: {os.path.basename(output_file)}")
    print(f"  - Total genomes compared: {len(genome_names)}")
    print(f"  - Total codons: {len(merged_data)}")


# ============================================================================
# R VISUALIZATION
# ============================================================================

def check_r_availability() -> bool:
    """Check if R is installed and available."""
    try:
        result = subprocess.run(['R', '--version'], 
                              capture_output=True, 
                              text=True, 
                              timeout=5)
        return result.returncode == 0
    except:
        return False


def check_r_packages() -> Tuple[bool, List[str]]:
    """Check if required R packages are installed."""
    required_packages = ['ggplot2', 'pheatmap', 'RColorBrewer', 'reshape2', 'grid', 'dplyr', 'tidyr', 'zoo']
    
    r_script = f"""
    required_packages <- c({', '.join([f"'{pkg}'" for pkg in required_packages])})
    installed <- installed.packages()[,"Package"]
    missing <- required_packages[!required_packages %in% installed]
    cat(paste(missing, collapse=","))
    """
    
    try:
        result = subprocess.run(['R', '--vanilla', '--slave', '-e', r_script],
                              capture_output=True,
                              text=True,
                              timeout=10)
        
        missing = [pkg.strip() for pkg in result.stdout.strip().split(',') if pkg.strip()]
        return len(missing) == 0, missing
    except:
        return False, required_packages


def install_r_packages(packages: List[str]) -> bool:
    """Attempt to install missing R packages."""
    print(f"\n  → Installing R packages: {', '.join(packages)}")
    
    r_script = f"""
    packages <- c({', '.join([f"'{pkg}'" for pkg in packages])})
    install.packages(packages, repos='http://cran.rstudio.com/', quiet=TRUE)
    """
    
    try:
        result = subprocess.run(['R', '--vanilla', '--slave', '-e', r_script],
                              capture_output=True,
                              text=True,
                              timeout=300)
        return result.returncode == 0
    except:
        return False


def generate_r_visualizations(merged_file: str, output_folder: str, genome_count: int):
    """Generate publication-quality visualizations using R."""
    print(f"\n{'='*70}")
    print("GENERATING R VISUALIZATIONS")
    print(f"{'='*70}")
    
    # Check R availability
    if not check_r_availability():
        print("  ⚠ R is not installed or not in PATH")
        print("  → Skipping visualization")
        print("  → To enable: Install R from https://www.r-project.org/")
        return
    
    print("  ✓ R is available")
    
    # Create figures folder first
    figures_folder = os.path.join(output_folder, FIGURES_FOLDER)
    os.makedirs(figures_folder, exist_ok=True)
    
    # Check required packages
    packages_ok, missing_packages = check_r_packages()
    
    if not packages_ok:
        print(f"  ⚠ Missing R packages: {', '.join(missing_packages)}")
        print(f"  → Attempting automatic installation...")
        
        if install_r_packages(missing_packages):
            print(f"  ✓ Successfully installed R packages")
        else:
            print(f"  ✗ Failed to install packages automatically")
            print(f"  → Please install manually in R:")
            pkg_list = "', '".join(missing_packages)
            print(f"     install.packages(c('{pkg_list}'))")
            return
    else:
        print("  ✓ All required R packages are installed")
    
    # First, need to convert Excel to CSV for R (keep both files)
    print(f"  → Converting Excel to CSV...")
    try:
        df = pd.read_excel(merged_file, sheet_name='Merged_RSCU')
        csv_file = merged_file.replace('.xlsx', '.csv')
        df.to_csv(csv_file, index=False)
        print(f"  ✓ Created CSV for R processing: {os.path.basename(csv_file)}")
        print(f"    - Rows: {len(df)}, Columns: {len(df.columns)}")
    except Exception as e:
        print(f"  ✗ Error converting to CSV: {e}")
        return
    
    # Generate R script
    r_script_path = os.path.join(output_folder, "generate_plots.R")
    
    # Determine optimal figure dimensions based on genome count
    heatmap_width = max(8, min(20, 6 + genome_count * 0.3))
    heatmap_height = 12
    
    r_script_content = f'''
# CGAS Module 8: R Visualization Script  
# Generated automatically - Publication-quality RSCU visualizations

cat("\\n========================================\\n")
cat("CGAS MODULE 8: R VISUALIZATION\\n")
cat("========================================\\n\\n")

# Set working directory
setwd("{os.path.abspath(output_folder)}")
cat("Working directory:", getwd(), "\\n")

# Read merged RSCU data from CSV
cat("Reading merged RSCU data from CSV...\\n")
csv_file <- "{os.path.basename(csv_file)}"
cat("Looking for CSV file:", csv_file, "\\n")

if (file.exists(csv_file)) {{
  cat(paste("  ✓ CSV file found:", csv_file, "\\n"))
  data <- read.csv(csv_file, check.names=FALSE, stringsAsFactors=FALSE)
}} else {{
  cat("  ✗ ERROR: CSV file not found!\\n")
  cat("  Current files in directory:\\n")
  print(list.files())
  quit(save="no", status=1)
}}

cat(paste("  ✓ Loaded data:", nrow(data), "rows,", ncol(data), "columns\\n\\n"))

# Clean up species names: replace underscores with spaces for italics
colnames(data)[3:ncol(data)] <- gsub("_", " ", colnames(data)[3:ncol(data)])

# Prepare data for heatmap
heatmap_data <- data[, -c(1,2)]
rownames(heatmap_data) <- data$Codon

# Convert to numeric matrix
heatmap_matrix <- as.matrix(heatmap_data)
heatmap_matrix <- apply(heatmap_matrix, 2, as.numeric)
rownames(heatmap_matrix) <- data$Codon

# Replace any NA or infinite values
heatmap_matrix[is.na(heatmap_matrix)] <- 0
heatmap_matrix[is.infinite(heatmap_matrix)] <- 0

cat("Data summary:\\n")
cat(paste("  - Codons:", nrow(heatmap_matrix), "\\n"))
cat(paste("  - Genomes:", ncol(heatmap_matrix), "\\n"))
cat(paste("  - RSCU range:", round(min(heatmap_matrix), 2), "-", 
          round(max(heatmap_matrix), 2), "\\n\\n"))

# Create amino acid grouping for row labels
library(zoo)

# Convert full amino acid names to 3-letter codes
aa_name_to_code <- c(
  "Alanine" = "Ala", "Arginine" = "Arg", "Asparagine" = "Asn",
  "Aspartate" = "Asp", "Cysteine" = "Cys", "Glutamine" = "Gln",
  "Glutamate" = "Glu", "Glycine" = "Gly", "Histidine" = "His",
  "Isoleucine" = "Ile", "Leucine" = "Leu", "Lysine" = "Lys",
  "Methionine" = "Met", "Phenylalanine" = "Phe", "Proline" = "Pro",
  "Serine" = "Ser", "Threonine" = "Thr", "Tryptophan" = "Trp",
  "Tyrosine" = "Tyr", "Valine" = "Val", "Stop" = "Stop"
)

amino_acids <- data$Amino_Acid
amino_acids[amino_acids == ""] <- NA
amino_acids <- na.locf(amino_acids)  # Fill forward with zoo package

# Convert to 3-letter codes
amino_acids_3letter <- aa_name_to_code[amino_acids]
amino_acids_3letter[is.na(amino_acids_3letter)] <- amino_acids[is.na(amino_acids_3letter)]

# Create row labels that show amino acid name only once per group
row_labels <- character(length(data$Codon))
for (i in 1:length(data$Codon)) {{
  if (i == 1 || amino_acids_3letter[i] != amino_acids_3letter[i-1]) {{
    # First occurrence of this amino acid - show the name
    row_labels[i] <- paste0(amino_acids_3letter[i], " - ", data$Codon[i])
  }} else {{
    # Not first occurrence - show just codon with spacing
    row_labels[i] <- paste0("       ", data$Codon[i])  # Indent to align with codons
  }}
}}

# Update rownames to include amino acid labels
rownames(heatmap_matrix) <- row_labels
# Find positions where amino acid changes (for gaps)
aa_change_positions <- which(amino_acids_3letter[-1] != amino_acids_3letter[-length(amino_acids_3letter)])

# Define colors for gaps between amino acid groups (for visual separation)
aa_colors <- c(
  "Ala" = "#E6F5FF", "Arg" = "#FFE6E6", "Asn" = "#E6FFE6",
  "Asp" = "#FFFFE6", "Cys" = "#FFE6FF", "Gln" = "#E6FFFF",
  "Glu" = "#FFF0E6", "Gly" = "#F0E6FF", "His" = "#E6F0FF",
  "Ile" = "#FFE6F0", "Leu" = "#F0FFE6", "Lys" = "#FFEBCD",
  "Met" = "#E6E6FA", "Phe" = "#FFF0F5", "Pro" = "#F0FFF0",
  "Ser" = "#FFFACD", "Thr" = "#F5F5DC", "Trp" = "#FDF5E6",
  "Tyr" = "#FAF0E6", "Val" = "#F0FFFF", "Stop" = "#D3D3D3"
)

# ============================================================================
# FIGURE 1: RSCU HEATMAP WITH AMINO ACID NAMES ON LEFT
# ============================================================================

cat("Generating RSCU heatmap with amino acid labels...\\n")

library(pheatmap)
library(RColorBrewer)

# Color palette for RSCU values
color_palette <- colorRampPalette(c("navy", "blue", "white", "red", "darkred"))(100)

# Custom label function for italicizing species names in heatmap
make_italic <- function(x) {{
  as.expression(lapply(x, function(label) bquote(italic(.(label)))))
}}

# Generate heatmap with amino acid labels on left
pdf("Figures/RSCU_Heatmap.pdf", width={heatmap_width}, height={heatmap_height})
pheatmap(heatmap_matrix,
         main="Relative Synonymous Codon Usage (RSCU)",
         color=color_palette,
         breaks=seq(0, max(heatmap_matrix, na.rm=TRUE), length.out=101),
         cluster_rows=FALSE,  # Don't cluster - keep amino acid grouping
         cluster_cols=TRUE,   # Cluster genomes
         clustering_distance_cols="euclidean",
         clustering_method="complete",
         labels_col=make_italic(colnames(heatmap_matrix)),  # Italicize species names
         fontsize=10,
         fontsize_row=6,      # Slightly smaller for better fit
         fontsize_col=9,
         angle_col=90,
         cellwidth=NA,
         cellheight=10,       # Fixed height for alignment
         border_color="grey80",  # Add grid lines
         gaps_row=aa_change_positions,  # Add gaps between amino acid groups
         legend=TRUE,
         legend_breaks=seq(0, max(heatmap_matrix, na.rm=TRUE), length.out=5),
         legend_labels=round(seq(0, max(heatmap_matrix, na.rm=TRUE), length.out=5), 2),
         show_rownames=TRUE,
         show_colnames=TRUE,
         treeheight_col=30)
dev.off()

# Generate PNG version (600 DPI)
png("Figures/RSCU_Heatmap.png", 
    width={heatmap_width}, 
    height={heatmap_height}, 
    units="in",
    res=600)
pheatmap(heatmap_matrix,
         main="Relative Synonymous Codon Usage (RSCU)",
         color=color_palette,
         breaks=seq(0, max(heatmap_matrix, na.rm=TRUE), length.out=101),
         cluster_rows=FALSE,
         cluster_cols=TRUE,
         clustering_distance_cols="euclidean",
         clustering_method="complete",
         labels_col=make_italic(colnames(heatmap_matrix)),
         fontsize=12,
         fontsize_row=9,
         fontsize_col=9,
         angle_col=90,
         cellwidth=NA,
         cellheight=10,
         border_color="grey80",  # Add grid lines
         gaps_row=aa_change_positions,  # Add gaps between amino acid groups
         legend=TRUE,
         legend_breaks=seq(0, max(heatmap_matrix, na.rm=TRUE), length.out=5),
         legend_labels=round(seq(0, max(heatmap_matrix, na.rm=TRUE), length.out=5), 2),
         show_rownames=TRUE,
         show_colnames=TRUE,
         treeheight_col=30)

cat("  ✓ RSCU heatmap saved: Figures/RSCU_Heatmap.pdf and .png\\n\\n")

# ============================================================================
# FIGURE 2: STACKED BAR PLOT - THREE REPRESENTATIVE SPECIES ONLY
# ============================================================================

cat("Generating stacked bar plot for 3 representative species...\\n")

library(ggplot2)
library(dplyr)
library(tidyr)

# Select 3 representative species
n_genomes <- ncol(heatmap_matrix)

if (n_genomes >= 3) {{
  # Cluster genomes to find representatives
  genome_dist <- dist(t(heatmap_matrix), method="euclidean")
  genome_clust <- hclust(genome_dist, method="complete")
  genome_order <- genome_clust$order
  
  # Select 3 representative genomes: first, middle, last from cluster
  if (n_genomes == 3) {{
    selected_idx <- c(1, 2, 3)
  }} else if (n_genomes < 10) {{
    # For small datasets, pick evenly spaced
    selected_idx <- c(genome_order[1], 
                      genome_order[round(n_genomes/2)], 
                      genome_order[n_genomes])
  }} else {{
    # For large datasets, pick from different clusters
    selected_idx <- c(genome_order[1], 
                      genome_order[round(n_genomes/2)], 
                      genome_order[n_genomes])
  }}
  
  selected_species <- colnames(heatmap_matrix)[selected_idx]
  cat(paste("  → Selected 3 representative species:\\n"))
  for (sp in selected_species) {{
    cat(paste("     -", sp, "\\n"))
  }}
  
}} else {{
  # If less than 3 genomes, use all
  selected_species <- colnames(heatmap_matrix)
  cat(paste("  → Using all", n_genomes, "species\\n"))
}}

# Filter data for selected species
data_subset <- data[, c("Amino_Acid", "Codon", selected_species)]

# Convert to long format
data_long <- data_subset %>%
  pivot_longer(
    cols = -c(Amino_Acid, Codon),
    names_to = "Species",
    values_to = "RSCU"
  ) %>%
  mutate(
    Species = factor(Species, levels = selected_species),
    AA = amino_acids_3letter[match(Codon, data$Codon)],
    AA = factor(AA, levels = unique(amino_acids_3letter)),
    Codon = factor(Codon, levels = unique(Codon))
  )

# Remove rows with NA in AA
data_long <- data_long[!is.na(data_long$AA),]

# Compute label positions for stacked bars
data_long <- data_long %>%
  group_by(AA, Species) %>%
  arrange(Codon) %>%
  mutate(
    cum_RSCU = cumsum(RSCU),
    ypos = cum_RSCU - RSCU / 2
  ) %>%
  ungroup()

# Define colors for 3 species
species_colors <- c("#87CEFA", "#FFA07A", "#90EE90")
if (length(selected_species) != 3) {{
  # If not exactly 3, create color palette
  species_colors <- brewer.pal(max(3, length(selected_species)), "Set3")[1:length(selected_species)]
}}

# Create plot
p <- ggplot(data_long, aes(x = Species, y = RSCU, fill = Species)) +
  geom_bar(stat = "identity", width = 0.7, color = "black", linewidth = 0.5) +
  geom_text(aes(label = Codon, y = ypos), size = 5, angle = 90, vjust = 0.5, fontface = "bold") +
  facet_wrap(~ AA, nrow = 1, strip.position = "bottom") +
  scale_fill_manual(values = species_colors, name = "") +
  scale_y_continuous(breaks = seq(0, ceiling(max(data_long$cum_RSCU, na.rm=TRUE)), 1), 
                     expand = expansion(mult = c(0, 0.05))) +
  labs(y = "Relative Synonymous Codon Usage", x = "Amino Acid") +
  theme_minimal(base_size = 20) +
  theme(
    axis.text.x = element_blank(),
    axis.ticks.x = element_blank(),
    strip.text = element_text(size = 18, face = "bold"),
    panel.grid.major.y = element_line(color = "gray90", linewidth = 0.5),
    panel.grid.minor = element_blank(),
    panel.border = element_rect(color = "black", fill = NA, linewidth = 0.8),
    axis.line = element_line(color = "black", linewidth = 0.8),
    axis.ticks.y = element_line(color = "black", linewidth = 0.8),
    legend.position = "top",
    legend.text = element_text(size = 18, face = "italic"),
    legend.key.size = unit(1.5, "cm"),
    plot.background = element_rect(fill = "white", color = NA),
    plot.margin = margin(10, 10, 10, 10)
  ) +
  geom_hline(yintercept = 0, color = "black", linewidth = 0.8)

# Save plot
pdf("Figures/Codon_Usage_Comparison.pdf", width = 28, height = 12)
print(p)
dev.off()

# Save PNG version (600 DPI)
ggsave("Figures/Codon_Usage_Comparison.png", plot = p,
       width = 28, height = 12, dpi = 600)

cat("  ✓ Codon usage comparison saved: Figures/Codon_Usage_Comparison.pdf and .png\\n\\n")

# ============================================================================
# SUMMARY
# ============================================================================

cat("========================================\\n")
cat("All visualizations generated successfully!\\n")
cat("========================================\\n\\n")
cat("Files created:\\n")
cat("  1. RSCU_Heatmap.pdf and .png (all genomes, 600 DPI)\\n")
cat("  2. Codon_Usage_Comparison.pdf and .png (3 representative species, 600 DPI)\\n\\n")
cat("Representative species for comparison:\\n")
for (sp in selected_species) {{
  cat(paste("  -", sp, "\\n"))
}}
cat("\\n")
'''
    
    # Save R script
    with open(r_script_path, 'w') as f:
        f.write(r_script_content)
    
    print(f"  ✓ R script created: {os.path.basename(r_script_path)}")
    
    # Execute R script
    print(f"  → Executing R script...")
    try:
        result = subprocess.run(['Rscript', r_script_path],
                              capture_output=True,
                              text=True,
                              timeout=300,
                              cwd=output_folder)
        
        if result.returncode == 0:
            print(f"  ✓ R visualization completed successfully")
            print(f"\n  Generated figures in {FIGURES_FOLDER}/:")
            print(f"    - RSCU_Heatmap.pdf and .png (600 DPI)")
            print(f"    - Codon_Usage_Comparison.pdf and .png (600 DPI)")
            
            # Check if files were actually created
            import glob
            figures = glob.glob(os.path.join(figures_folder, "*"))
            if figures:
                print(f"\n  Files created:")
                for fig in sorted(figures):
                    size = os.path.getsize(fig)
                    print(f"    ✓ {os.path.basename(fig)} ({size:,} bytes)")
                
            if os.path.exists(comparison_pdf):
                print(f"  ✓ Codon_Usage_Comparison.pdf created ({os.path.getsize(comparison_pdf):,} bytes)")
            else:
                print(f"  ⚠ Codon_Usage_Comparison.pdf not found!")
                
        else:
            print(f"  ✗ R script execution failed with return code {result.returncode}")
            print(f"  Error output:")
            if result.stderr:
                print(result.stderr[:1000])
            if result.stdout:
                # Look for error messages in stdout
                error_lines = [line for line in result.stdout.split('\n') if 'ERROR' in line or 'error' in line]
                if error_lines:
                    print("Errors in output:")
                    for line in error_lines[:5]:
                        print(f"  {line}")
    
    except subprocess.TimeoutExpired:
        print(f"  ✗ R script execution timed out (>5 minutes)")
        print(f"  → This may happen with very large datasets")
    except Exception as e:
        print(f"  ✗ Error executing R script: {e}")
        import traceback
        traceback.print_exc()


# ============================================================================
# MAIN PIPELINE
# ============================================================================

def main():
    """Main pipeline execution function."""
    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description='CGAS Module 8: Codon Usage Analysis (RSCU) with R Visualization',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  %(prog)s                          # Process current directory
  %(prog)s -i genbank_files/        # Process specific directory
  %(prog)s -i data/ -o results/     # Custom input and output
  %(prog)s --no-figures             # Skip R visualization

For more information, see the CGAS documentation.
        '''
    )
    
    parser.add_argument('-i', '--input',
                       type=str,
                       default='.',
                       help='Input directory containing GenBank files (default: current directory)')
    
    parser.add_argument('-o', '--output',
                       type=str,
                       default=None,
                       help='Output directory (default: Module8_Codon_Usage_Analysis in input directory)')
    
    parser.add_argument('--no-figures',
                       action='store_true',
                       help='Skip R figure generation')
    
    args = parser.parse_args()
    
    # Set up directories
    input_dir = os.path.abspath(args.input)
    
    if args.output:
        output_dir = os.path.abspath(args.output)
    else:
        output_dir = os.path.join(input_dir, OUTPUT_FOLDER)
    
    print(f"\n{'='*70}")
    print("CGAS MODULE 8: CODON USAGE ANALYSIS (RSCU)")
    print(f"{'='*70}")
    print(f"\nInput directory: {input_dir}")
    print(f"Output directory: {output_dir}")
    
    # Create output folder
    os.makedirs(output_dir, exist_ok=True)
    
    # Find GenBank files
    genbank_files = find_genbank_files(input_dir)
    
    if not genbank_files:
        print("\n❌ ERROR: No GenBank files found!")
        print("Please place GenBank files (.gb, .gbf, .gbk) in the input directory.")
        return
    
    print(f"\nFound {len(genbank_files)} GenBank file(s):")
    for gf in genbank_files[:10]:  # Show first 10
        print(f"  - {os.path.basename(gf)}")
    if len(genbank_files) > 10:
        print(f"  ... and {len(genbank_files) - 10} more")
    
    # Process each GenBank file
    print(f"\n{'='*70}")
    print("CALCULATING RSCU VALUES")
    print(f"{'='*70}")
    
    rscu_files = []
    for gb_file in genbank_files:
        try:
            output_file, _, _ = process_single_genbank_file(gb_file, output_dir)
            if output_file:
                rscu_files.append(output_file)
        except Exception as e:
            print(f"  ❌ ERROR processing {os.path.basename(gb_file)}: {e}")
            continue
    
    # Merge RSCU files
    if rscu_files:
        merged_file = os.path.join(output_dir, "Complete_Codon_Usage_Analysis.xlsx")
        merge_rscu_files(rscu_files, merged_file)
        
        # Generate R visualizations
        if not args.no_figures:
            generate_r_visualizations(merged_file, output_dir, len(rscu_files))
        else:
            print(f"\n  → Skipping R visualization (--no-figures flag)")
    
    # Summary
    print(f"\n{'='*70}")
    print("CGAS MODULE 8 COMPLETE")
    print(f"{'='*70}")
    print(f"✓ Successfully processed: {len(rscu_files)} genome(s)")
    print(f"✓ Individual RSCU files: {len(rscu_files)}")
    if rscu_files:
        print(f"✓ Merged analysis (Excel): Complete_Codon_Usage_Analysis.xlsx")
        print(f"✓ Merged analysis (CSV): Complete_Codon_Usage_Analysis.csv")
        if not args.no_figures:
            print(f"✓ R visualizations: {FIGURES_FOLDER}/")
            print(f"    - RSCU_Heatmap.pdf")
            print(f"    - Codon_Usage_Comparison.pdf")
    print(f"\nAll output files saved in: {output_dir}/")
    print(f"{'='*70}\n")


# ============================================================================
# SCRIPT ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠ Analysis interrupted by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\n\n❌ FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
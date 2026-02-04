#!/usr/bin/env python3
"""
CGAS Module 9: Comprehensive Amino Acid Analysis with R Visualization
=====================================================================

This script performs comprehensive amino acid analysis on chloroplast genomes by:
1. Extracting and translating coding sequences (CDS) from GenBank files
2. Calculating amino acid composition percentages
3. Generating individual composition reports for each genome
4. Creating merged comparative analysis across all genomes
5. Generating publication-quality R-based visualizations:
   - Heatmap of amino acid composition across all genomes
   - Bar plot comparison of 3 representative species

Author: Abdullah
Version: 1.0.1
Date: January 2026

Dependencies:
    Python: biopython, pandas, openpyxl
    R: ggplot2, pheatmap, RColorBrewer, reshape2, dplyr, tidyr

Usage:
    python cgas_module9.py
    python cgas_module9.py -i genbank_files/
    python cgas_module9.py -i genbank_files/ -o custom_output/
    python cgas_module9.py --no-figures  # Skip R visualization

Output:
    Module9_Amino_Acid_Analysis/
        - Individual files: [genome_name]_AminoAcid.xlsx
        - Merged analysis: Complete_Amino_Acid_Analysis.xlsx
        - Figures/ (if R available):
            - AA_Composition_Heatmap.pdf
            - Amino_Acid_Comparison.pdf
"""

import os
import sys
import argparse
import subprocess
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from collections import Counter

try:
    from Bio import SeqIO
    from Bio.Seq import Seq
    from Bio.SeqUtils import ProtParam
    import pandas as pd
except ImportError as e:
    print(f"Error: Required package not installed: {e}")
    print("Please install required packages using:")
    print("pip install biopython pandas openpyxl --break-system-packages")
    sys.exit(1)


# ============================================================================
# CONSTANTS
# ============================================================================

OUTPUT_FOLDER = "Module9_Amino_Acid_Analysis"
FIGURES_FOLDER = "Figures"

# Full amino acid names mapping
AMINO_ACID_NAMES = {
    'A': 'Alanine',
    'R': 'Arginine',
    'N': 'Asparagine',
    'D': 'Aspartate',
    'C': 'Cysteine',
    'Q': 'Glutamine',
    'E': 'Glutamate',
    'G': 'Glycine',
    'H': 'Histidine',
    'I': 'Isoleucine',
    'L': 'Leucine',
    'K': 'Lysine',
    'M': 'Methionine',
    'F': 'Phenylalanine',
    'P': 'Proline',
    'S': 'Serine',
    'T': 'Threonine',
    'W': 'Tryptophan',
    'Y': 'Tyrosine',
    'V': 'Valine'
}

# Amino acid order for publication (alphabetical)
AA_ORDER = ['A', 'R', 'N', 'D', 'C', 'Q', 'E', 'G', 'H', 'I', 
            'L', 'K', 'M', 'F', 'P', 'S', 'T', 'W', 'Y', 'V']

# Color scheme for 3 species (matches codon usage analysis)
SPECIES_COLORS = {
    'species1': '#0072B2',  # Blue
    'species2': '#E69F00',  # Orange
    'species3': '#009E73'   # Green
}


# ============================================================================
# SEQUENCE EXTRACTION AND TRANSLATION
# ============================================================================

def find_genbank_files(directory: str) -> List[str]:
    """Find all GenBank files in the specified directory."""
    extensions = ['.gb', '.gbf', '.gbk', '.genbank']
    genbank_files = []
    
    for ext in extensions:
        genbank_files.extend(Path(directory).glob(f'*{ext}'))
    
    return [str(f) for f in sorted(genbank_files)]


def get_genome_name(genbank_file: str) -> str:
    """Extract genome/species name from GenBank file."""
    try:
        record = SeqIO.read(genbank_file, "genbank")
        organism = record.annotations.get('organism', None)
        
        if organism:
            return organism.replace(' ', '_')
        else:
            return Path(genbank_file).stem
    except:
        return Path(genbank_file).stem


def extract_and_translate_cds(genbank_file: str) -> Tuple[str, int]:
    """
    Extract all CDS features and translate them to amino acid sequences.
    
    Returns:
        combined_amino_acid_sequence: str - Combined amino acid sequence
        cds_count: int - Number of CDS features processed
    """
    all_amino_acid_sequences = []
    cds_count = 0
    
    try:
        record = SeqIO.read(genbank_file, "genbank")
        
        for feature in record.features:
            if feature.type == "CDS":
                try:
                    # Extract CDS sequence
                    cds_seq = feature.extract(record.seq)
                    
                    # Translate to amino acids (stop at first stop codon)
                    amino_acid_sequence = cds_seq.translate(to_stop=True)
                    
                    all_amino_acid_sequences.append(str(amino_acid_sequence))
                    cds_count += 1
                    
                except Exception as e:
                    print(f"    ⚠ Warning: Error translating CDS: {e}")
                    continue
        
        print(f"    - Translated {cds_count} coding sequences")
        
    except Exception as e:
        print(f"    ✗ Error reading GenBank file: {e}")
        return "", 0
    
    # Combine all amino acid sequences
    combined_sequence = "".join(all_amino_acid_sequences)
    
    return combined_sequence, cds_count


# ============================================================================
# AMINO ACID COMPOSITION ANALYSIS
# ============================================================================

def analyze_amino_acid_composition(amino_acid_sequence: str) -> pd.DataFrame:
    """
    Analyze amino acid composition and calculate percentages.
    
    Returns:
        DataFrame with columns: AA_Code, Amino_Acid, Percentage
    """
    if not amino_acid_sequence:
        return pd.DataFrame()
    
    # Use Biopython's ProteinAnalysis
    protein_analysis = ProtParam.ProteinAnalysis(amino_acid_sequence)
    composition = protein_analysis.amino_acids_percent
    
    # Convert to structured data
    composition_data = []
    for aa_code in AA_ORDER:
        if aa_code in composition:
            percentage = composition[aa_code]  # Already in percentage (0-100)
            composition_data.append({
                'AA_Code': aa_code,
                'Amino_Acid': AMINO_ACID_NAMES[aa_code],
                'Percentage': round(percentage, 4)
            })
    
    df = pd.DataFrame(composition_data)
    
    total_aa = len(amino_acid_sequence)
    print(f"    - Analyzed {total_aa:,} amino acids")
    
    return df


# ============================================================================
# FILE PROCESSING
# ============================================================================

def process_single_genbank_file(genbank_file: str, output_folder: str) -> Tuple[Optional[str], str, pd.DataFrame]:
    """Process a single GenBank file and generate amino acid analysis."""
    filename = os.path.basename(genbank_file)
    genome_name = get_genome_name(genbank_file)
    
    print(f"\nProcessing: {filename}")
    print(f"  Genome: {genome_name}")
    
    try:
        # Extract and translate CDS
        amino_acid_sequence, cds_count = extract_and_translate_cds(genbank_file)
        
        if not amino_acid_sequence or cds_count == 0:
            print(f"  ⚠ WARNING: No coding sequences found or failed to translate!")
            return None, genome_name, pd.DataFrame()
        
        # Analyze composition
        aa_df = analyze_amino_acid_composition(amino_acid_sequence)
        
        if aa_df.empty:
            return None, genome_name, pd.DataFrame()
        
        # Prepare output for individual file
        output_data = aa_df.copy()
        
        # Save individual file
        output_file = os.path.join(output_folder, f"{genome_name}_AminoAcid.xlsx")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Main composition sheet
            output_data.to_excel(writer, sheet_name='AA_Composition', index=False)
            
            # Summary sheet
            summary_data = {
                'Metric': [
                    'Organism',
                    'Total CDS sequences',
                    'Total amino acids',
                    'Number of AA types'
                ],
                'Value': [
                    genome_name.replace('_', ' '),
                    cds_count,
                    len(amino_acid_sequence),
                    len(output_data)
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format the worksheet
            workbook = writer.book
            worksheet = writer.sheets['AA_Composition']
            
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # Header formatting
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data formatting
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                for col_idx, cell in enumerate(row):
                    cell.border = thin_border
                    
                    if col_idx == 0:  # AA_Code column
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(bold=True, size=10)
                    elif col_idx == 1:  # Amino_Acid column
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.font = Font(size=10)
                    else:  # Percentage column
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if cell.value is not None:
                            cell.number_format = '0.0000'
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 8    # AA_Code
            worksheet.column_dimensions['B'].width = 15   # Amino_Acid
            worksheet.column_dimensions['C'].width = 12   # Percentage
            
            worksheet.freeze_panes = 'A2'
        
        print(f"  ✓ Amino acid analysis saved: {os.path.basename(output_file)}")
        return output_file, genome_name, aa_df
        
    except Exception as e:
        print(f"  ✗ Error: {str(e)}")
        return None, genome_name, pd.DataFrame()


def merge_aa_files(aa_files: List[str], output_file: str):
    """Merge multiple amino acid files into comparative analysis."""
    print(f"\n{'='*70}")
    print("CREATING MERGED COMPARATIVE ANALYSIS")
    print(f"{'='*70}")
    
    all_data = []
    genome_names = []
    
    for aa_file in aa_files:
        try:
            df = pd.read_excel(aa_file, sheet_name='AA_Composition')
            
            # Get genome name from filename
            genome_name = os.path.basename(aa_file).replace('_AminoAcid.xlsx', '')
            genome_names.append(genome_name)
            
            all_data.append(df)
            print(f"  - Loaded: {genome_name}")
            
        except Exception as e:
            print(f"  ✗ Error loading {os.path.basename(aa_file)}: {e}")
            continue
    
    if not all_data:
        print("  ✗ No amino acid files could be loaded for merging")
        return
    
    # Create merged structure
    base_data = all_data[0][['AA_Code', 'Amino_Acid']].copy()
    
    # Build final merged dataframe
    merged_data = pd.DataFrame()
    merged_data['AA_Code'] = base_data['AA_Code']
    merged_data['Amino_Acid'] = base_data['Amino_Acid']
    
    # Add percentage columns from each genome
    for genome_name, data in zip(genome_names, all_data):
        merged_data[genome_name] = data['Percentage'].values
    
    # Save merged file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        merged_data.to_excel(writer, sheet_name='Merged_AA_Composition', index=False)
        
        # Format worksheet
        workbook = writer.book
        worksheet = writer.sheets['Merged_AA_Composition']
        
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Make species name headers italic
        for col_idx in range(3, len(genome_names) + 3):
            header_cell = worksheet.cell(row=1, column=col_idx)
            header_cell.font = Font(bold=True, italic=True, color='FFFFFF', size=11)
        
        # Data formatting
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format data rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            for col_idx, cell in enumerate(row):
                cell.border = thin_border
                
                if col_idx == 0:  # AA_Code
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True, size=10)
                elif col_idx == 1:  # Amino_Acid
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # Percentage values
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if cell.value is not None:
                        try:
                            cell.number_format = '0.0000'
                        except:
                            pass
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 8    # AA_Code
        worksheet.column_dimensions['B'].width = 15   # Amino_Acid
        
        for col_idx in range(3, len(genome_names) + 3):
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[col_letter].width = 15
        
        worksheet.freeze_panes = 'C2'
    
    print(f"\n  ✓ Merged analysis saved: {os.path.basename(output_file)}")
    print(f"  - Total genomes compared: {len(genome_names)}")
    print(f"  - Total amino acids: {len(merged_data)}")


# ============================================================================
# R VISUALIZATION
# ============================================================================

def check_r_availability() -> bool:
    """Check if R is installed and available."""
    try:
        result = subprocess.run(['R', '--version'], 
                              capture_output=True, 
                              text=True, 
                              timeout=5)
        return result.returncode == 0
    except:
        return False


def check_r_packages() -> Tuple[bool, List[str]]:
    """Check if required R packages are installed."""
    required_packages = ['ggplot2', 'pheatmap', 'RColorBrewer', 'reshape2', 'dplyr', 'tidyr']
    
    r_script = f"""
    required_packages <- c({', '.join([f"'{pkg}'" for pkg in required_packages])})
    installed <- installed.packages()[,"Package"]
    missing <- required_packages[!required_packages %in% installed]
    cat(paste(missing, collapse=","))
    """
    
    try:
        result = subprocess.run(['R', '--vanilla', '--slave', '-e', r_script],
                              capture_output=True,
                              text=True,
                              timeout=10)
        
        missing = [pkg.strip() for pkg in result.stdout.strip().split(',') if pkg.strip()]
        return len(missing) == 0, missing
    except:
        return False, required_packages


def install_r_packages(packages: List[str]) -> bool:
    """Attempt to install missing R packages."""
    print(f"\n  → Installing R packages: {', '.join(packages)}")
    
    r_script = f"""
    packages <- c({', '.join([f"'{pkg}'" for pkg in packages])})
    install.packages(packages, repos='http://cran.rstudio.com/', quiet=TRUE)
    """
    
    try:
        result = subprocess.run(['R', '--vanilla', '--slave', '-e', r_script],
                              capture_output=True,
                              text=True,
                              timeout=300)
        return result.returncode == 0
    except:
        return False


def generate_r_visualizations(merged_file: str, output_folder: str, genome_count: int):
    """Generate publication-quality visualizations using R."""
    print(f"\n{'='*70}")
    print("GENERATING R VISUALIZATIONS")
    print(f"{'='*70}")
    
    # Check R availability
    if not check_r_availability():
        print("  ⚠ R is not installed or not in PATH")
        print("  → Skipping visualization")
        print("  → To enable: Install R from https://www.r-project.org/")
        return
    
    print("  ✓ R is available")
    
    # Create figures folder first
    figures_folder = os.path.join(output_folder, FIGURES_FOLDER)
    os.makedirs(figures_folder, exist_ok=True)
    
    # Check required packages
    packages_ok, missing_packages = check_r_packages()
    
    if not packages_ok:
        print(f"  ⚠ Missing R packages: {', '.join(missing_packages)}")
        print(f"  → Attempting automatic installation...")
        
        if install_r_packages(missing_packages):
            print(f"  ✓ Successfully installed R packages")
        else:
            print(f"  ✗ Failed to install packages automatically")
            print(f"  → Please install manually in R:")
            pkg_list = "', '".join(missing_packages)
            print(f"     install.packages(c('{pkg_list}'))")
            return
    else:
        print("  ✓ All required R packages are installed")
    
    # Convert Excel to CSV for R
    print(f"  → Converting Excel to CSV...")
    try:
        df = pd.read_excel(merged_file, sheet_name='Merged_AA_Composition')
        csv_file = merged_file.replace('.xlsx', '.csv')
        df.to_csv(csv_file, index=False)
        print(f"  ✓ Created CSV for R processing: {os.path.basename(csv_file)}")
        print(f"    - Rows: {len(df)}, Columns: {len(df.columns)}")
    except Exception as e:
        print(f"  ✗ Error converting to CSV: {e}")
        return
    
    # Generate R script
    r_script_path = os.path.join(output_folder, "generate_aa_plots.R")
    
    # Determine optimal figure dimensions
    heatmap_width = max(8, min(20, 6 + genome_count * 0.3))
    heatmap_height = 8
    
    r_script_content = f'''
# CGAS Module 9: R Visualization Script  
# Generated automatically - Publication-quality amino acid visualizations

suppressPackageStartupMessages({{
  library(ggplot2)
  library(pheatmap)
  library(RColorBrewer)
  library(reshape2)
  library(dplyr)
  library(tidyr)
}})

# Set working directory
setwd("{os.path.abspath(output_folder)}")

cat("\\n========================================\\n")
cat("CGAS MODULE 9: AMINO ACID VISUALIZATION\\n")
cat("========================================\\n\\n")

# Read merged amino acid data from CSV
cat("Reading merged amino acid data from CSV...\\n")
csv_file <- "{os.path.basename(csv_file)}"

if (file.exists(csv_file)) {{
  cat(paste("  - Reading from CSV:", csv_file, "\\n"))
  data <- read.csv(csv_file, check.names=FALSE, stringsAsFactors=FALSE)
}} else {{
  stop("ERROR: CSV file not found!")
}}

cat(paste("  ✓ Loaded data:", nrow(data), "rows,", ncol(data), "columns\\n\\n"))

# Clean up species names: replace underscores with spaces for italics
species_cols <- colnames(data)[3:ncol(data)]
colnames(data)[3:ncol(data)] <- gsub("_", " ", species_cols)

# Prepare data for heatmap
heatmap_data <- data[, -c(1,2)]
rownames(heatmap_data) <- data$Amino_Acid  # Only full amino acid names

# Convert to numeric matrix
heatmap_matrix <- as.matrix(heatmap_data)
heatmap_matrix <- apply(heatmap_matrix, 2, as.numeric)
rownames(heatmap_matrix) <- data$Amino_Acid  # Only full amino acid names

# Replace any NA or infinite values
heatmap_matrix[is.na(heatmap_matrix)] <- 0
heatmap_matrix[is.infinite(heatmap_matrix)] <- 0

cat("Data summary:\\n")
cat(paste("  - Amino acids:", nrow(heatmap_matrix), "\\n"))
cat(paste("  - Genomes:", ncol(heatmap_matrix), "\\n"))
cat(paste("  - Percentage range:", round(min(heatmap_matrix), 2), "-", 
          round(max(heatmap_matrix), 2), "%\\n\\n"))

# ============================================================================
# FIGURE 1: AMINO ACID COMPOSITION HEATMAP
# ============================================================================

cat("Generating amino acid composition heatmap...\\n")

# Color palette for percentages
color_palette <- colorRampPalette(c("white", "yellow", "orange", "red", "darkred"))(100)

# Custom label function for italicizing species names in heatmap
make_italic <- function(x) {{
  as.expression(lapply(x, function(label) bquote(italic(.(label)))))
}}

# Generate PDF heatmap
pdf("Figures/AA_Composition_Heatmap.pdf", width={heatmap_width}, height={heatmap_height})
pheatmap(heatmap_matrix,
         main="Amino Acid Composition Across Genomes",
         color=color_palette,
         breaks=seq(0, max(heatmap_matrix, na.rm=TRUE), length.out=101),
         cluster_rows=TRUE,
         cluster_cols=TRUE,
         clustering_distance_rows="euclidean",
         clustering_distance_cols="euclidean",
         clustering_method="complete",
         labels_col=make_italic(colnames(heatmap_matrix)),
         fontsize=10,
         fontsize_row=9,
         fontsize_col=9,
         angle_col=90,
         cellwidth=NA,
         cellheight=NA,
         border_color="grey80",  # Add grid lines
         legend=TRUE,
         legend_labels="Percentage (%)",
         show_rownames=TRUE,
         show_colnames=TRUE,
         treeheight_row=20,
         treeheight_col=30)
dev.off()

# Generate high-quality PNG
png("Figures/AA_Composition_Heatmap.png", width={heatmap_width}, height={heatmap_height}, 
    units="in", res=600, type="cairo")
pheatmap(heatmap_matrix,
         main="Amino Acid Composition Across Genomes",
         color=color_palette,
         breaks=seq(0, max(heatmap_matrix, na.rm=TRUE), length.out=101),
         cluster_rows=TRUE,
         cluster_cols=TRUE,
         clustering_distance_rows="euclidean",
         clustering_distance_cols="euclidean",
         clustering_method="complete",
         labels_col=make_italic(colnames(heatmap_matrix)),
         fontsize=10,
         fontsize_row=9,
         fontsize_col=9,
         angle_col=90,
         cellwidth=NA,
         cellheight=NA,
         border_color="grey80",
         legend=TRUE,
         legend_labels="Percentage (%)",
         show_rownames=TRUE,
         show_colnames=TRUE,
         treeheight_row=20,
         treeheight_col=30)
dev.off()

cat("  ✓ Amino acid heatmap saved:\\n")
cat("    - Figures/AA_Composition_Heatmap.pdf\\n")
cat("    - Figures/AA_Composition_Heatmap.png\\n\\n")

# ============================================================================
# FIGURE 2: THREE SPECIES COMPARISON (BAR PLOT)
# ============================================================================

cat("Generating three species comparison bar plot...\\n")

# Select 3 representative species
n_genomes <- ncol(heatmap_matrix)

if (n_genomes >= 3) {{
  # Cluster genomes to find representatives
  genome_dist <- dist(t(heatmap_matrix), method="euclidean")
  genome_clust <- hclust(genome_dist, method="complete")
  genome_order <- genome_clust$order
  
  # Select 3 representative genomes: first, middle, last from cluster
  if (n_genomes == 3) {{
    selected_idx <- c(1, 2, 3)
  }} else if (n_genomes < 10) {{
    # For small datasets, pick evenly spaced
    selected_idx <- c(genome_order[1], 
                      genome_order[round(n_genomes/2)], 
                      genome_order[n_genomes])
  }} else {{
    # For large datasets, pick from different clusters
    selected_idx <- c(genome_order[1], 
                      genome_order[round(n_genomes/2)], 
                      genome_order[n_genomes])
  }}
  
  selected_species <- colnames(heatmap_matrix)[selected_idx]
  cat(paste("  → Selected 3 representative species:\\n"))
  for (sp in selected_species) {{
    cat(paste("     -", sp, "\\n"))
  }}
  
}} else {{
  # If less than 3 genomes, use all
  selected_species <- colnames(heatmap_matrix)
  cat(paste("  → Using all", n_genomes, "species\\n"))
}}

# Prepare data for bar plot
plot_data <- data[, c("AA_Code", "Amino_Acid", selected_species)]

# Convert to long format for ggplot
aa_data <- plot_data %>%
  pivot_longer(
    cols = -c(AA_Code, Amino_Acid),
    names_to = "Species",
    values_to = "Percentage"
  ) %>%
  mutate(
    Species = factor(Species, levels = selected_species),
    AA_Code = factor(AA_Code)
  )

# Rename columns to match your preferred style
colnames(aa_data)[colnames(aa_data) == "AA_Code"] <- "AminoAcid"

# Convert single-letter codes to 3-letter codes
aa_1to3 <- c(
  "A" = "Ala", "R" = "Arg", "N" = "Asn", "D" = "Asp", "C" = "Cys",
  "Q" = "Gln", "E" = "Glu", "G" = "Gly", "H" = "His", "I" = "Ile",
  "L" = "Leu", "K" = "Lys", "M" = "Met", "F" = "Phe", "P" = "Pro",
  "S" = "Ser", "T" = "Thr", "W" = "Trp", "Y" = "Tyr", "V" = "Val"
)
aa_data$AminoAcid <- aa_1to3[as.character(aa_data$AminoAcid)]

# Reorder AminoAcid by average percentage
avg_order <- aa_data %>%
  group_by(AminoAcid) %>%
  summarise(avg = mean(Percentage)) %>%
  arrange(desc(avg)) %>%
  pull(AminoAcid)

aa_data$AminoAcid <- factor(aa_data$AminoAcid, levels = avg_order)

# y-position for labels — slightly higher to avoid overlap
aa_data <- aa_data %>% mutate(label_y = Percentage + 1)  # Reduced from 0.8 for tighter layout

# Define colors for 3 species (matches your preferred style)
species_colors <- c("#87CEFA", "#FFA07A", "#90EE90")  # Light blue, light salmon, light green

# Create bar plot with your preferred styling
p <- ggplot(aa_data, aes(x = AminoAcid, y = Percentage, fill = Species)) +
  geom_bar(stat = "identity",
           position = position_dodge(width = 0.9),
           width = 0.8,
           color = "black", linewidth = 0.5) +  # Thin black border for clarity
  geom_text(aes(y = label_y, label = sprintf("%.2f", Percentage)),
            position = position_dodge(width = 0.9),
            size = 5,                            # Slightly larger for readability
            angle = 90,
            vjust = 0.5,
            hjust = 0.5) +                   # Bold percentages for emphasis
  scale_fill_manual(
    values = species_colors,
    name = ""
  ) +
  scale_y_continuous(
    expand = expansion(mult = c(0, 0.1)),
    limits = c(0, 12),
    breaks = seq(0, 12, 2)
  ) +
  labs(x = "Amino acid", y = "Amino acid frequency") +
  theme_classic(base_size = 16) +
  theme(
    axis.line = element_line(color = "black", size = 0.8),
    axis.ticks = element_line(color = "black", size = 0.8),
    
    # --- X-AXIS LABEL ---
    axis.title.x = element_text(size = 18, face = "bold", margin = margin(t = 20), hjust = 0.5),
    
    # --- Y-AXIS LABEL AND TICKS ---
    axis.title.y = element_text(size = 18, face = "bold", margin = margin(r = 15)),
    axis.text.y = element_text(size = 18, color = "black"),
    
    # --- X-AXIS TICKS (Amino Acid Abbreviations) ---
    axis.text.x = element_text(angle = 0, hjust = 0.5, vjust = 0.5, size = 22, color = "black"),
    
    # --- LEGEND ---
    legend.position = "top",
    legend.direction = "horizontal",
    legend.text = element_text(size = 18, face = "italic"),
    legend.title = element_blank(),
    legend.key.size = unit(1.2, "cm"),
    legend.background = element_rect(fill = "white", color = NA),
    legend.box.margin = margin(t = 10),
    
    # --- GENERAL ---
    panel.grid = element_blank(),
    plot.margin = margin(t = 1, r = 1.5, b = 2.5, l = 1.5, unit = "cm"),  # Extra bottom margin for legend
    panel.border = element_rect(color = "black", fill = NA, linewidth = 0.8)
  )

# Save plot
pdf("Figures/Amino_Acid_Comparison.pdf", width = 24, height = 7)
print(p)
dev.off()

# Also save as PNG with high DPI
ggsave("Figures/Amino_Acid_Comparison.png", plot = p,
       width = 24, height = 7, dpi = 600)

cat("  ✓ Three species comparison saved: Figures/Amino_Acid_Comparison.pdf and .png\\n\\n")

# ============================================================================
# SUMMARY
# ============================================================================

cat("========================================\\n")
cat("All visualizations generated successfully!\\n")
cat("========================================\\n\\n")
cat("Files created:\\n")
cat("  1. AA_Composition_Heatmap.pdf (all genomes, species italicized)\\n")
cat("  2. Amino_Acid_Comparison.pdf and .png (3 representative species, publication style)\\n\\n")
cat("Representative species for comparison:\\n")
for (sp in selected_species) {{
  cat(paste("  -", sp, "\\n"))
}}
cat("\\nNote: Species names formatted with spaces and italics\\n")
cat("      Heatmap shows all genomes\\n")
cat("      Bar plot shows 3 representatives for clarity\\n")
cat("      Publication-quality styling with black borders and custom colors\\n")
cat("========================================\\n\\n")
'''
    
    # Save R script
    with open(r_script_path, 'w') as f:
        f.write(r_script_content)
    
    print(f"  → R script created: {os.path.basename(r_script_path)}")
    
    # Execute R script
    print(f"  → Executing R script...")
    try:
        result = subprocess.run(['Rscript', r_script_path],
                              capture_output=True,
                              text=True,
                              timeout=300,
                              cwd=output_folder)
        
        if result.returncode == 0:
            print(f"  ✓ R visualization completed successfully")
            print(f"\n  Generated figures in {FIGURES_FOLDER}/:")
            print(f"    - AA_Composition_Heatmap.pdf")
            print(f"    - Amino_Acid_Comparison.pdf")
            print(f"    - Amino_Acid_Comparison.png")
            
            # Verify files were created
            import glob
            pdf_files = glob.glob(os.path.join(figures_folder, "*.pdf"))
            png_files = glob.glob(os.path.join(figures_folder, "*.png"))
            if pdf_files:
                for pdf in pdf_files:
                    size = os.path.getsize(pdf)
                    print(f"    ✓ {os.path.basename(pdf)} ({size:,} bytes)")
            if png_files:
                for png in png_files:
                    size = os.path.getsize(png)
                    print(f"    ✓ {os.path.basename(png)} ({size:,} bytes)")
            else:
                print(f"    ⚠ No PDF or PNG files found in {FIGURES_FOLDER}/")
                
        else:
            print(f"  ✗ R script execution failed")
            if result.stderr:
                print(f"  Error output (first 500 chars):")
                print(result.stderr[:500])
    
    except subprocess.TimeoutExpired:
        print(f"  ✗ R script execution timed out (>5 minutes)")
        print(f"  → This may happen with very large datasets")
    except Exception as e:
        print(f"  ✗ Error executing R script: {e}")


# ============================================================================
# MAIN PIPELINE
# ============================================================================

def main():
    """Main pipeline execution function."""
    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description='CGAS Module 9: Comprehensive Amino Acid Analysis with R Visualization',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  %(prog)s                          # Process current directory
  %(prog)s -i genbank_files/        # Process specific directory
  %(prog)s -i data/ -o results/     # Custom input and output
  %(prog)s --no-figures             # Skip R visualization

For more information, see the CGAS documentation.
        '''
    )
    
    parser.add_argument('-i', '--input',
                       type=str,
                       default='.',
                       help='Input directory containing GenBank files (default: current directory)')
    
    parser.add_argument('-o', '--output',
                       type=str,
                       default=None,
                       help='Output directory (default: Module9_Amino_Acid_Analysis in input directory)')
    
    parser.add_argument('--no-figures',
                       action='store_true',
                       help='Skip R figure generation')
    
    args = parser.parse_args()
    
    # Set up directories
    input_dir = os.path.abspath(args.input)
    
    if args.output:
        output_dir = os.path.abspath(args.output)
    else:
        output_dir = os.path.join(input_dir, OUTPUT_FOLDER)
    
    print(f"\n{'='*70}")
    print("CGAS MODULE 9: COMPREHENSIVE AMINO ACID ANALYSIS")
    print(f"{'='*70}")
    print(f"\nInput directory: {input_dir}")
    print(f"Output directory: {output_dir}")
    
    # Create output folder
    os.makedirs(output_dir, exist_ok=True)
    
    # Find GenBank files
    genbank_files = find_genbank_files(input_dir)
    
    if not genbank_files:
        print("\n❌ ERROR: No GenBank files found!")
        print("Please place GenBank files (.gb, .gbf, .gbk) in the input directory.")
        return
    
    print(f"\nFound {len(genbank_files)} GenBank file(s):")
    for gf in genbank_files[:10]:  # Show first 10
        print(f"  - {os.path.basename(gf)}")
    if len(genbank_files) > 10:
        print(f"  ... and {len(genbank_files) - 10} more")
    
    # Process each GenBank file
    print(f"\n{'='*70}")
    print("CALCULATING AMINO ACID COMPOSITION")
    print(f"{'='*70}")
    
    aa_files = []
    for gb_file in genbank_files:
        try:
            output_file, _, _ = process_single_genbank_file(gb_file, output_dir)
            if output_file:
                aa_files.append(output_file)
        except Exception as e:
            print(f"  ❌ ERROR processing {os.path.basename(gb_file)}: {e}")
            continue
    
    # Merge amino acid files
    if aa_files:
        merged_file = os.path.join(output_dir, "Complete_Amino_Acid_Analysis.xlsx")
        merge_aa_files(aa_files, merged_file)
        
        # Generate R visualizations
        if not args.no_figures:
            generate_r_visualizations(merged_file, output_dir, len(aa_files))
        else:
            print(f"\n  → Skipping R visualization (--no-figures flag)")
    
    # Summary
    print(f"\n{'='*70}")
    print("CGAS MODULE 9 COMPLETE")
    print(f"{'='*70}")
    print(f"✓ Successfully processed: {len(aa_files)} genome(s)")
    print(f"✓ Individual AA files: {len(aa_files)}")
    if aa_files:
        print(f"✓ Merged analysis (Excel): Complete_Amino_Acid_Analysis.xlsx")
        print(f"✓ Merged analysis (CSV): Complete_Amino_Acid_Analysis.csv")
        if not args.no_figures:
            print(f"✓ R visualizations: {FIGURES_FOLDER}/")
            print(f"    - AA_Composition_Heatmap.pdf (all genomes)")
            print(f"    - Amino_Acid_Comparison.pdf (3 representatives)")
    print(f"\nAll output files saved in: {output_dir}/")
    print(f"{'='*70}\n")


# ============================================================================
# SCRIPT ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠ Analysis interrupted by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\n\n❌ FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
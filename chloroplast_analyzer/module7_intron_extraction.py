#!/usr/bin/env python3
"""
Gene and tRNA Intron Extraction from GenBank Files

This script extracts intron positions and lengths from both gene (CDS/gene features)
and tRNA features in GenBank format files. Results are saved in a single Excel file
with separate sheets for gene introns and tRNA introns.

Author: Abdullah
Date: December 2025
Version: 1.0

Requirements:
    - Python 3.7+
    - biopython
    - pandas
    - openpyxl

Installation:
    pip install biopython pandas openpyxl

Usage:
    python module7_intron_extraction.py

Output:
    Module7_Intron_Analysis/
        intron_data_YYYYMMDD_HHMMSS.xlsx - Excel file with two sheets:
            - Sheet 1: Gene Introns (CDS/gene features)
            - Sheet 2: tRNA Introns (tRNA features)
"""

import os
import sys
import logging
from datetime import datetime
from typing import List, Tuple
import pandas as pd
from Bio import SeqIO
from Bio.SeqFeature import CompoundLocation

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Handle Windows console encoding
if sys.platform == 'win32':
    try:
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8')
    except (AttributeError, OSError):
        pass

# Global constants
MAX_INTRON_LENGTH = 15000
OUTPUT_FOLDER = "Module7_Intron_Analysis"
VALID_EXTENSIONS = ('.gb', '.gbff', '.genbank', '.gbk')


def to_roman(num):
    """Convert integer to Roman numeral (for intron numbering)."""
    val = [1000, 900, 500, 400, 100, 90, 50, 40, 10, 9, 5, 4, 1]
    syms = ['M', 'CM', 'D', 'CD', 'C', 'XC', 'L', 'XL', 'X', 'IX', 'V', 'IV', 'I']
    roman_num = ''
    for i in range(len(val)):
        count = int(num / val[i])
        if count:
            roman_num += syms[i] * count
            num -= val[i] * count
    return roman_num


def extract_gene_introns_from_genbank(gb_file: str) -> List[List]:
    """
    Extract gene/CDS intron information from a GenBank file.
    
    Args:
        gb_file (str): Path to GenBank format file
        
    Returns:
        List[List]: List of gene records with intron data
    """
    gene_introns = []
    filename = os.path.splitext(os.path.basename(gb_file))[0]  # Get filename without extension
    
    for record in SeqIO.parse(gb_file, "genbank"):
        species = record.annotations.get('organism', 'Unknown species')
        
        for feature in record.features:
            if feature.type in ['CDS', 'gene']:
                gene = feature.qualifiers.get('gene', ['Unknown'])[0]
                
                if isinstance(feature.location, CompoundLocation):
                    exons = sorted(feature.location.parts, key=lambda x: x.start)
                    
                    # Count valid introns first
                    valid_introns = []
                    for i in range(len(exons) - 1):
                        intron_start = int(exons[i].end) + 1
                        intron_end = int(exons[i + 1].start)
                        intron_length = intron_end - intron_start + 1
                        
                        if 0 < intron_length <= MAX_INTRON_LENGTH:
                            valid_introns.append((intron_start, intron_end, intron_length))
                    
                    # Now add intron data with Roman numerals
                    if valid_introns:
                        intron_info = []
                        num_introns = len(valid_introns)
                        
                        for idx, (intron_start, intron_end, intron_length) in enumerate(valid_introns, 1):
                            if num_introns == 1:
                                intron_label = "Intron"
                            else:
                                intron_label = f"Intron {to_roman(idx)}"
                            
                            intron_info.extend([
                                intron_label,
                                intron_start,
                                intron_end,
                                intron_length
                            ])
                        
                        gene_data = [filename, species, gene] + intron_info
                        gene_introns.append(gene_data)
    
    return gene_introns


def extract_tRNA_introns_from_genbank(gb_file: str) -> List[List]:
    """
    Extract tRNA intron information from a GenBank file.
    
    Args:
        gb_file (str): Path to GenBank format file
        
    Returns:
        List[List]: List of tRNA records with intron data
    """
    trna_introns = []
    filename = os.path.splitext(os.path.basename(gb_file))[0]  # Get filename without extension
    
    for record in SeqIO.parse(gb_file, "genbank"):
        species = record.annotations.get('organism', 'Unknown species')
        
        for feature in record.features:
            if feature.type == 'tRNA':
                gene = feature.qualifiers.get('gene', feature.qualifiers.get('product', ['Unknown']))[0]
                
                if isinstance(feature.location, CompoundLocation):
                    exons = sorted(feature.location.parts, key=lambda x: x.start)
                    
                    # Count valid introns first
                    valid_introns = []
                    for i in range(len(exons) - 1):
                        intron_start = exons[i].end + 1
                        intron_end = exons[i + 1].start - 1
                        intron_length = int(intron_end - intron_start + 1)
                        
                        if intron_length > 0 and intron_length <= MAX_INTRON_LENGTH:
                            valid_introns.append((int(intron_start), int(intron_end), intron_length))
                    
                    # Now add intron data with Roman numerals
                    if valid_introns:
                        intron_info = []
                        num_introns = len(valid_introns)
                        
                        for idx, (intron_start, intron_end, intron_length) in enumerate(valid_introns, 1):
                            if num_introns == 1:
                                intron_label = "Intron"
                            else:
                                intron_label = f"Intron {to_roman(idx)}"
                            
                            intron_info.extend([
                                intron_label,
                                intron_start,
                                intron_end,
                                intron_length
                            ])
                        
                        gene_data = [filename, species, gene] + intron_info
                        trna_introns.append(gene_data)
    
    return trna_introns


def add_abbreviation_footnotes(worksheet, start_row):
    """
    Add abbreviation footnotes to Excel worksheet for publication quality.
    
    Args:
        worksheet: openpyxl worksheet object
        start_row: Row number to start adding footnotes
    """
    from openpyxl.styles import Font, Alignment
    
    footnotes = [
        "Abbreviations:",
        "tRNA: Transfer RNA",
        "CDS: Coding Sequence",
        "bp: Base pairs"
    ]
    
    row = start_row
    for footnote in footnotes:
        cell = worksheet.cell(row, 1, footnote)
        if footnote == "Abbreviations:":
            cell.font = Font(bold=True, size=10)
        else:
            cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        row += 1
    
    return row


def create_combined_excel_output(gene_data: List[List], trna_data: List[List], 
                                 output_excel: str) -> None:
    """
    Create Excel output with separate sheets for gene and tRNA introns.
    Species are grouped with merged header rows (like module8).
    
    Args:
        gene_data (List[List]): Gene intron data rows [filename, species, gene, ...]
        trna_data (List[List]): tRNA intron data rows [filename, species, gene, ...]
        output_excel (str): Output Excel filename
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    species_font = Font(italic=True, bold=True, size=11)
    species_alignment = Alignment(horizontal='center', vertical='center')
    species_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    italic_font = Font(italic=True)
    regular_font = Font(italic=False)
    center_alignment = Alignment(horizontal='center')
    left_alignment = Alignment(horizontal='left')
    
    # ========================================================================
    # Create Gene Introns sheet
    # ========================================================================
    if gene_data:
        ws = wb.create_sheet('Gene Introns')
        
        # Determine max introns
        max_introns = max((len(row) - 3) // 4 for row in gene_data)
        
        # Create headers (without File column)
        headers = ["Gene"]
        for i in range(1, max_introns + 1):
            if max_introns == 1:
                headers += ["Intron", "Start", "End", "Length (bp)"]
            else:
                roman = to_roman(i)
                headers += [f"Intron {roman}", f"Start", f"End", f"Length (bp)"]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.alignment = header_alignment
        
        row = 2
        
        # Group data by species
        data_by_species = {}
        for record in gene_data:
            filename, species, gene = record[0], record[1], record[2]
            intron_data = record[3:]
            if species not in data_by_species:
                data_by_species[species] = []
            # Store without filename
            data_by_species[species].append([gene] + intron_data)
        
        # Write data grouped by species
        for species in sorted(data_by_species.keys()):
            species_records = data_by_species[species]
            
            # Add species row (merged across all columns)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            species_cell = ws.cell(row, 1, species)
            species_cell.font = species_font
            species_cell.alignment = species_alignment
            species_cell.fill = species_fill
            row += 1
            
            # Add data rows for this species
            for record in species_records:
                # Pad record to match column count
                while len(record) < len(headers):
                    record.append("")
                
                # Write gene (italic)
                ws.cell(row, 1, record[0]).font = italic_font
                ws.cell(row, 1, record[0]).alignment = left_alignment
                
                # Write intron data (centered)
                for col_idx, value in enumerate(record[1:], 2):
                    cell = ws.cell(row, col_idx, value)
                    cell.font = regular_font
                    cell.alignment = center_alignment
                
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 3, 50)
        
        # Add footnotes
        add_abbreviation_footnotes(ws, row + 2)
    
    # ========================================================================
    # Create tRNA Introns sheet
    # ========================================================================
    if trna_data:
        ws = wb.create_sheet('tRNA Introns')
        
        # Determine max introns
        max_introns = max((len(row) - 3) // 4 for row in trna_data)
        
        # Create headers (without File column)
        headers = ["tRNA Gene"]
        for i in range(1, max_introns + 1):
            if max_introns == 1:
                headers += ["Intron", "Start", "End", "Length (bp)"]
            else:
                roman = to_roman(i)
                headers += [f"Intron {roman}", f"Start", f"End", f"Length (bp)"]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.alignment = header_alignment
        
        row = 2
        
        # Group data by species
        data_by_species = {}
        for record in trna_data:
            filename, species, gene = record[0], record[1], record[2]
            intron_data = record[3:]
            if species not in data_by_species:
                data_by_species[species] = []
            # Store without filename
            data_by_species[species].append([gene] + intron_data)
        
        # Write data grouped by species
        for species in sorted(data_by_species.keys()):
            species_records = data_by_species[species]
            
            # Add species row (merged across all columns)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            species_cell = ws.cell(row, 1, species)
            species_cell.font = species_font
            species_cell.alignment = species_alignment
            species_cell.fill = species_fill
            row += 1
            
            # Add data rows for this species
            for record in species_records:
                # Pad record to match column count
                while len(record) < len(headers):
                    record.append("")
                
                # Write gene (italic)
                ws.cell(row, 1, record[0]).font = italic_font
                ws.cell(row, 1, record[0]).alignment = left_alignment
                
                # Write intron data (centered)
                for col_idx, value in enumerate(record[1:], 2):
                    cell = ws.cell(row, col_idx, value)
                    cell.font = regular_font
                    cell.alignment = center_alignment
                
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 3, 50)
        
        # Add footnotes
        add_abbreviation_footnotes(ws, row + 2)
    
    wb.save(output_excel)


def process_all_genbank_files() -> None:
    """
    Process all GenBank files in the current working directory.
    """
    logger.info("="*70)
    logger.info("MODULE 7: INTRON EXTRACTION ANALYSIS")
    logger.info("="*70)
    
    all_gene_introns = []
    all_trna_introns = []
    current_dir = os.getcwd()
    
    gb_files = [f for f in os.listdir(current_dir) if f.endswith(VALID_EXTENSIONS)]
    
    if not gb_files:
        logger.error(f"No GenBank files found in {current_dir}")
        raise FileNotFoundError("No GenBank files (.gb, .gbff, .genbank, .gbk) found")
    
    logger.info(f"Found {len(gb_files)} GenBank file(s)")
    logger.info("-"*70)
    
    # Process each file
    for idx, gb_file in enumerate(gb_files, 1):
        logger.info(f"[{idx}/{len(gb_files)}] Processing: {gb_file}")
        try:
            gene_rows = extract_gene_introns_from_genbank(gb_file)
            trna_rows = extract_tRNA_introns_from_genbank(gb_file)
            all_gene_introns.extend(gene_rows)
            all_trna_introns.extend(trna_rows)
            logger.info(f"  Genes with introns: {len(gene_rows)}, tRNAs with introns: {len(trna_rows)}")
        except Exception as e:
            logger.warning(f"  Skipping {gb_file}: {str(e)}")
            continue
    
    logger.info("-"*70)
    
    if not all_gene_introns and not all_trna_introns:
        raise ValueError("No introns found in any file")
    
    # Create output folder
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    
    # Generate timestamped filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(OUTPUT_FOLDER, f"intron_data_{timestamp}.xlsx")
    
    # Create Excel file
    logger.info(f"Creating: {output_file}")
    create_combined_excel_output(all_gene_introns, all_trna_introns, output_file)
    
    logger.info("="*70)
    logger.info(f"✓ SUCCESS: {output_file} created")
    logger.info(f"  Gene introns: {len(all_gene_introns)}")
    logger.info(f"  tRNA introns: {len(all_trna_introns)}")
    logger.info("="*70)


def main():
    """Main execution function."""
    try:
        process_all_genbank_files()
    except Exception as e:
        logger.error(f"Error: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
